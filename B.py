# ----------------------------------------------------- Partie 1 : Configuration et Imports ------------------------------------------------------
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import os
import unicodedata
import subprocess
import re
import numpy as np
import keyboard
try:
    from fuzzywuzzy import fuzz, process
except ImportError:
    subprocess.run(["pip", "install", "fuzzywuzzy"])
    from fuzzywuzzy import fuzz, process
import platform
import threading


# ----------------------------------------------------- Partie 2 : Fonctions utilitaires ----------------------------------------------------------------
# ----------------------------Cette section définit des fonctions qui seront utilisées par les fonctions principales-------------------------------------


def remove_accents(input_str):
    nfkd_form = unicodedata.normalize('NFKD', str(input_str))
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])

def close_excel_file(file_path):
    """Close the Excel file if it is open."""
    import psutil
    import os
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if proc.name().lower() in ['excel.exe', 'open', 'xdg-open']:
                for open_file in proc.open_files():
                    if os.path.normpath(open_file.path) == os.path.normpath(file_path):
                        proc.terminate()
                        print(f"Fichier Excel '{file_path}' fermé via terminaison du processus {proc.name()}.")
                        return
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    print(f"Aucun processus Excel ouvert trouvé pour '{file_path}'.")


# ----------------------------------------------------- Partie 3 : Normalisation du fichier Excel ----------------------------------------------------------------
# Cette section définit la fonction principale pour normaliser les noms des colonnes et des lignes dans le fichier Excel.
# -------------------------------------------Objectif : Standardiser le fichier Excel généré par le code A --------------------------------------------------------

def normalize_excel(input_file=None):
    EXPECTED_COLUMNS = ["CATEGORIES", "GROUPE", "A.TRAVAIL", "INCENDIE", "RISQUES DIVERS", "TRANSPORT", "AVIATION", "AUTOMOBILE", "ACCEPTATION", "TOTAL"]
    EXPECTED_ROWS = [
        "PRIMES ACQUISES", "PRIMES EMISES", "VARIATION DES PRIMES NON ACQUISES", 
        "CHARGES DE PRESTATION", "PRESTATIONS ET FRAIS PAYES", 
        "CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE", "SOLDE DE SOUSCRIPTION", 
        "FRAIS D'ACQUISITION", "AUTRES CHARGES DE GESTION NETTES", 
        "CHARGES D'ACQUISITION ET DE GESTION NETTES", "PRODUITS NETS DE PLACEMENTS", 
        "PARTICIPATION AUX RESULTATS", "SOLDE FINANCIER", 
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRIMES ACQUISES", 
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRESTATIONS PAYES", 
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LES CHARGES DE PROVI. POUR PRESTATIONS", 
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LA PARTICIPATION AUX RESULTATS", 
        "COMMISSIONS REÇUES DES REASSUREURS /RETROCESS", "SOLDE DE REASSURANCE / RETROCESSION", 
        "RESULTAT TECHNIQUE", "INFORMATIONS COMPLEMENTAIRES", 
        "PROVISIONS POUR PRIMES NON ACQUISES CLOTURE", 
        "PROVISIONS POUR PRIMES NON ACQUISES REOUVERTURE", 
        "PROVISIONS POUR SINISTRES A PAYER CLOTURE", 
        "PROVISIONS POUR SINISTRES A PAYER REOUVERTURE"
    ]

    script_dir = os.path.dirname(os.path.abspath(__file__))
    if input_file is None:
        input_file = input("Saisissez le nom du fichier Excel (.xlsx) : ").strip()
    input_path = os.path.join(script_dir, input_file) if not os.path.isabs(input_file) else input_file
    
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Le fichier '{input_file}' n'a pas été trouvé dans le dossier du projet.")
    
    if not input_file.lower().endswith('.xlsx'):
        raise ValueError(f"Le fichier '{input_file}' n'est pas un fichier Excel valide (.xlsx).")
    
    base, ext = os.path.splitext(os.path.basename(input_file))
    output_file = os.path.join(script_dir, f"output_{base}{ext}")

    df = pd.read_excel(input_path, thousands=" ")
    print(f"Colonnes détectées : {df.columns.tolist()}")

    # Normalisation des noms de colonnes
    print("Normalisation des noms de colonnes...")
    normalized_columns = []
    total_col = None
    expected_columns_no_accents = [remove_accents(col) for col in EXPECTED_COLUMNS]
    original_columns = df.columns.tolist()
    for col_idx, col in enumerate(original_columns):
        col_str = str(col).strip()
        if pd.isna(col) or col_str == '' or col_str.startswith('Unnamed:'):
            if col_idx == len(df.columns) - 1:
                normalized_columns.append("TOTAL")
                total_col = "TOTAL"
                print(f"Colonne sans nom ou 'Unnamed' (dernière, index {col_idx}) normalisée en 'TOTAL'.")
            else:
                raise ValueError(f"Erreur : Colonne sans nom ou 'Unnamed' détectée à l'index {col_idx + 1} (non dernière colonne).")
        else:
            col_no_accents = remove_accents(col_str)
            match, score = process.extractOne(col_no_accents, expected_columns_no_accents, scorer=fuzz.token_sort_ratio)
            if score >= 80:
                match_idx = expected_columns_no_accents.index(match)
                normalized_columns.append(EXPECTED_COLUMNS[match_idx])
                if EXPECTED_COLUMNS[match_idx] == "TOTAL":
                    total_col = EXPECTED_COLUMNS[match_idx]
            else:
                normalized_columns.append(col_str)
                print(f"Colonne '{col_str}' non normalisée (score={score})")

    if total_col and total_col in normalized_columns:
        normalized_columns.remove("TOTAL")
        normalized_columns.append("TOTAL")

    # Réaligner les données avec les nouvelles colonnes
    column_mapping = dict(zip(original_columns, normalized_columns))
    df = df.rename(columns=column_mapping)
    print(f"Noms de colonnes normalisés : {df.columns.tolist()}")

    # Normalisation des noms de lignes
    print("Normalisation des noms de lignes...")
    expected_rows_no_accents = [remove_accents(row) for row in EXPECTED_ROWS]
    normalized_rows = []
    prefix = None
    skip_prefix = False
    indices_to_drop = []
    seen_rows = set()
    for idx, row_name in enumerate(df['CATEGORIES']):
        row_name_str = str(row_name).strip()
        if not row_name_str:
            indices_to_drop.append(idx)
            print(f"Ligne vide à l'index {idx}, supprimée.")
            continue
        numeric_values = df.iloc[idx, 1:].infer_objects(copy=False)
        is_numeric_empty = numeric_values.isna().all() or (numeric_values.eq(0).sum() == 1 and numeric_values.isna().sum() == len(numeric_values) - 1)

        if is_numeric_empty:
            if row_name_str.lower().endswith(('dans', 'pour', 'de', 'au')):
                prefix = row_name_str
                skip_prefix = True
                indices_to_drop.append(idx)
                print(f"Ligne '{row_name_str}' avec valeurs vides ou un seul 0, utilisée comme préfixe.")
            else:
                indices_to_drop.append(idx)
                print(f"Ligne '{row_name_str}' avec valeurs vides ou un seul 0, supprimée (ne se termine pas par 'dans', 'pour', 'de', ou 'au').")
            continue
        else:
            row_name_no_accents = remove_accents(row_name_str)
            if skip_prefix and row_name_str.lower().startswith(('le ', 'la ', 'les ')):
                combined_name = f"{prefix} {row_name_str}"
                match, score = process.extractOne(remove_accents(combined_name), expected_rows_no_accents, scorer=fuzz.token_sort_ratio)
                if score >= 65:
                    match_idx = expected_rows_no_accents.index(match)
                    normalized_row = EXPECTED_ROWS[match_idx]
                else:
                    normalized_row = combined_name
                    print(f"Ligne '{combined_name}' non normalisée (score={score})")
            else:
                match, score = process.extractOne(row_name_no_accents, expected_rows_no_accents, scorer=fuzz.token_sort_ratio)
                if score >= 65 and not row_name_str.lower().startswith('part réassureurs /rétrocessionnaires dans commissions'):
                    match_idx = expected_rows_no_accents.index(match)
                    normalized_row = EXPECTED_ROWS[match_idx]
                    print(f"Ligne '{row_name_str}' normalisée en '{normalized_row}' (score={score})")
                else:
                    if row_name_no_accents.lower().startswith('commissions reçues'):
                        normalized_row = "COMMISSIONS REÇUES DES REASSUREURS /RETROCESS"
                        print(f"Ligne '{row_name_str}' normalisée en 'COMMISSIONS REÇUES DES REASSUREURS /RETROCESS' (règle spéciale)")
                    else:
                        normalized_row = row_name_str
                        print(f"Ligne '{row_name_str}' non normalisée (score={score})")
                skip_prefix = False
                prefix = None

            numeric_tuple = tuple(numeric_values.fillna(0).infer_objects(copy=False).astype(str))
            row_identifier = (normalized_row, numeric_tuple)
            if row_identifier in seen_rows:
                print(f"Doublon détecté : '{normalized_row}' à l'index {idx} avec données {numeric_tuple}, supprimé.")
                indices_to_drop.append(idx)
                continue
            seen_rows.add(row_identifier)
            normalized_rows.append(normalized_row)

    df = df.drop(indices_to_drop).reset_index(drop=True)
    if len(normalized_rows) != len(df):
        raise ValueError(f"Incohérence : {len(normalized_rows)} lignes normalisées vs {len(df)} lignes dans le DataFrame")
    df['CATEGORIES'] = normalized_rows
    print(f"Noms de lignes normalisés : {df['CATEGORIES'].tolist()}")

    # Mise à jour du fichier Excel
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    ws.delete_rows(2, ws.max_row)
    output_cols = normalized_columns
    col_map = {col: idx + 1 for idx, col in enumerate(output_cols)}

    for col_idx, col_name in enumerate(output_cols, 1):
        ws.cell(row=1, column=col_idx).value = col_name

    for row_idx, (row, norm_row_name) in enumerate(zip(df.itertuples(index=False), normalized_rows), 2):
        for col_idx, col_name in enumerate(output_cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == 'CATEGORIES':
                cell.value = norm_row_name
                print(f"Écriture ligne {row_idx-1} CATEGORIES : {norm_row_name}")
            else:
                value = row[df.columns.get_loc(col_name)] if col_name in df.columns else None
                cell.value = value if not pd.isna(value) else None
            if pd.notnull(cell.value):
                value_str = str(int(cell.value)) if isinstance(cell.value, (int, float)) else str(cell.value)
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="0077CC", end_color="0077CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    for col_idx in range(1, len(output_cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    try:
        wb.save(output_file)
        print(f"Fichier sauvegardé : {output_file}")
    except Exception as e:
        raise ValueError(f"Erreur lors de l'enregistrement du fichier Excel '{output_file}' : {str(e)}")

    return output_file




# ----------------------------------------------------- Partie 4 : Validation des données financières ----------------------------------------------------------------
# Cette section définit une fonction qui valide les données financières dans le fichier Excel en effectuant des calculs de contrôle (C1 à C9).
# ---------------------------------Objectif : Vérifier que les données sont cohérentes et signaler les erreurs visuellement--------------------------------------------

def validate_excel(input_file):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_filename = os.path.basename(input_file)
    input_path = os.path.join(script_dir, input_filename)
    

    # Vérifie si le fichier existe-----------
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Le fichier '{input_filename}' n'a pas été trouvé dans le dossier du projet.")
    

    # Crée le nom du fichier de sortie--------
    base, ext = os.path.splitext(input_filename)
    output_file = os.path.join(script_dir, f"output_{base}{ext}")


# Charge le fichier Excel----------------------
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    has_symbol_or_letter = False
    corrected_cells = []
# Nettoie les cellules contenant des lettres ou des symboles dans les colonnes numériques------
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(2, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if pd.notnull(cell_value):
                cell_str = str(cell_value)
                if re.search(r'[a-zA-Z!@#$%^&*(),.?":{}|<>]', cell_str):# Vérifie si la cellule contient des lettres ou symboles
                    cleaned_value = re.sub(r'[a-zA-Z!@#$%^&*(),.?":{}|<>]', '', cell_str)# Supprime les caractères non numériques
                    try:
                        ws.cell(row=row_idx, column=col_idx).value = float(cleaned_value) if cleaned_value else 0
                        ws.cell(row=row_idx, column=col_idx).number_format = '0'
                    except ValueError:
                        ws.cell(row=row_idx, column=col_idx).value = 0
                    print(f"Cellule corrigée à la ligne {row_idx}, colonne {col_idx} : '{cell_str}' -> '{cleaned_value}'")
                    has_symbol_or_letter = True
                    corrected_cells.append((row_idx, col_idx))


# Sauvegarde le fichier si des corrections ont été faites-------------------------------------------
    if has_symbol_or_letter:
        try:
            wb.save(output_file)
        except Exception as e:
            raise ValueError(f"Erreur lors de l'enregistrement du fichier Excel '{output_file}' : {str(e)}")
        print(f"Cellules corrigées : {corrected_cells}")
        print(f"Fichier sauvegardé : {output_file} avec les cellules corrigées.")
        df = pd.read_excel(output_file, thousands=" ")
    else:
        df = pd.read_excel(input_path, thousands=" ")
        print(f"Aucune lettre ou symbole détecté, utilisation du fichier d'entrée : {input_path}")

    print(f"Colonnes détectées : {df.columns.tolist()}")
    print(f"Premières lignes :\n{df.head()}")


# Identifie la colonne "TOTAL" (la considère comme dernière colonne)
    total_col = None
    numeric_cols = [col for col in df.columns[1:] if pd.api.types.is_numeric_dtype(df[col])]
    if numeric_cols:
        total_col = numeric_cols[-1]
        print(f"Détection automatique de '{total_col}' comme colonne TOTAL.")
    if not total_col:
        raise ValueError("Aucune colonne numérique trouvée pour TOTAL.")

    numeric_cols = [col for col in df.columns[1:] if col != total_col and pd.api.types.is_numeric_dtype(df[col])]


# Liste des noms de lignes à vérifier pour les calculs.
    target_rows = [
        "PRIMES ACQUISES", "PRIMES EMISES", "VARIATION DES PRIMES NON ACQUISES",
        "CHARGES DE PRESTATION", "PRESTATIONS ET FRAIS PAYES", "CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE",
        "SOLDE DE SOUSCRIPTION", "CHARGES D'ACQUISITION ET DE GESTION NETTES", "FRAIS D'ACQUISITION", 
        "AUTRES CHARGES DE GESTION NETTES", "PRODUITS NETS DE PLACEMENTS", "PARTICIPATION AUX RESULTATS", 
        "SOLDE FINANCIER", "PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRIMES ACQUISES",
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRESTATIONS PAYES",
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LES CHARGES DE PROVI. POUR PRESTATIONS",
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LA PARTICIPATION AUX RESULTATS",
        "COMMISSIONS REÇUES DES REASSUREURS /RETROCESS",
        "SOLDE DE REASSURANCE / RETROCESSION", "RESULTAT TECHNIQUE",
        "PROVISIONS POUR PRIMES NON ACQUISES CLOTURE", "PROVISIONS POUR PRIMES NON ACQUISES REOUVERTURE"
    ]



    # -----------------------------C1 :TOTAL = Σ(valeurs des colonnes numériques)---------------------------------
    c1_values = [float('nan')] * len(df)
    for idx, row in df.iterrows():
        if row.iloc[0] in target_rows:
            try:
                numeric_sum = sum(row[col] for col in numeric_cols if pd.notnull(row[col]))
                total_value = row[total_col] if pd.notnull(row[total_col]) else 0
                C1 = float(total_value) - numeric_sum
                if abs(C1) > 5:
                    print(f"Avertissement : Ligne {idx + 2} - Somme ({numeric_sum}) ≠ Total ({total_value}), C1 = {C1}")
                c1_values[idx] = C1
            except (TypeError, ValueError):
                c1_values[idx] = float('nan')


# Ajoute la colonne C1 au tableau------------------------
    output_cols = list(df.columns) + ['C1'] if total_col == df.columns[-1] else list(df.columns[:-1]) + [total_col, 'C1']
    df_output = df.copy()
    df_output['C1'] = c1_values

# Charge à nouveau le fichier Excel pour appliquer les modifications-------------
    wb = openpyxl.load_workbook(output_file) if has_symbol_or_letter else openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # Supprime la colonne C1 existante si elle existe.
    max_col = ws.max_column
    if ws.cell(row=1, column=max_col).value and ws.cell(row=1, column=max_col).value.lower() == 'c1':
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=max_col, max_col=max_col):
            for cell in row:
                cell.value = None


# Écrit les nouvelles colonnes, y compris C1.
    col_map = {col: idx + 1 for idx, col in enumerate(output_cols)}
    for col_idx, col_name in enumerate(output_cols, 1):
        ws.cell(row=1, column=col_idx).value = col_name
        for row_idx, value in enumerate(df_output[col_name], 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == 'C1' and pd.notnull(value):
                cell.value = int(value) if not pd.isna(value) else None
                cell.number_format = '0'
            else:
                cell.value = value if not pd.isna(value) else None
            if pd.notnull(value):
                value_str = str(int(value)) if isinstance(value, (int, float)) else str(value)
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="0077CC", end_color="0077CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    for col_idx in range(1, len(output_cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

# Définit les couleurs pour signaler les erreurs ou corrections.
    c1_col_idx = col_map['C1']
    ws.column_dimensions[openpyxl.utils.get_column_letter(c1_col_idx)].width = 15

    intersection_red_fill = PatternFill(start_color="FF4040", end_color="FF4040", fill_type="solid")
    corrected_green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
    medium_orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    bold_white_font = Font(color="FFFFFF", bold=True)
    bold_black_font = Font(color="000000", bold=True)


  # -----------------------------C2 :Primes Acquises = Primes Émises + Variation des Primes non acquises---------------------------------
  

    primes_rows = ["PRIMES ACQUISES", "PRIMES EMISES", "VARIATION DES PRIMES NON ACQUISES"]
    primes_acquises_idx = df.index[df['CATEGORIES'].str.contains("PRIMES ACQUISES", case=False, na=False)].tolist()
    primes_emises_idx = df.index[df['CATEGORIES'].str.contains("PRIMES EMISES", case=False, na=False)].tolist()
    var_acquises_idx = df.index[df['CATEGORIES'].str.contains("VARIATION DES PRIMES NON ACQUISES", case=False, na=False)].tolist()
    
    c2_values = [None]
    invalid_c2_cols = []
    if primes_acquises_idx and primes_emises_idx and var_acquises_idx:
        print(f"Indices détectés - Primes acquises : {primes_acquises_idx}, Primes émises : {primes_emises_idx}, Variation : {var_acquises_idx}")
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            primes_acquises = df.at[primes_acquises_idx[0], col_name] if primes_acquises_idx else float('nan')
            primes_emises = df.at[primes_emises_idx[0], col_name] if primes_emises_idx else float('nan')
            var_acquises = df.at[var_acquises_idx[0], col_name] if var_acquises_idx else float('nan')
            
            print(f"Debug - Colonne {col_name} : Primes acquises={primes_acquises}, Primes émises={primes_emises}, Variation={var_acquises}")
            if pd.notnull(primes_acquises) and pd.notnull(primes_emises) and pd.notnull(var_acquises):
                C2 = primes_acquises - (primes_emises + var_acquises)
                print(f"C2 calculé pour {col_name} : {C2}")
                c2_values.append(C2)
                if abs(C2) > 5:
                    invalid_c2_cols.append(col_idx)
            else:
                c2_values.append(float('nan'))
                print(f"Données manquantes dans la colonne {col_name} pour le calcul de C2")
        print(f"Colonnes C2 invalides (indices Excel) : {invalid_c2_cols}")
    else:
        print(f"Avertissement : Lignes manquantes pour C2 : {primes_rows}")



# -----------------------------C3 :Charges de prestation= Prestations et Frais payés + Charges des provisions pour prestations diverse---------------------------------
  
    charges_rows = ["CHARGES DE PRESTATION", "PRESTATIONS ET FRAIS PAYES", "CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE"]
    charges_prestation_idx = df.index[df['CATEGORIES'].str.contains("CHARGES DE PRESTATION", case=False, na=False)].tolist()
    prestations_payes_idx = df.index[df['CATEGORIES'].str.contains("PRESTATIONS ET FRAIS PAYES", case=False, na=False)].tolist()
    charges_provisions_idx = df.index[df['CATEGORIES'].str.contains("CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE", case=False, na=False)].tolist()
    
    c3_values = [None]
    invalid_c3_cols = []
    if charges_prestation_idx and prestations_payes_idx and charges_provisions_idx:
        print(f"Indices détectés - Charges de prestation : {charges_prestation_idx}, Prestations payés : {prestations_payes_idx}, Charges provisions : {charges_provisions_idx}")
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            charges_prestation = df.at[charges_prestation_idx[0], col_name] if charges_prestation_idx else float('nan')
            prestations_payes = df.at[prestations_payes_idx[0], col_name] if prestations_payes_idx else float('nan')
            charges_provisions = df.at[charges_provisions_idx[0], col_name] if charges_provisions_idx else float('nan')
            
            print(f"Debug - Colonne {col_name} : Charges de prestation={charges_prestation}, Prestations payés={prestations_payes}, Charges provisions={charges_provisions}")
            if pd.notnull(charges_prestation) and pd.notnull(prestations_payes) and pd.notnull(charges_provisions):
                C3 = charges_prestation - (prestations_payes + charges_provisions)
                print(f"C3 calculé pour {col_name} : {C3}")
                c3_values.append(C3)
                if abs(C3) > 5:
                    invalid_c3_cols.append(col_idx)
            else:
                c3_values.append(float('nan'))
                print(f"Données manquantes dans la colonne {col_name} pour le calcul de C3")
        print(f"Colonnes C3 invalides (indices Excel) : {invalid_c3_cols}")
    else:
        print(f"Avertissement : Lignes manquantes pour C3 : {charges_rows}")


# -----------------------------C4 :Solde de souscription = Primes Acquises + Charges de prestation---------------------------------

    solde_rows = ["SOLDE DE SOUSCRIPTION"]
    solde_souscription_idx = df.index[df['CATEGORIES'].str.contains("SOLDE DE SOUSCRIPTION", case=False, na=False)].tolist()
    primes_acquises_idx = df.index[df['CATEGORIES'].str.contains("PRIMES ACQUISES", case=False, na=False)].tolist()
    charges_prestation_idx = df.index[df['CATEGORIES'].str.contains("CHARGES DE PRESTATION", case=False, na=False)].tolist()
    
    c4_values = [None]
    invalid_c4_cols = []
    if solde_souscription_idx and primes_acquises_idx and charges_prestation_idx:
        print(f"Indices détectés - Solde de souscription : {solde_souscription_idx}, Primes acquises : {primes_acquises_idx}, Charges de prestation : {charges_prestation_idx}")
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            solde_souscription = df.at[solde_souscription_idx[0], col_name] if solde_souscription_idx else float('nan')
            primes_acquises = df.at[primes_acquises_idx[0], col_name] if primes_acquises_idx else float('nan')
            charges_prestation = df.at[charges_prestation_idx[0], col_name] if charges_prestation_idx else float('nan')
            
            print(f"Debug - Colonne {col_name} : Solde de souscription={solde_souscription}, Primes acquises={primes_acquises}, Charges de prestation={charges_prestation}")
            if pd.notnull(solde_souscription) and pd.notnull(primes_acquises) and pd.notnull(charges_prestation):
                C4 = solde_souscription - (primes_acquises - charges_prestation)
                print(f"C4 calculé pour {col_name} : {C4}")
                c4_values.append(C4)
                if abs(C4) > 5:
                    invalid_c4_cols.append(col_idx)
            else:
                c4_values.append(float('nan'))
                print(f"Données manquantes dans la colonne {col_name} pour le calcul de C4")
        print(f"Colonnes C4 invalides (indices Excel) : {invalid_c4_cols}")
    else:
        print(f"Avertissement : Lignes manquantes pour C4 : {solde_rows + ['PRIMES ACQUISES', 'CHARGES DE PRESTATION']}")



# -----------------------------C5 : Charges d'acquisition et de gestion nettes = Frais d'acquisition + Autres charges de gestion nettes---------------------------------

    acquisition_rows = ["CHARGES D'ACQUISITION ET DE GESTION NETTES", "FRAIS D'ACQUISITION", "AUTRES CHARGES DE GESTION NETTES"]
    charges_acquisition_idx = df.index[df['CATEGORIES'].str.contains("CHARGES D'ACQUISITION ET DE GESTION NETTES", case=False, na=False)].tolist()
    frais_acquisition_idx = df.index[df['CATEGORIES'].str.contains("FRAIS D'ACQUISITION", case=False, na=False)].tolist()
    autres_charges_idx = df.index[df['CATEGORIES'].str.contains("AUTRES CHARGES DE GESTION NETTES", case=False, na=False)].tolist()
    
    c5_values = [None]
    invalid_c5_cols = []
    if charges_acquisition_idx and frais_acquisition_idx and autres_charges_idx:
        print(f"Indices détectés - Charges d'acquisition : {charges_acquisition_idx}, Frais d'acquisition : {frais_acquisition_idx}, Autres charges : {autres_charges_idx}")
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            charges_acquisition = df.at[charges_acquisition_idx[0], col_name] if charges_acquisition_idx else float('nan')
            frais_acquisition = df.at[frais_acquisition_idx[0], col_name] if frais_acquisition_idx else float('nan')
            autres_charges = df.at[autres_charges_idx[0], col_name] if autres_charges_idx else float('nan')
            
            print(f"Debug - Colonne {col_name} : Charges d'acquisition={charges_acquisition}, Frais d'acquisition={frais_acquisition}, Autres charges={autres_charges}")
            if pd.notnull(charges_acquisition) and pd.notnull(frais_acquisition) and pd.notnull(autres_charges):
                C5 = charges_acquisition - (frais_acquisition + autres_charges)
                print(f"C5 calculé pour {col_name} : {C5}")
                c5_values.append(C5)
                if abs(C5) > 5:
                    invalid_c5_cols.append(col_idx)
            else:
                c5_values.append(float('nan'))
                print(f"Données manquantes dans la colonne {col_name} pour le calcul de C5")
        print(f"Colonnes C5 invalides (indices Excel) : {invalid_c5_cols}")
    else:
        print(f"Avertissement : Lignes manquantes pour C5 : {acquisition_rows}")


# -----------------------------C6 : Solde Financier = Produits nets de placements + Participation aux résultats ---------------------------------


    financier_rows = ["SOLDE FINANCIER", "PRODUITS NETS DE PLACEMENTS", "PARTICIPATION AUX RESULTATS"]
    solde_financier_idx = df.index[df['CATEGORIES'].str.contains("SOLDE FINANCIER", case=False, na=False)].tolist()
    produits_nets_idx = df.index[df['CATEGORIES'].str.contains("PRODUITS NETS DE PLACEMENTS", case=False, na=False)].tolist()
    participation_idx = df.index[df['CATEGORIES'].str.contains("PARTICIPATION AUX RESULTATS", case=False, na=False)].tolist()
    
    c6_values = [None]
    invalid_c6_cols = []
    if solde_financier_idx and produits_nets_idx and participation_idx:
        participation_idx = [min(participation_idx)]
        print(f"Indices détectés - Solde Financier : {solde_financier_idx}, Produits nets : {produits_nets_idx}, Participation : {participation_idx}")
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            solde_financier = df.at[solde_financier_idx[0], col_name] if solde_financier_idx else float('nan')
            produits_nets = df.at[produits_nets_idx[0], col_name] if produits_nets_idx else float('nan')
            participation = df.at[participation_idx[0], col_name] if participation_idx else float('nan')
            
            print(f"Debug - Colonne {col_name} : Solde Financier={solde_financier}, Produits nets={produits_nets}, Participation={participation}")
            if pd.notnull(solde_financier) and pd.notnull(produits_nets) and pd.notnull(participation):
                C6 = solde_financier - (produits_nets - participation)
                print(f"C6 calculé pour {col_name} : {C6}")
                c6_values.append(C6)
                if abs(C6) > 5:
                    invalid_c6_cols.append(col_idx)
            else:
                c6_values.append(float('nan'))
                print(f"Données manquantes dans la colonne {col_name} pour le calcul de C6")
        print(f"Colonnes C6 invalides (indices Excel) : {invalid_c6_cols}")
    else:
        print(f"Avertissement : Lignes manquantes pour C6 : {financier_rows}")

# -----------------------------C7 : Solde de réassurance / rétrocession=Part réassureurs /rétrocessionnaires dans les primes acquises + Part réassureurs /rétrocessionnaires dans les prestations payés+ Part réassureurs /rétrocessionnaires dans les charges de provi.  pour prestations+ Part réassureurs /rétrocessionnaires dans Commissions reçues des réassureurs /rétrocess ---------------------------------

    reassurance_rows = [
        "SOLDE DE REASSURANCE / RETROCESSION",
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRIMES ACQUISES",
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRESTATIONS PAYES",
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LES CHARGES DE PROVI. POUR PRESTATIONS",
        "COMMISSIONS REÇUES DES REASSUREURS /RETROCESS",
        "PART REASSUREURS /RETROCESSIONNAIRES DANS LA PARTICIPATION AUX RESULTATS"
    ]
    solde_reassurance_idx = df.index[df['CATEGORIES'].str.contains("SOLDE DE REASSURANCE / RETROCESSION", case=False, na=False, regex=True)].tolist()
    part_primes_idx = df.index[df['CATEGORIES'].str.contains("PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRIMES ACQUISES", case=False, na=False, regex=True)].tolist()
    part_prestations_idx = df.index[df['CATEGORIES'].str.contains("PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRESTATIONS PAYES", case=False, na=False, regex=True)].tolist()
    part_provisions_idx = df.index[df['CATEGORIES'].str.contains("PART REASSUREURS /RETROCESSIONNAIRES DANS LES CHARGES DE PROVI. POUR PRESTATIONS", case=False, na=False, regex=True)].tolist()
    part_commissions_idx = df.index[df['CATEGORIES'].str.contains("COMMISSIONS REÇUES DES REASSUREURS /RETROCESS", case=False, na=False, regex=True)].tolist()
    part_participation_idx = df.index[df['CATEGORIES'].str.contains("PART REASSUREURS /RETROCESSIONNAIRES DANS LA PARTICIPATION AUX RESULTATS", case=False, na=False, regex=True)].tolist()
    
    c7_values = [None]
    invalid_c7_cols = []
    missing_rows = []
    if not solde_reassurance_idx:
        missing_rows.append("SOLDE DE REASSURANCE / RETROCESSION")
    if not part_primes_idx:
        missing_rows.append("PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRIMES ACQUISES")
    if not part_prestations_idx:
        missing_rows.append("PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRESTATIONS PAYES")
    if not part_provisions_idx:
        missing_rows.append("PART REASSUREURS /RETROCESSIONNAIRES DANS LES CHARGES DE PROVI. POUR PRESTATIONS")
    if not part_commissions_idx:
        missing_rows.append("COMMISSIONS REÇUES DES REASSUREURS /RETROCESS")
    if not part_participation_idx:
        missing_rows.append("PART REASSUREURS /RETROCESSIONNAIRES DANS LA PARTICIPATION AUX RESULTATS")
    
    if missing_rows:
        print(f"Avertissement : Lignes manquantes pour C7 : {missing_rows}")
    else:
        print(f"Indices détectés - Solde de réassurance : {solde_reassurance_idx}, Part primes : {part_primes_idx}, Part prestations : {part_prestations_idx}, Part provisions : {part_provisions_idx}, Part participation : {part_participation_idx}, Part commissions : {part_commissions_idx}")
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            solde_reassurance = df.at[solde_reassurance_idx[0], col_name] if solde_reassurance_idx else float('nan')
            part_primes = df.at[part_primes_idx[0], col_name] if part_primes_idx else float('nan')
            part_prestations = df.at[part_prestations_idx[0], col_name] if part_prestations_idx else float('nan')
            part_provisions = df.at[part_provisions_idx[0], col_name] if part_provisions_idx else float('nan')
            part_commissions = df.at[part_commissions_idx[0], col_name] if part_commissions_idx else float('nan')
            part_participation = df.at[part_participation_idx[0], col_name] if part_participation_idx else float('nan')
            
            print(f"Debug - Colonne {col_name} : Solde de réassurance={solde_reassurance}, Part primes={part_primes}, Part prestations={part_prestations}, Part provisions={part_provisions}, Part commissions={part_commissions}, Part participation={part_participation}")
            if pd.notnull(solde_reassurance) and pd.notnull(part_primes) and pd.notnull(part_prestations) and pd.notnull(part_provisions) and pd.notnull(part_commissions) and pd.notnull(part_participation):
                C7 = solde_reassurance - (part_primes + part_prestations + part_provisions - part_commissions + part_participation)
                print(f"C7 calculé pour {col_name} : {C7}")
                c7_values.append(C7)
                if abs(C7) > 5:
                    invalid_c7_cols.append(col_idx)
            else:
                c7_values.append(float('nan'))
                print(f"Données manquantes dans la colonne {col_name} pour le calcul de C7")
        print(f"Colonnes C7 invalides (indices Excel) : {invalid_c7_cols}")

# -----------------------------C8 : Résultat technique=Solde de souscription+Charges d'acquisition et de gestion nettes+ Solde Financier+ Solde de réassurance / rétrocession ---------------------------------

    resultat_tech_rows = ["RESULTAT TECHNIQUE", "SOLDE DE SOUSCRIPTION", "CHARGES D'ACQUISITION ET DE GESTION NETTES", "SOLDE FINANCIER", "SOLDE DE REASSURANCE / RETROCESSION"]
    resultat_technique_idx = df.index[df['CATEGORIES'].str.contains("RESULTAT TECHNIQUE", case=False, na=False)].tolist()
    
    c8_values = [None]
    invalid_c8_cols = []
    if resultat_technique_idx and solde_souscription_idx and charges_acquisition_idx and solde_financier_idx and solde_reassurance_idx:
        print(f"Indices détectés - Résultat technique : {resultat_technique_idx}, Solde de souscription : {solde_souscription_idx}, Charges d'acquisition : {charges_acquisition_idx}, Solde Financier : {solde_financier_idx}, Solde de réassurance : {solde_reassurance_idx}")
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            resultat_technique = df.at[resultat_technique_idx[0], col_name] if resultat_technique_idx else float('nan')
            solde_souscription = df.at[solde_souscription_idx[0], col_name] if solde_souscription_idx else float('nan')
            charges_acquisition = df.at[charges_acquisition_idx[0], col_name] if charges_acquisition_idx else float('nan')
            solde_financier = df.at[solde_financier_idx[0], col_name] if solde_financier_idx else float('nan')
            solde_reassurance = df.at[solde_reassurance_idx[0], col_name] if solde_reassurance_idx else float('nan')
            
            print(f"Debug - Colonne {col_name} : Résultat technique={resultat_technique}, Solde de souscription={solde_souscription}, Charges d'acquisition={charges_acquisition}, Solde Financier={solde_financier}, Solde de réassurance={solde_reassurance}")
            if pd.notnull(resultat_technique) and pd.notnull(solde_souscription) and pd.notnull(charges_acquisition) and pd.notnull(solde_financier) and pd.notnull(solde_reassurance):
                C8 = resultat_technique - (solde_souscription + charges_acquisition + solde_financier + solde_reassurance)
                print(f"C8 calculé pour {col_name} : {C8}")
                c8_values.append(C8)
                if abs(C8) > 5:
                    invalid_c8_cols.append(col_idx)
            else:
                c8_values.append(float('nan'))
                print(f"Données manquantes dans la colonne {col_name} pour le calcul de C8")
        print(f"Colonnes C8 invalides (indices Excel) : {invalid_c8_cols}")
    else:
        print(f"Avertissement : Lignes manquantes pour C8 : {resultat_tech_rows}")


# -----------------------------C9 : Provisions pour Primes non Acquises clôture=Provisions pour Primes non Acquises Réouverture+Variation des Primes non acquises ---------------------------------

    provisions_rows = ["PROVISIONS POUR PRIMES NON ACQUISES CLOTURE", "PROVISIONS POUR PRIMES NON ACQUISES REOUVERTURE", "VARIATION DES PRIMES NON ACQUISES"]
    provisions_cloture_idx = df.index[df['CATEGORIES'].str.contains("PROVISIONS POUR PRIMES NON ACQUISES CLOTURE", case=False, na=False)].tolist()
    provisions_reouverture_idx = df.index[df['CATEGORIES'].str.contains("PROVISIONS POUR PRIMES NON ACQUISES REOUVERTURE", case=False, na=False)].tolist()
    var_acquises_idx = df.index[df['CATEGORIES'].str.contains("VARIATION DES PRIMES NON ACQUISES", case=False, na=False)].tolist()
    
    c9_values = [None]
    invalid_c9_cols = []
    if provisions_cloture_idx and provisions_reouverture_idx and var_acquises_idx:
        print(f"Indices détectés - Provisions clôture : {provisions_cloture_idx}, Provisions réouverture : {provisions_reouverture_idx}, Variation : {var_acquises_idx}")
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            provisions_cloture = df.at[provisions_cloture_idx[0], col_name] if provisions_cloture_idx else float('nan')
            provisions_reouverture = df.at[provisions_reouverture_idx[0], col_name] if provisions_reouverture_idx else float('nan')
            var_acquises = df.at[var_acquises_idx[0], col_name] if var_acquises_idx else float('nan')
            
            print(f"Debug - Colonne {col_name} : Provisions clôture={provisions_cloture}, Provisions réouverture={provisions_reouverture}, Variation={var_acquises}")
            if pd.notnull(provisions_cloture) and pd.notnull(provisions_reouverture) and pd.notnull(var_acquises):
                C9 = provisions_cloture - (provisions_reouverture - var_acquises)
                print(f"C9 calculé pour {col_name} : {C9}")
                c9_values.append(C9)
                if abs(C9) > 5:
                    invalid_c9_cols.append(col_idx)
            else:
                c9_values.append(float('nan'))
                print(f"Données manquantes dans la colonne {col_name} pour le calcul de C9")
        print(f"Colonnes C9 invalides (indices Excel) : {invalid_c9_cols}")
    else:
        print(f"Avertissement : Lignes manquantes pour C9 : {provisions_rows}")

# Ajoute les lignes C2 à C9 dans le fichier Excel-------------------------------
    c2_row_idx = ws.max_row + 1
    ws.cell(row=c2_row_idx, column=1, value="C2: PRIMES ACQUISES - (PRIMES EMISES + VARIATION DES PRIMES NON ACQUISES)")
    for col_idx, c2_val in enumerate(c2_values, 1):
        cell = ws.cell(row=c2_row_idx, column=col_idx)
        cell.value = int(c2_val) if pd.notnull(c2_val) and not np.isnan(c2_val) else None
        cell.number_format = '0'
        if pd.notnull(c2_val):
            value_str = str(int(c2_val)) if isinstance(c2_val, (int, float)) else str(c2_val)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    c3_row_idx = ws.max_row + 1
    ws.cell(row=c3_row_idx, column=1, value="C3: CHARGES DE PRESTATION - (PRESTATIONS ET FRAIS PAYES + CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE)")
    for col_idx, c3_val in enumerate(c3_values, 1):
        cell = ws.cell(row=c3_row_idx, column=col_idx)
        cell.value = int(c3_val) if pd.notnull(c3_val) and not np.isnan(c3_val) else None
        cell.number_format = '0'
        if pd.notnull(c3_val):
            value_str = str(int(c3_val)) if isinstance(c3_val, (int, float)) else str(c3_val)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    c4_row_idx = ws.max_row + 1
    ws.cell(row=c4_row_idx, column=1, value="C4: SOLDE DE SOUSCRIPTION - (PRIMES ACQUISES - CHARGES DE PRESTATION)")
    for col_idx, c4_val in enumerate(c4_values, 1):
        cell = ws.cell(row=c4_row_idx, column=col_idx)
        cell.value = int(c4_val) if pd.notnull(c4_val) and not np.isnan(c4_val) else None
        cell.number_format = '0'
        if pd.notnull(c4_val):
            value_str = str(int(c4_val)) if isinstance(c4_val, (int, float)) else str(c4_val)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    c5_row_idx = ws.max_row + 1
    ws.cell(row=c5_row_idx, column=1, value="C5: CHARGES D'ACQUISITION ET DE GESTION NETTES - (FRAIS D'ACQUISITION + AUTRES CHARGES DE GESTION NETTES)")
    for col_idx, c5_val in enumerate(c5_values, 1):
        cell = ws.cell(row=c5_row_idx, column=col_idx)
        cell.value = int(c5_val) if pd.notnull(c5_val) and not np.isnan(c5_val) else None
        cell.number_format = '0'
        if pd.notnull(c5_val):
            value_str = str(int(c5_val)) if isinstance(c5_val, (int, float)) else str(c5_val)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    c6_row_idx = ws.max_row + 1
    ws.cell(row=c6_row_idx, column=1, value="C6: SOLDE FINANCIER - (PRODUITS NETS DE PLACEMENTS - PARTICIPATION AUX RESULTATS)")
    for col_idx, c6_val in enumerate(c6_values, 1):
        cell = ws.cell(row=c6_row_idx, column=col_idx)
        cell.value = int(c6_val) if pd.notnull(c6_val) and not np.isnan(c6_val) else None
        cell.number_format = '0'
        if pd.notnull(c6_val):
            value_str = str(int(c6_val)) if isinstance(c6_val, (int, float)) else str(c6_val)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    c7_row_idx = ws.max_row + 1
    ws.cell(row=c7_row_idx, column=1, value="C7: SOLDE DE REASSURANCE / RETROCESSION - (PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRIMES ACQUISES + PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRESTATIONS PAYES + PART REASSUREURS /RETROCESSIONNAIRES DANS LES CHARGES DE PROVI. POUR PRESTATIONS - COMMISSIONS REÇUES DES REASSUREURS /RETROCESS + PART REASSUREURS /RETROCESSIONNAIRES DANS LA PARTICIPATION AUX RESULTATS)")
    for col_idx, c7_val in enumerate(c7_values, 1):
        cell = ws.cell(row=c7_row_idx, column=col_idx)
        cell.value = int(c7_val) if pd.notnull(c7_val) and not np.isnan(c7_val) else None
        cell.number_format = '0'
        if pd.notnull(c7_val):
            value_str = str(int(c7_val)) if isinstance(c7_val, (int, float)) else str(c7_val)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    c8_row_idx = ws.max_row + 1
    ws.cell(row=c8_row_idx, column=1, value="C8: RESULTAT TECHNIQUE - (SOLDE DE SOUSCRIPTION + CHARGES D'ACQUISITION ET DE GESTION NETTES + SOLDE FINANCIER + SOLDE DE REASSURANCE / RETROCESSION)")
    for col_idx, c8_val in enumerate(c8_values, 1):
        cell = ws.cell(row=c8_row_idx, column=col_idx)
        cell.value = int(c8_val) if pd.notnull(c8_val) and not np.isnan(c8_val) else None
        cell.number_format = '0'
        if pd.notnull(c8_val):
            value_str = str(int(c8_val)) if isinstance(c8_val, (int, float)) else str(c8_val)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)

    c9_row_idx = ws.max_row + 1
    ws.cell(row=c9_row_idx, column=1, value="C9: PROVISIONS POUR PRIMES NON ACQUISES CLOTURE - (PROVISIONS POUR PRIMES NON ACQUISES REOUVERTURE - VARIATION DES PRIMES NON ACQUISES)")
    for col_idx, c9_val in enumerate(c9_values, 1):
        cell = ws.cell(row=c9_row_idx, column=col_idx)
        cell.value = int(c9_val) if pd.notnull(c9_val) and not np.isnan(c9_val) else None
        cell.number_format = '0'
        if pd.notnull(c9_val):
            value_str = str(int(c9_val)) if isinstance(c9_val, (int, float)) else str(c9_val)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, len(value_str) + 2)


# Applique des couleurs pour signaler les erreurs dans C1 à C9------------------------------------
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=c1_col_idx)
        if pd.notnull(cell.value) and ws.cell(row=row_idx, column=1).value in target_rows:
            c1_value = abs(float(cell.value)) if pd.notnull(cell.value) else 0
            if c1_value > 5:
                if c1_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")

    for col_idx in invalid_c2_cols:
        cell = ws.cell(row=c2_row_idx, column=col_idx)
        if pd.notnull(cell.value):
            c2_value = abs(float(cell.value))
            if c2_value > 5:
                if c2_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")

    for col_idx in invalid_c3_cols:
        cell = ws.cell(row=c3_row_idx, column=col_idx)
        if pd.notnull(cell.value):
            c3_value = abs(float(cell.value))
            if c3_value > 5:
                if c3_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")

    for col_idx in invalid_c4_cols:
        cell = ws.cell(row=c4_row_idx, column=col_idx)
        if pd.notnull(cell.value):
            c4_value = abs(float(cell.value))
            if c4_value > 5:
                if c4_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")

    for col_idx in invalid_c5_cols:
        cell = ws.cell(row=c5_row_idx, column=col_idx)
        if pd.notnull(cell.value):
            c5_value = abs(float(cell.value))
            if c5_value > 5:
                if c5_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")

    for col_idx in invalid_c6_cols:
        cell = ws.cell(row=c6_row_idx, column=col_idx)
        if pd.notnull(cell.value):
            c6_value = abs(float(cell.value))
            if c6_value > 5:
                if c6_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")

    for col_idx in invalid_c7_cols:
        cell = ws.cell(row=c7_row_idx, column=col_idx)
        if pd.notnull(cell.value):
            c7_value = abs(float(cell.value))
            if c7_value > 5:
                if c7_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")

    for col_idx in invalid_c8_cols:
        cell = ws.cell(row=c8_row_idx, column=col_idx)
        if pd.notnull(cell.value):
            c8_value = abs(float(cell.value))
            if c8_value > 5:
                if c8_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")

    for col_idx in invalid_c9_cols:
        cell = ws.cell(row=c9_row_idx, column=col_idx)
        if pd.notnull(cell.value):
            c9_value = abs(float(cell.value))
            if c9_value > 5:
                if c9_value < 1000:
                    cell.fill = light_orange_fill
                    cell.font = bold_black_font
                else:
                    cell.fill = medium_orange_fill
                    cell.font = bold_white_font
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(color="000000")



# Identifie les cellules erronées pour C1 à C9.
    ListeCellulesRougesFinales = []

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value in target_rows:
            c1_value = float(ws.cell(row=row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c1_value) > 5:
                numeric_sum = sum(float(ws.cell(row=row_idx, column=col_map[col]).value) for col in numeric_cols if pd.notnull(ws.cell(row=row_idx, column=col_map[col]).value))
                total_value = float(ws.cell(row=row_idx, column=col_map[total_col]).value) if pd.notnull(ws.cell(row=row_idx, column=col_map[total_col]).value) else 0
                max_contribution = 0
                target_cell = None
                for col_idx, col_name in enumerate(numeric_cols, 2):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if pd.notnull(cell_value) and cell_value != 0:
                        contribution = abs(cell_value) / abs(numeric_sum) if numeric_sum != 0 else 0
                        if contribution > 0.3:
                            max_contribution = contribution
                            target_cell = (row_idx, col_idx)
                if target_cell and max_contribution > 0.3:
                    ListeCellulesRougesFinales.append(target_cell)
                    print(f"Cellule erronée détectée pour C1 à la ligne {row_idx}, colonne {target_cell[1]} (contribution={max_contribution})")


# --------------------------------------------La double validation commence ici pour chaque condition : -------------------------------------------------------------------------------------------------------

# Identifie les cellules erronées pour C2
    if primes_acquises_idx and primes_emises_idx and var_acquises_idx:
        for col_idx in invalid_c2_cols:
            c2_value = float(ws.cell(row=c2_row_idx, column=col_idx).value) if pd.notnull(ws.cell(row=c2_row_idx, column=col_idx).value) else float('nan')
            c1_value = float(ws.cell(row=c2_row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=c2_row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c2_value) > 5 and abs(c1_value) > 5:
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value in primes_rows:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if pd.notnull(cell_value) and cell_value != 0:
                            contribution = abs(cell_value) / abs(c2_value) if c2_value != 0 else 0
                            if contribution > 0.3:
                                ListeCellulesRougesFinales.append((row_idx, col_idx))
                                print(f"Cellule erronée détectée pour C2 à la ligne {row_idx}, colonne {col_idx} (contribution={contribution})")
                                break


# Identifie les cellules erronées pour C3
    if charges_prestation_idx and prestations_payes_idx and charges_provisions_idx:
        for col_idx in invalid_c3_cols:
            c3_value = float(ws.cell(row=c3_row_idx, column=col_idx).value) if pd.notnull(ws.cell(row=c3_row_idx, column=col_idx).value) else float('nan')
            c1_value = float(ws.cell(row=c3_row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=c3_row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c3_value) > 5 and abs(c1_value) > 5:
                max_contribution = 0
                target_cell = None
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value in charges_rows:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if pd.notnull(cell_value) and cell_value != 0:
                            contribution = abs(cell_value) / abs(c3_value) if c3_value != 0 else 0
                            if contribution > 0.3:
                                max_contribution = contribution
                                target_cell = (row_idx, col_idx)
                                break
                if target_cell and max_contribution > 0.3:
                    ListeCellulesRougesFinales.append(target_cell)
                    print(f"Cellule erronée détectée pour C3 à la ligne {target_cell[0]}, colonne {col_idx} (contribution={max_contribution})")

# Identifie les cellules erronées pour C4
    if solde_souscription_idx and primes_acquises_idx and charges_prestation_idx:
        for col_idx in invalid_c4_cols:
            c4_value = float(ws.cell(row=c4_row_idx, column=col_idx).value) if pd.notnull(ws.cell(row=c4_row_idx, column=col_idx).value) else float('nan')
            c1_value = float(ws.cell(row=c4_row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=c4_row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c4_value) > 5 and abs(c1_value) > 5:
                max_contribution = 0
                target_cell = None
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value in solde_rows + ["PRIMES ACQUISES", "CHARGES DE PRESTATION"]:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if pd.notnull(cell_value) and cell_value != 0:
                            contribution = abs(cell_value) / abs(c4_value) if c4_value != 0 else 0
                            if contribution > 0.3:
                                max_contribution = contribution
                                target_cell = (row_idx, col_idx)
                                break
                if target_cell and max_contribution > 0.3:
                    ListeCellulesRougesFinales.append(target_cell)
                    print(f"Cellule erronée détectée pour C4 à la ligne {target_cell[0]}, colonne {col_idx} (contribution={max_contribution})")

# Identifie les cellules erronées pour C5
    if charges_acquisition_idx and frais_acquisition_idx and autres_charges_idx:
        for col_idx in invalid_c5_cols:
            c5_value = float(ws.cell(row=c5_row_idx, column=col_idx).value) if pd.notnull(ws.cell(row=c5_row_idx, column=col_idx).value) else float('nan')
            c1_value = float(ws.cell(row=c5_row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=c5_row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c5_value) > 5 and abs(c1_value) > 5:
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value in acquisition_rows:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if pd.notnull(cell_value) and cell_value != 0:
                            contribution = abs(cell_value) / abs(c5_value) if c5_value != 0 else 0
                            if contribution > 0.3:
                                ListeCellulesRougesFinales.append((row_idx, col_idx))
                                print(f"Cellule erronée détectée pour C5 à la ligne {row_idx}, colonne {col_idx} (contribution={contribution})")
                                break

# Identifie les cellules erronées pour C6
    if solde_financier_idx and produits_nets_idx and participation_idx:
        for col_idx in invalid_c6_cols:
            c6_value = float(ws.cell(row=c6_row_idx, column=col_idx).value) if pd.notnull(ws.cell(row=c6_row_idx, column=col_idx).value) else float('nan')
            c1_value = float(ws.cell(row=c6_row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=c6_row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c6_value) > 5 and abs(c1_value) > 5:
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value in financier_rows:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if pd.notnull(cell_value) and cell_value != 0:
                            contribution = abs(cell_value) / abs(c6_value) if c6_value != 0 else 0
                            if contribution > 0.3:
                                ListeCellulesRougesFinales.append((row_idx, col_idx))
                                print(f"Cellule erronée détectée pour C6 à la ligne {row_idx}, colonne {col_idx} (contribution={contribution})")
                                break

# Identifie les cellules erronées pour C7
    if solde_reassurance_idx and part_primes_idx and part_prestations_idx and part_provisions_idx and part_commissions_idx and part_participation_idx:
        for col_idx in invalid_c7_cols:
            c7_value = float(ws.cell(row=c7_row_idx, column=col_idx).value) if pd.notnull(ws.cell(row=c7_row_idx, column=col_idx).value) else float('nan')
            c1_value = float(ws.cell(row=c7_row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=c7_row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c7_value) > 5 and abs(c1_value) > 5:
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value in reassurance_rows:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if pd.notnull(cell_value) and cell_value != 0:
                            contribution = abs(cell_value) / abs(c7_value) if c7_value != 0 else 0
                            if contribution > 0.3:
                                ListeCellulesRougesFinales.append((row_idx, col_idx))
                                print(f"Cellule erronée détectée pour C7 à la ligne {row_idx}, colonne {col_idx} (contribution={contribution})")
                                break

 # Identifie les cellules erronées pour C8
    if resultat_technique_idx and solde_souscription_idx and charges_acquisition_idx and solde_financier_idx and solde_reassurance_idx:
        for col_idx in invalid_c8_cols:
            c8_value = float(ws.cell(row=c8_row_idx, column=col_idx).value) if pd.notnull(ws.cell(row=c8_row_idx, column=col_idx).value) else float('nan')
            c1_value = float(ws.cell(row=c8_row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=c8_row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c8_value) > 5 and abs(c1_value) > 5:
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value in resultat_tech_rows:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if pd.notnull(cell_value) and cell_value != 0:
                            contribution = abs(cell_value) / abs(c8_value) if c8_value != 0 else 0
                            if contribution > 0.3:
                                ListeCellulesRougesFinales.append((row_idx, col_idx))
                                print(f"Cellule erronée détectée pour C8 à la ligne {row_idx}, colonne {col_idx} (contribution={contribution})")
                                break

# Identifie les cellules erronées pour C9
    if provisions_cloture_idx and provisions_reouverture_idx and var_acquises_idx:
        for col_idx in invalid_c9_cols:
            c9_value = float(ws.cell(row=c9_row_idx, column=col_idx).value) if pd.notnull(ws.cell(row=c9_row_idx, column=col_idx).value) else float('nan')
            c1_value = float(ws.cell(row=c9_row_idx, column=c1_col_idx).value) if pd.notnull(ws.cell(row=c9_row_idx, column=c1_col_idx).value) else float('nan')
            if abs(c9_value) > 5 and abs(c1_value) > 5:
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value in provisions_rows:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if pd.notnull(cell_value) and cell_value != 0:
                            contribution = abs(cell_value) / abs(c9_value) if c9_value != 0 else 0
                            if contribution > 0.3:
                                ListeCellulesRougesFinales.append((row_idx, col_idx))
                                print(f"Cellule erronée détectée pour C9 à la ligne {row_idx}, colonne {col_idx} (contribution={contribution})")
                                break

    ListeCellulesRougesFinales = list(set(ListeCellulesRougesFinales))

 # Applique la couleur rouge aux cellules erronées détectées
    for row_idx, col_idx in corrected_cells:
        cell = ws.cell(row=row_idx, column=col_idx)
        if (row_idx, col_idx) in ListeCellulesRougesFinales:
            cell.fill = intersection_red_fill
            cell.font = bold_white_font
        else:
            cell.fill = corrected_green_fill
            cell.font = bold_black_font

    # Applique la couleur verte aux cellules corrigées (symboles ou lettres supprimés)
    for row_idx, col_idx in ListeCellulesRougesFinales:
        if (row_idx, col_idx) not in corrected_cells:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = intersection_red_fill
            cell.font = bold_white_font

    for row_idx, col_idx in ListeCellulesRougesFinales:
        cell = ws.cell(row=row_idx, column=col_idx)
        category = ws.cell(row=row_idx, column=1).value
        if category in primes_rows and col_idx in invalid_c2_cols:
            cell.comment = openpyxl.comments.Comment("Erreur C2: Incohérence dans PRIMES ACQUISES - (PRIMES EMISES + VARIATION DES PRIMES NON ACQUISES)", "Validation Script")
        elif category in charges_rows and col_idx in invalid_c3_cols:
            cell.comment = openpyxl.comments.Comment("Erreur C3: Incohérence dans CHARGES DE PRESTATION - (PRESTATIONS ET FRAIS PAYES + CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE)", "Validation Script")
        elif category in solde_rows and col_idx in invalid_c4_cols:
            cell.comment = openpyxl.comments.Comment("Erreur C4: Incohérence dans SOLDE DE SOUSCRIPTION - (PRIMES ACQUISES - CHARGES DE PRESTATION)", "Validation Script")
        elif category in acquisition_rows and col_idx in invalid_c5_cols:
            cell.comment = openpyxl.comments.Comment("Erreur C5: Incohérence dans CHARGES D'ACQUISITION ET DE GESTION NETTES - (FRAIS D'ACQUISITION + AUTRES CHARGES DE GESTION NETTES)", "Validation Script")
        elif category in financier_rows and col_idx in invalid_c6_cols:
            cell.comment = openpyxl.comments.Comment("Erreur C6: Incohérence dans SOLDE FINANCIER - (PRODUITS NETS DE PLACEMENTS - PARTICIPATION AUX RESULTATS)", "Validation Script")
        elif category in reassurance_rows and col_idx in invalid_c7_cols:
            cell.comment = openpyxl.comments.Comment("Erreur C7: Incohérence dans SOLDE DE REASSURANCE / RETROCESSION - (PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRIMES ACQUISES + PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRESTATIONS PAYES + PART REASSUREURS /RETROCESSIONNAIRES DANS LES CHARGES DE PROVI. POUR PRESTATIONS - COMMISSIONS REÇUES DES REASSUREURS /RETROCESS + PART REASSUREURS /RETROCESSIONNAIRES DANS LA PARTICIPATION AUX RESULTATS)", "Validation Script")
        elif category in resultat_tech_rows and col_idx in invalid_c8_cols:
            cell.comment = openpyxl.comments.Comment("Erreur C8: Incohérence dans RESULTAT TECHNIQUE - (SOLDE DE SOUSCRIPTION + CHARGES D'ACQUISITION ET DE GESTION NETTES + SOLDE FINANCIER + SOLDE DE REASSURANCE / RETROCESSION)", "Validation Script")
        elif category in provisions_rows and col_idx in invalid_c9_cols:
            cell.comment = openpyxl.comments.Comment("Erreur C9: Incohérence dans PROVISIONS POUR PRIMES NON ACQUISES CLOTURE - (PROVISIONS POUR PRIMES NON ACQUISES REOUVERTURE - VARIATION DES PRIMES NON ACQUISES)", "Validation Script")

    #------------------------------------ Vérification de la validité du fichier (Valide si aucune cellule rouge, Invalide sinon)-------------------------------------
    file_status = "Valide" if not ListeCellulesRougesFinales else "Invalide"
    print(f"Statut du fichier : {file_status}")



                                #---------------Si fichie Valide ------------------#
    try:
        wb.save(output_file)
        print(f"Fichier final sauvegardé : {output_file}")
        # Ouvre automatiquement le fichier de sortie
        if platform.system() == "Windows":
            os.startfile(output_file)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(["open", output_file])
        else:  # Linux et autres
            subprocess.run(["xdg-open", output_file])
        

                                #---------------Si fichier Invalide ------------------#


        # Si le fichier est invalide, attendre CTRL+S pour fermer
        if file_status == "Invalide":
            print("Fichier invalide détecté. Appuyez sur CTRL+S dans le fichier Excel pour le fermer et terminer l'exécution.")
            while True:
                if keyboard.is_pressed('ctrl+s'):
                    print("CTRL+S détecté, fermeture du fichier Excel...")
                    wb.close()
                    print(f"Fichier Excel '{output_file}' fermé.")
                    break
    except Exception as e:
        raise ValueError(f"Erreur lors de l'enregistrement ou de l'ouverture du fichier Excel '{output_file}' : {str(e)}")

    return output_file, file_status
 



                                #---------------Boucle d'exécution du code ------------------#


def run_validation_loop(input_file):
    import time
    import openpyxl
    import pandas as pd
    import mysql.connector
    from mysql.connector import Error
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_filename = os.path.basename(input_file)
    input_path = os.path.join(script_dir, input_filename)
    
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Le fichier '{input_filename}' n'a pas été trouvé dans le dossier du projet.")
    
    while True:
        # Charger le fichier Excel 
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        
        # Trouver et supprimer la colonne C1 si existante pour que si le fichier est invalide 
        # et qu'il sera envoyé vers la fonction de validation son nouveau C1 sera recalculé après toute modification faite
        max_col = ws.max_column
        c1_col_idx = None
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=1, column=col_idx).value == 'C1':
                c1_col_idx = col_idx
                break
        
        if c1_col_idx:
            ws.delete_cols(c1_col_idx, 1)
            print(f"Colonne C1 supprimée à l'index {c1_col_idx}.")
        

         #Même travail pour les lignes C2-C9
        c_rows = [
            "C2: PRIMES ACQUISES - (PRIMES EMISES + VARIATION DES PRIMES NON ACQUISES)",
            "C3: CHARGES DE PRESTATION - (PRESTATIONS ET FRAIS PAYES + CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE)",
            "C4: SOLDE DE SOUSCRIPTION - (PRIMES ACQUISES - CHARGES DE PRESTATION)",
            "C5: CHARGES D'ACQUISITION ET DE GESTION NETTES - (FRAIS D'ACQUISITION + AUTRES CHARGES DE GESTION NETTES)",
            "C6: SOLDE FINANCIER - (PRODUITS NETS DE PLACEMENTS - PARTICIPATION AUX RESULTATS)",
            "C7: SOLDE DE REASSURANCE / RETROCESSION - (PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRIMES ACQUISES + PART REASSUREURS /RETROCESSIONNAIRES DANS LES PRESTATIONS PAYES + PART REASSUREURS /RETROCESSIONNAIRES DANS LES CHARGES DE PROVI. POUR PRESTATIONS - COMMISSIONS REÇUES DES REASSUREURS /RETROCESS + PART REASSUREURS /RETROCESSIONNAIRES DANS LA PARTICIPATION AUX RESULTATS)",
            "C8: RESULTAT TECHNIQUE - (SOLDE DE SOUSCRIPTION + CHARGES D'ACQUISITION ET DE GESTION NETTES + SOLDE FINANCIER + SOLDE DE REASSURANCE / RETROCESSION)",
            "C9: PROVISIONS POUR PRIMES NON ACQUISES CLOTURE - (PROVISIONS POUR PRIMES NON ACQUISES REOUVERTURE - VARIATION DES PRIMES NON ACQUISES)"
        ]
        rows_to_delete = []
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value in c_rows:
                rows_to_delete.append(row_idx)
                print(f"Ligne {row_idx} supprimée (contenait {cell_value}).")
        
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx, 1)
        
        # Sauvegarder le fichier nettoyé
        wb.save(input_path)
        wb.close()
        
        # S'assurer que C1-C9 ont bien été enlevées
        df = pd.read_excel(input_path, thousands=" ")
        if 'C1' in df.columns:
            df = df.drop(columns=['C1'])
            print("Colonne C1 supprimée du DataFrame.")
        
        df = df[~df['CATEGORIES'].isin(c_rows)]
        print(f"Lignes C2-C9 supprimées du DataFrame : {c_rows}")
        
        # Save the DataFrame back to the Excel file
        with pd.ExcelWriter(input_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False)
        
        # Valider le fichier nettoyé avant revalidation
        output_file, file_status = validate_excel(input_path)
        print(f"Fichier validé : {output_file}, Statut : {file_status}")


        #----------------------------------------------------------------(CAS VALIDE)------------------------------------------------
        if file_status == "Valide":
            print("Fichier valide, fin du processus.")
            # Insert the valid file into the MySQL database
            try:
                connection = mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="",
                    database="cmf"
                )
                cursor = connection.cursor()
                insert_query = "INSERT INTO documentvalide (filename) VALUES (%s)"
                cursor.execute(insert_query, (os.path.basename(output_file),))
                connection.commit()
                print(f"Fichier '{os.path.basename(output_file)}' inséré dans la table documentvalide.")
            except Error as e:
                print(f"Erreur lors de l'insertion dans la base de données : {e}")
            finally:
                if connection.is_connected():
                    cursor.close()
                    connection.close()
                    print("Connexion MySQL fermée.")
            break
         #----------------------------------------------------------------(CAS InALIDE)------------------------------------------------
        print("Fichier invalide détecté. Appuyez sur CTRL+S dans le fichier Excel pour le révalider.")
        while True:
            if keyboard.is_pressed('ctrl+s'):
                print("CTRL+S détecté, fermeture du fichier Excel...")
                close_excel_file(output_file)
                print(f"Fichier Excel '{output_file}' fermé.")
                input_path = output_file  # output file ywali howa el new input_path mte3na 
                break
            time.sleep(0.1)  # Prevent excessive CPU usage
    
    return output_file, file_status
        #------------------------------------------------------Partie Main du code :---------------------------------------------------------------
if __name__ == "__main__":
    import sys
    try:
        input_file = sys.argv[1] if len(sys.argv) > 1 else None # ici vérifie si un argument est fourni 
        output_file = normalize_excel(input_file)
        run_validation_loop(output_file)
    except Exception as e:
        print(f"Une erreur s'est produite : {str(e)}")