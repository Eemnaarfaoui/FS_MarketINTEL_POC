"""
Excel Exporter Module
Exports hierarchical financial data to Excel with formatting
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import shutil
import traceback

def export_to_excel(hierarchical_data, company_name, pdf_path, output_name, year_n, year_n_1):
    """
    Export hierarchical data to Excel with proper structure and formatting
    
    Args:
        hierarchical_data: List of dicts with hierarchy information
        output_name: Output Excel filename
        year_n: Current year (e.g., 2024)
        year_n_1: Previous year (e.g., 2023)
    """
    # Create company folder
        # ----------------------------------
    safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in company_name)
      # Use abbreviated name if too long
    if len(safe_name) > 30:
        safe_name = safe_name[:27] + "_"
    folder_path = os.path.join(os.getcwd(), "outputs", safe_name)
    os.makedirs(folder_path, exist_ok=True)

        # ----------------------------------
        #  Copy PDF into the folder
        # ----------------------------------
    pdf_filename = os.path.basename(pdf_path)
    pdf_dest = os.path.join(folder_path, pdf_filename)
    if os.path.abspath(pdf_path) != os.path.abspath(pdf_dest):
        shutil.copy2(pdf_path, pdf_dest)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "CAPITAUX PROPRES ET PASSIF"
        
        # Styles
        header_font = Font(bold=True, size=11)
        total_font = Font(bold=True, size=10)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        
        number_format = '#,##0'
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws.append(['Type', 'Sous-catégorie', 'Code', 'Description', f'31/12/{year_n}', f'31/12/{year_n_1}', 'ValidationResult'])
        for col in range(1, 8):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        current_row = 2
        for item in hierarchical_data:
            level = item.get('level', 2)
            code = item.get('code', '')
            description = item.get('description', '')
            is_total = item.get('is_total', False)
            category = item.get('category', '')
            subcategory = item.get('subcategory', '')
            values = item.get('values', [])
            validation_result = item.get('ValidationResult', '')

            # Skip section headers (level 1)
            if level == 1 and category == "SECTION":
                continue
            # Skip main title (level 0)
            if level == 0:
                continue
            indent = "  " * max(0, level - 1)
            ws.cell(current_row, 1, category)
            ws.cell(current_row, 2, subcategory)
            ws.cell(current_row, 3, code)
            ws.cell(current_row, 4, f"{indent}{description}")
            for i, value in enumerate(values[:2], start=5):
                if value and value != '':
                    ws.cell(current_row, i, value)
                    ws.cell(current_row, i).number_format = number_format
            ws.cell(current_row, 7, validation_result)
            for col in range(1, 8):
                cell = ws.cell(current_row, col)
                cell.border = border
                if is_total or level == 2:
                    cell.font = total_font if is_total else bold_font
                else:
                    cell.font = normal_font
                if col <= 4:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            current_row += 1
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:G{current_row-1}"
        excel_path = os.path.join(folder_path, output_name)
        wb.save(excel_path)
        
        '''wb.save(output_name)
        print(f"✅ Fichier Excel créé : {os.path.abspath(output_name)}")
        return True'''
        return True
    except Exception as e:
        tb = traceback.format_exc()
        print(f"❌ Erreur export Excel : {str(e)}\n{tb}")
        return f"{str(e)}\n{tb}"

def beautify_excel_layout(excel_path, company_name=None):
    """
    Apply better formatting to an Excel file
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    # Insert company name as a field above the header if provided
    if company_name:
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        cell = ws.cell(1, 1)
        cell.value = f"Company: {company_name}"
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Define styles
    header_font = Font(bold=True, size=11)
    total_font = Font(bold=True, size=10)
    bold_font = Font(bold=True, size=10)
    
    return True
    normal_font = Font(size=10)
    number_format = '#,##0'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header styling
    header_row = 2 if company_name else 1
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows styling
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.border = border
            if col <= 4:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.font = normal_font

    # Set column widths (adjust as needed)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 55
    for col in range(5, ws.max_column + 1):
        ws.column_dimensions[chr(64+col)].width = 15

    # Freeze first row after company name (or header if no company name)
    ws.freeze_panes = f'A{header_row+1}'
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)
    print(f'Beautified layout saved to: {excel_path}')
