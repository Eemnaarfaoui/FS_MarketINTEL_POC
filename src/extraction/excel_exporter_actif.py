from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def export_actif_to_excel(actif_data, output_name_actif, year, year_1):
    wb = Workbook()
    ws = wb.active
    ws.title = "ACTIF"
    # Styles
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    number_format = '#,##0'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    headers = ["DESIGNATION", "BRUT", "AMORT_PROV", f"NET_N {year}", f"NET_N1 {year_1}"]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    for i, row in enumerate(actif_data, start=2):
        ws.cell(i, 1, row.get("DESIGNATION", ""))
        ws.cell(i, 2, row.get("BRUT", ""))
        ws.cell(i, 3, row.get("AMORT_PROV", ""))
        ws.cell(i, 4, row.get("NET_N", ""))
        ws.cell(i, 5, row.get("NET_N1", ""))
        
        for col in range(1, 7):
            cell = ws.cell(i, col)
            cell.font = normal_font
            cell.border = border
            if col > 1:
                cell.number_format = number_format
    ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
    wb.save(output_name_actif)
    print(f"✅ Fichier Excel ACTIF généré : {output_name_actif}")
    return True
