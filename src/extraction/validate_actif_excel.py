import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os


TOLERANCE = 5


# ==========================
# Nettoyage numérique robuste
# ==========================
def clean_number(x):
    """
    Convertit en float.
    Valeur vide ou invalide → 0.0
    """
    if pd.isna(x):
        return 0.0

    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).replace("\u00a0", "").replace(" ", "").replace(",", ".").strip()

    if s == "":
        return 0.0

    try:
        return float(s)
    except:
        return 0.0


# ==========================
# Extraction code AC
# ==========================
def extract_ac_code(designation):
    if pd.isna(designation):
        return None

    text = str(designation).upper()
    match = re.search(r"\bAC\s*(\d+)\b", text)

    if match:
        return "AC" + match.group(1)

    return None


# ==========================
# Merge doublons AC
# ==========================
def merge_duplicate_ac_rows(df):

    df["AC_CODE"] = df["DESIGNATION"].apply(extract_ac_code)

    df_with_code = df[df["AC_CODE"].notna()].copy()
    df_without_code = df[df["AC_CODE"].isna()].copy()

    cols_numeric = ["BRUT", "AMORT_PROV", "NET_N", "NET_N1"]

    merged_rows = []

    for code, group in df_with_code.groupby("AC_CODE", sort=False):

        # meilleure designation (la plus longue)
        designation_best = group["DESIGNATION"].astype(str).iloc[
            group["DESIGNATION"].astype(str).str.len().argmax()
        ]

        row = {
            "DESIGNATION": designation_best,
            "AC_CODE": code
        }

        # prendre la première valeur non nulle
        for col in cols_numeric:
            values = group[col].dropna().tolist()
            row[col] = float(values[0]) if values else 0.0

        merged_rows.append(row)

    df_merged = pd.DataFrame(merged_rows)

    df_final = pd.concat([df_merged, df_without_code], ignore_index=True)

    return df_final


# ==========================
# Validation principale
# ==========================
def validate_actif_from_data(data_actifs, assurance_name, annee, output_xlsx):

    df = pd.DataFrame(data_actifs)

    required_cols = ["DESIGNATION", "BRUT", "AMORT_PROV", "NET_N", "NET_N1"]
    for col in required_cols:
        if col not in df.columns:
            raise Exception(f"Colonne manquante dans ACTIF : {col}")

    # ==========================
    # Nettoyage numérique sécurisé
    # ==========================
    numeric_cols = ["BRUT", "AMORT_PROV", "NET_N", "NET_N1"]

    for col in numeric_cols:
        df[col] = df[col].apply(clean_number)
        df[col] = df[col].fillna(0.0).astype(float)

    df = df.dropna(subset=["DESIGNATION"])

    # ==========================
    # Merge doublons AC
    # ==========================
    df = merge_duplicate_ac_rows(df)

    # ==========================
    # Ajout metadata
    # ==========================
    df.insert(0, "ASSURANCE", assurance_name)
    df.insert(1, "ANNEE", annee)

    # ==========================
    # Validation calcul
    # ==========================
    df["CALC_NET"] = (df["BRUT"] - df["AMORT_PROV"]).round(2)
    df["DIFF"] = (df["NET_N"] - df["CALC_NET"]).round(2)

    df["STATUS"] = df["DIFF"].apply(
        lambda x: "OK" if abs(x) <= TOLERANCE else "NOT_OK"
    )

    # ==========================
    # Export Excel
    # ==========================
    wb = Workbook()
    ws = wb.active
    ws.title = "ACTIF_VALIDATION"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.append(df.columns.tolist())

    for _, row in df.iterrows():
        ws.append(row.tolist())
        current_row = ws.max_row

        fill = green_fill if row["STATUS"] == "OK" else red_fill

        cols_to_color = ["BRUT", "AMORT_PROV", "NET_N", "CALC_NET", "DIFF", "STATUS"]

        for col_name in cols_to_color:
            col_index = df.columns.get_loc(col_name) + 1
            ws.cell(row=current_row, column=col_index).fill = fill

    # ==========================
    # Format currency Excel
    # ==========================
    currency_format = '#,##0.00'

    for col_name in ["BRUT", "AMORT_PROV", "NET_N", "NET_N1", "CALC_NET", "DIFF"]:
        col_index = df.columns.get_loc(col_name) + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_index).number_format = currency_format

    # ==========================
    # Sauvegarde
    # ==========================
    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    wb.save(output_xlsx)

    return output_xlsx