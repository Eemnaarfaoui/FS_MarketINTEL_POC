import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from src.extraction.validator_passifs import ValidatorPassifs


def validate_capitaux_propres_passif(excel_path: str, company_name: str):
    EXCEL_PATH = excel_path
    COMPANY_NAME = company_name

    df = pd.read_excel(EXCEL_PATH)

    year_cols = [col for col in df.columns if col.startswith('31/12/')]

    def is_row_empty(row):
        code_empty = pd.isna(row.get('Code')) or str(row.get('Code')).strip() == ''
        desc_empty = pd.isna(row.get('Sous-catégorie')) or str(row.get('Sous-catégorie')).strip() == ''

        def is_empty_or_year(val):
            if pd.isna(val) or val == 0:
                return True
            if isinstance(val, (int, float)) and 2020 <= val <= 2030:
                return True
            return False

        years_empty = all(is_empty_or_year(row.get(col)) for col in year_cols)

        desc = str(row.get('Description', '')).strip().upper()
        is_header_junk = desc in ('', 'DESIGNATION', 'MONTANT MONTANT', 'MONTANT')

        return (code_empty and desc_empty and years_empty) or (code_empty and is_header_junk)

    filtered_df = df[~df.apply(is_row_empty, axis=1)].copy()

    # ══════════════════════════════════════════════════════════════
    # BUILD CONTEXT: collect all code → value mappings from the data
    # The validator needs this to check parent = sum(children)
    # ══════════════════════════════════════════════════════════════
    first_year_col = year_cols[0] if year_cols else None

    extracted_data_context = {}
    for _, row in filtered_df.iterrows():
        code = str(row.get('Code', '')).strip()
        if not code or code == 'nan':
            continue

        # Get the first year column value
        value = None
        if first_year_col:
            raw_val = row.get(first_year_col)
            if pd.notna(raw_val):
                try:
                    value = float(raw_val)
                except (TypeError, ValueError):
                    value = None

        # Don't overwrite an existing value with None
        # (e.g. CP5 appears twice: once with value, once without)
        if code in extracted_data_context and value is None:
            pass  # Keep existing value
        else:
            extracted_data_context[code] = {'value': value}

    print(f"\n📊 Context built: {len(extracted_data_context)} codes")
    for code, data in sorted(extracted_data_context.items()):
        v = data['value']
        print(f"  {code}: {v:,.0f}" if v is not None else f"  {code}: None")

    # ══════════════════════════════════════════════════════════════
    # VALIDATE: pass context to validator so rules can check sums
    # ══════════════════════════════════════════════════════════════
    validator = ValidatorPassifs(extracted_data_context=extracted_data_context)

    def validate_row(row):
        # Build a dict the validator understands:
        #   'description' → from Description column
        #   'values' → list of year column values
        validator_row = {
            'description': str(row.get('Description', '')),
            'values': [row.get(col) for col in year_cols],
            'code': str(row.get('Code', '')).strip(),
        }
        result = validator.validate(validator_row)
        return 'PASS' if result else 'FAIL'

    filtered_df['ValidationResult'] = filtered_df.apply(validate_row, axis=1)
    filtered_df['Assurance'] = COMPANY_NAME

    # Convert year columns to numeric so Excel treats them as numbers
    for col in year_cols:
        filtered_df[col] = pd.to_numeric(filtered_df[col], errors='coerce')

    # Count results
    pass_count = (filtered_df['ValidationResult'] == 'PASS').sum()
    fail_count = (filtered_df['ValidationResult'] == 'FAIL').sum()
    print(f"\n✅ PASS: {pass_count}  ❌ FAIL: {fail_count}")

    # Save raw data first
    output_path = EXCEL_PATH.replace('.xlsx', '_validated.xlsx')
    filtered_df.to_excel(output_path, index=False)

    # ── Beautify with openpyxl ──
    wb = load_workbook(output_path)
    ws = wb.active

    # Style definitions
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    pass_font = Font(color="006100", bold=True)
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    fail_font = Font(color="9C0006", bold=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    number_fmt = '#,##0'

    # Find column indices
    col_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        col_map[cell.value] = col_idx

    validation_col = col_map.get('ValidationResult')
    year_col_indices = [col_map[c] for c in year_cols if c in col_map]

    # Style headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Style data rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin_border

            # Number formatting for year columns
            if col_idx in year_col_indices:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal="right")

        # Color the validation result cell
        if validation_col:
            val_cell = ws.cell(row_idx, validation_col)
            if val_cell.value == 'PASS':
                val_cell.fill = pass_fill
                val_cell.font = pass_font
            elif val_cell.value == 'FAIL':
                val_cell.fill = fail_fill
                val_cell.font = fail_font
            val_cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths (approximate)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = ws.cell(1, col_idx).column_letter
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            val = ws.cell(row_idx, col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f'\nValidation complete. Output saved to: {output_path}')
    return output_path


if __name__ == "__main__":
    validate_capitaux_propres_passif(
        excel_path=r"C:/Users/USER/Documents/GitHub/FS_MarketINTEL_POC/outputs/Sté__TUNISIENNE_D_ASSURANCES_-_LLOYD_TUNISIEN_-/LLOYD_TUNISIE_2024_passif.xlsx",
        company_name="LLOYD TUNISIE"
    )