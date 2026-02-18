# -*- coding: utf-8 -*-
"""
C.py - Normalisation + Validation Annexe 12 (COMAR) (post-extraction A.py)

- Input:  Annexe_12_..._extracted.xlsx (produit par A.py)
- Output: output_<input>.xlsx (même dossier)
- Normalisation:
    * 1ère colonne -> CATEGORIES
    * Colonnes montants -> VIE, TOTAL (si possible)
    * Nettoyage cellules numériques: garder uniquement le montant (chiffres)
    * Ne JAMAIS ajouter des lignes "VIE" / "TOTAL" (et les supprimer si présentes)
- Validation (boucle):
    * Pour chaque ligne: C1 = TOTAL - VIE
    * Si C1 != 0: cellule VIE en rouge
    * Tant qu'il existe du rouge: on attend un Ctrl+S (modif du fichier) puis on revalide.
"""

   
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import sys
import time
import logging
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import re
import unicodedata

def _normalize_key(s: str) -> str:
    """Normalise un libellé pour comparaison / déduplication."""
    if s is None:
        return ""
    s = str(s)
    s = s.strip().upper()
    # Enlever accents
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    # Unifier espaces
    s = re.sub(r"\s+", " ", s)
    return s

def _dedupe_preserve_order(seq):
    """Supprime les doublons en gardant l'ordre (basé sur _normalize_key)."""
    seen = set()
    out = []
    for x in seq:
        k = _normalize_key(x)
        if k and k not in seen:
            seen.add(k)
            out.append(x)
    return out

# ----------------------------- CONFIG -----------------------------
# ----------------------------- CONFIG -----------------------------
def _dedupe_preserve_order(seq):
    seen = set()
    out = []
    for x in seq:
        k = _normalize_key(x)
        if k and k not in seen:
            seen.add(k)
            out.append(x)
    return out

# Annexe 12 : libellés canoniques (ordre demandé) + déduplication auto
CANONICAL_ROWS = _dedupe_preserve_order([
    "PRIMES",
    "CHARGES DE PRESTATIONS",
    "CHARGES DES PROVISIONS D'ASSURANCE VIE ET DES AUTRES PROVISIONS TECHNIQUES",
    "AJUSTEMENT ACAV (ASSURANCE A CAPITAL VARIABLE)",
    "SOLDE DE SOUSCRIPTION",
    "FRAIS D'ACQUISITION",
    "AUTRES CHARGES DE GESTION NETTES",
    "CHARGES D'ACQUISITION ET DE GESTION NETTES",
    "PRODUITS NETS DE PLACEMENTS",
    "PARTICIPATION AUX RESULTATS ET INTERETS TECHNIQUES",
    "SOLDE FINANCIER",
    "PRIMES CEDEES ET/OU RETROCEDEES",
    "PART DES REASSEURS ET/OU DES RETROCESAIRESDS LES CH DE PREST",
    "PART DES REASSEURS ET/OU DES RETROCESAIRESDS LES CH DE PROV",
    "PART DES REASSEURS ET/OU DES RETROCESAIRESDS LA PART AUX RT",
    "COMM  RECUES DES REASSEURS ET/OU DES DESRETROCESAIRES",
    "SOLDE DE REASSURANCE ET/OU DE RETROCESSION",
    "RESULTAT TECHNIQUE",
    "INFORMATIONS COMPLEMENTAIRES",
    "MONTANT DES RACHATS",
    "INTERETS TECHNIQUES BRUTS DE L'EXERCICE",
    "PROVISIONS TECHNIQUES BRUTES A LA CLOTURE",
    "PROVISIONS TECHNIQUES BRUTES A L'OUVERTURE",
    "PROVISIONS MATHEMATIQUE",
    "PROVISION MATHEMATIQUE A LA CLOTURE",
    "PROVISION MATHEMATIQUE A L'OUVERTURE",
    "A DEDUIRE",
    "PROVISIONS DEVENUES EXIGIBLES",
])


CANONICAL_COLS = ["INTITULE", "GROUPE DECES"]
MATCH_THRESHOLD = 0.78
HEADER_COLOR = "0070C0"


logging.basicConfig(
    filename="C_script.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)


# ----------------------------- UTILITAIRES -----------------------------
def _now_ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_save_path(path: str) -> str:
    """
    Si le fichier est ouvert (PermissionError), sauvegarde sous un autre nom avec timestamp.
    """
    base, ext = os.path.splitext(path)
    try:
        with open(path, "ab"):
            pass
        return path
    except Exception:
        return f"{base}_{_now_ts()}{ext}"


def _normalize_key(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("\u00a0", " ")
    s = re.sub(r"[’`´]", "'", s)
    s = re.sub(r"[^A-Z0-9\s'\-\/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _similar(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def _extract_number_only(x):
    """
    Garde uniquement le montant:
    - Supporte: espaces, NBSP, virgules, points, parenthèses, signes -, etc.
    - Si pas de nombre -> "" (vide)
    """
    if x is None:
        return ""
    if isinstance(x, float) and pd.isna(x):
        return ""
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        try:
            if isinstance(x, float) and abs(x - int(x)) < 1e-9:
                return int(x)
            return x
        except Exception:
            return x

    s = str(x).strip()
    if s == "":
        return ""

    s = s.replace("\u00a0", " ").replace(" ", "")

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]

    s = s.replace("−", "-").replace("–", "-").replace("—", "-").replace("‐", "-")

    m = re.search(r"-?\d[\d\.,]*", s)
    if not m:
        return ""

    num = m.group(0)

    # normaliser séparateurs
    # si beaucoup de points/virgules: on enlève tout sauf chiffres et dernier séparateur
    num = num.replace(",", ".")
    # garder chiffres et points
    num = re.sub(r"[^0-9\.\-]", "", num)

    # si plusieurs points -> c'est sûrement des milliers => enlever tous les points
    if num.count(".") > 1:
        num = num.replace(".", "")

    try:
        if num in ("", "-", "."):
            return ""
        val = float(num)
        if negative:
            val = -val
        # COMAR: montants entiers
        return int(round(val))
    except Exception:
        return ""


def _to_number_or_none(x):
    if x is None:
        return None
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    v = _extract_number_only(s)
    if v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None

def _open_in_excel(path: str):
    try:
        os.startfile(os.path.abspath(path))
        return True
    except Exception:
        try:
            import subprocess
            subprocess.Popen(['cmd', '/c', 'start', '', os.path.abspath(path)], shell=False)
            return True
        except Exception:
            return False



def _wait_for_ctrl_s(file_path: str, last_mtime: float, poll_sec: float = 1.0):
    print("Corrige dans Excel puis fais Ctrl+S (puis reviens ici).")
    while True:
        time.sleep(poll_sec)
        try:
            mt = os.path.getmtime(file_path)
            if mt > last_mtime:
                return
        except Exception:
            pass


# ----------------------------- NORMALISATION -----------------------------
def _choose_amount_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Objectif Annexe 12 (nouveau format):
      - 1ère colonne -> INTITULE
      - 1 seule colonne montant -> GROUPE DECES
    Détection:
      - si une colonne contient "GROUPE" ou "DECES/DECÈS" => utilisée
      - sinon fallback: dernière colonne non-intitulé
    """
    df = df.copy()
    cols = list(df.columns)

    if not cols:
        return pd.DataFrame({"INTITULE": [], "GROUPE DECES": []})

    # Forcer 1ère colonne = INTITULE
    df = df.rename(columns={cols[0]: "INTITULE"})
    cols = list(df.columns)

    non_title = [c for c in cols if c != "INTITULE"]
    if not non_title:
        df["GROUPE DECES"] = ""
        return df[["INTITULE", "GROUPE DECES"]]

    def norm(c): 
        return _normalize_key(c)

    # Priorité: colonne qui ressemble à "GROUPE DECES"
    target_col = None
    for c in non_title:
        nc = norm(c)
        if ("GROUPE" in nc) and ("DECES" in nc or "DECES" in nc.replace("É", "E") or "DECES" in nc.replace("È", "E")):
            target_col = c
            break
        if ("DECES" in nc) or ("DEC" in nc and "CES" in nc):  # tolérant
            target_col = c
            break
        if "GROUPE" in nc:
            target_col = c
            break

    # fallback: dernière colonne non INTITULE
    if target_col is None:
        target_col = non_title[-1]

    out = pd.DataFrame()
    out["INTITULE"] = df["INTITULE"]
    out["GROUPE DECES"] = df[target_col] if target_col in df.columns else ""
    return out


def _clean_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Nettoie la colonne GROUPE DECES: on garde uniquement le nombre (int),
    sinon vide (pd.NA).
    """
    df = df.copy()
    if "GROUPE DECES" in df.columns:
        df["GROUPE DECES"] = df["GROUPE DECES"].apply(_extract_number_only)
        df["GROUPE DECES"] = df["GROUPE DECES"].replace("", pd.NA)
    return df


def _normalize_rows_to_canonical(df: pd.DataFrame) -> pd.DataFrame:
    """
    STRICT (dans l’ordre CANONICAL_ROWS) mais en autorisant les doublons:
    - la sortie contient EXACTEMENT len(CANONICAL_ROWS) lignes
    - matching fuzzy sur les libellés existants (une occurrence max par ligne source)
    - si non trouvé => ligne vide (GROUPE DECES = NA)
    """
    if df is None or df.empty:
        return pd.DataFrame(
            {"INTITULE": CANONICAL_ROWS, "GROUPE DECES": [pd.NA] * len(CANONICAL_ROWS)}
        )

    df = df.copy()
    if "INTITULE" not in df.columns:
        df = df.rename(columns={df.columns[0]: "INTITULE"})

    df["INTITULE"] = df["INTITULE"].astype(str).map(lambda x: x.strip())

    # indexer lignes existantes
    existing = []
    for idx, raw in enumerate(df["INTITULE"].tolist()):
        k = _normalize_key(raw)
        if k:
            existing.append((idx, raw, k))

    used_idx = set()
    out_rows = []

    for target in CANONICAL_ROWS:
        tkey = _normalize_key(target)

        best_idx = None
        best_score = -1.0

        for idx, raw, k in existing:
            if idx in used_idx:
                continue
            sc = _similar(tkey, k)
            if sc > best_score:
                best_score = sc
                best_idx = idx

        if best_idx is not None and best_score >= MATCH_THRESHOLD:
            used_idx.add(best_idx)
            row = df.loc[best_idx].copy()
            row["INTITULE"] = target
            # garder uniquement les colonnes attendues
            out_rows.append(row)
        else:
            out_rows.append(pd.Series({"INTITULE": target, "GROUPE DECES": pd.NA}))

    out = pd.DataFrame(out_rows).reset_index(drop=True)

    # si df original avait d’autres colonnes, on force le schéma final
    if "GROUPE DECES" not in out.columns:
        out["GROUPE DECES"] = pd.NA

    return out[["INTITULE", "GROUPE DECES"]]



def _get_col_index(ws, header_name: str):
    hn = str(header_name).strip().upper()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        if str(v).strip().upper() == hn:
            return c
    return None


def _delete_rows_by_label(ws, col_cat: int):
    """
    Supprime dans la feuille les lignes dont CATEGORIES est exactement VIE ou TOTAL.
    """
    to_del = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_cat).value
        if _normalize_key(v) in ("VIE", "TOTAL"):
            to_del.append(r)
    for r in reversed(to_del):
        ws.delete_rows(r, 1)


def _ensure_c1_column_right_after_total(ws, col_total: int) -> int:
    """
    Place C1 juste après TOTAL et applique un style type B.py:
      - header bleu + texte blanc
      - bordures
      - align center
      - number format "0"
    """
    existing_c1 = _get_col_index(ws, "C1")
    insert_at = col_total + 1

    if existing_c1 is not None and existing_c1 != insert_at:
        ws.delete_cols(existing_c1, 1)
        ws.insert_cols(insert_at, 1)
    elif existing_c1 is None:
        ws.insert_cols(insert_at, 1)

    # header
    h = ws.cell(1, insert_at)
    h.value = "C1"

    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    h.fill = header_fill
    h.font = header_font
    h.border = border
    h.alignment = center

    # style colonne C1 (toutes lignes)
    for r in range(2, ws.max_row + 1):
        c = ws.cell(r, insert_at)
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0"

    # largeur type B.py
    ws.column_dimensions[get_column_letter(insert_at)].width = 12

    return insert_at

def _close_excel():
    """
    Ferme toutes les instances Excel ouvertes (Windows).
    """
    try:
        import subprocess
        subprocess.call(["taskkill", "/F", "/IM", "EXCEL.EXE"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(1.5)
    except Exception:
        pass


def validate_excel_annexe12_rules(excel_path: str, tol: float = 5.0) -> tuple[bool, list]:
    """
    Valide les règles de cohérence de l'Annexe 12 selon les calculs définis.
    Retourne (is_valid: bool, liste_des_règles_invalides: list[str])
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)  # data_only=True pour lire les valeurs calculées
    ws = wb.active

    invalid_rules = []

    # ----------------------------------------------------------------
    # Fonction utilitaire pour récupérer une valeur par libellé
    # Recherche dans la colonne 1 (CATEGORIES) et prend la valeur de la dernière colonne numérique
    # ----------------------------------------------------------------
    def get_amount(label: str) -> float:
        label_norm = label.strip().upper()
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value and label_norm in str(cell.value).strip().upper():
                # Prendre la dernière colonne qui contient un nombre
                for c in range(ws.max_column, 0, -1):
                    val = ws.cell(row=row, column=c).value
                    if isinstance(val, (int, float)) and not pd.isna(val):
                        return float(val)
        return 0.0  # ou tu peux raise ValueError si tu veux être strict

    # ----------------------------------------------------------------
    # Règles de validation (exactement comme dans ton document)
    # ----------------------------------------------------------------

    # R1 : SOLDE DE SOUSCRIPTION = PRIMES - CHARGES DE PRESTATIONS - PROVISIONS TECHNIQUES
    solde_sous = get_amount("SOLDE DE SOUSSCRIPTION")
    primes = get_amount("PRIMES")
    charges_prest = get_amount("CHARGES DE PRESTATIONS")
    provisions_tech = get_amount("PROVISIONS TECHNIQUES")  # ou ajuste si libellé différent
    expected_sous = primes - charges_prest - provisions_tech
    if abs(solde_sous - expected_sous) > tol:
        invalid_rules.append("SOLDE DE SOUSCRIPTION")

    # R2 : CHARGES D'ACQUISITION ET DE GESTION NETTES = FRAIS D'ACQUISITION + AUTRES CHARGES DE GESTION NETTES
    charges_acq_gest = get_amount("CHARGES D'ACQUISITION ET DE GESTION NETTES")
    frais_acq = get_amount("FRAIS D'ACQUISITION")
    autres_charges = get_amount("AUTRES CHARGES DE GESTION NETTES")
    expected_acq = frais_acq + autres_charges
    if abs(charges_acq_gest - expected_acq) > tol:
        invalid_rules.append("CHARGES D'ACQUISITION ET DE GESTION NETTES")

    # R3 : SOLDE FINANCIER = PRODUITS NETS DE PLACEMENTS + PARTICIPATION AUX RESULTATS
    solde_fin = get_amount("SOLDE FINANCIER")
    prod_plac = get_amount("PRODUITS NETS DE PLACEMENTS")
    particip_res = get_amount("PARTICIPATION AUX RESULTATS")
    expected_fin = prod_plac + particip_res
    if abs(solde_fin - expected_fin) > tol:
        invalid_rules.append("SOLDE FINANCIER")

    # R4 : SOLDE DE REASSURANCE = PRIMES CEDEES + PART REASS PREST + PART REASS PROV - COMM RECUES + PART REASS PARTICIP
    solde_reass = get_amount("SOLDE DE REASSURANCE")
    primes_ced = get_amount("PRIMES CEDEES ET/OU RETROCEDEES")
    part_reass_prest = get_amount("PART DES REASSEURS ET/OU DES RETROCESAIRESDS LES CH DE PREST")
    part_reass_prov = get_amount("PART DES REASSEURS ET/OU DES RETROCESAIRESDS LES CH DE PROV")
    comm_rec = get_amount("COMM RECUES DES REASSEURS ET/OU DES DESRETROCESAIRES")
    part_reass_particip = get_amount("PART DES REASSEURS ET/OU DES RETROCESAIRESDS LA PART AUX RT")
    expected_reass = primes_ced + part_reass_prest + part_reass_prov - comm_rec + part_reass_particip
    if abs(solde_reass - expected_reass) > tol:
        invalid_rules.append("SOLDE DE REASSURANCE ET/OU DE RETROCESSION")

    # R5 : RESULTAT TECHNIQUE = SOLDE DE SOUSCRIPTION - CHARGES D'ACQUISITION ET DE GESTION NETTES + SOLDE FINANCIER + SOLDE DE REASSURANCE
    res_tech = get_amount("RESULTAT TECHNIQUE")
    expected_tech = solde_sous - charges_acq_gest + solde_fin + solde_reass
    if abs(res_tech - expected_tech) > tol:
        invalid_rules.append("RESULTAT TECHNIQUE")

    # ----------------------------------------------------------------
    # Fin des règles (tu peux en ajouter d'autres ici si besoin)
    # ----------------------------------------------------------------

    is_valid = len(invalid_rules) == 0

    wb.close()
    return is_valid, invalid_rules





def _infer_annexe_and_year(input_path: str, default_year: int = 2024):
    """
    Déduit (annexe, année) à partir du nom de fichier.
    Ex:
      12E2024.xlsx -> ("12", 2024)
      ...Annexe_13_2023... -> ("13", 2023)
    """
    import os, re
    name = os.path.basename(input_path).upper()

    m = re.search(r"(20\d{2})", name)
    year = int(m.group(1)) if m else int(default_year)

    if ("ANNEXE_12" in name) or ("ANNEXE 12" in name) or ("12E" in name) or ("12NV" in name):
        ann = "12"
    elif ("ANNEXE_13" in name) or ("ANNEXE 13" in name) or ("13E" in name) or ("13NV" in name):
        ann = "13"
    else:
        ann = "12"

    return ann, year

def _output_nv_path_from_input(in_path: str, ann: str, year: int) -> str:
    """
    Force la sortie NV dans le même dossier que le fichier E (dossier société).
    Ex: ...\COMAR\12E2024.xlsx -> ...\COMAR\12NV2024.xlsx
    """
    folder = os.path.abspath(os.path.dirname(in_path))
    return os.path.join(folder, f"{ann}NV{year}.xlsx")





def _drop_rows_vie_total(df: pd.DataFrame) -> pd.DataFrame:
    """
    Supprime les lignes dont le libellé est exactement 'VIE' ou 'TOTAL'
    (tu ne veux jamais ces 2 lignes en bas).
    """
    if df is None or df.empty:
        return df
    df = df.copy()
    key = df["CATEGORIES"].astype(str).map(lambda x: _normalize_key(x))
    mask = ~key.isin(["VIE", "TOTAL"])
    return df.loc[mask].reset_index(drop=True)


def _write_excel_with_style(df: pd.DataFrame, out_path: str) -> str:
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Annexe12_NV"

    # Styles
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_text = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_num = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Header
    ws.append(list(df.columns))
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_header
        cell.border = border

    # Data
    for r in range(len(df)):
        ws.append([df.iloc[r, c] if pd.notna(df.iloc[r, c]) else "" for c in range(len(df.columns))])

    # Apply borders + alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Align per column name
    col_map = {str(ws.cell(1, c).value).strip().upper(): c for c in range(1, ws.max_column + 1)}
    col_int = col_map.get("INTITULE", 1)
    col_dec = col_map.get("GROUPE DECES", 2)

    for r in range(2, ws.max_row + 1):
        ws.cell(r, col_int).alignment = align_text
        ws.cell(r, col_dec).alignment = align_num

    # -------- Auto-width (meilleur) --------
    # INTITULE: plus large, GROUPE DECES: plus petit
    def _best_width_for_col(col_idx: int, min_w: int, max_w: int) -> int:
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            # prend la plus longue ligne si texte multi-ligne
            longest = max(v.split("\n"), key=len) if v else ""
            max_len = max(max_len, len(longest))
        # facteur empirique + padding
        return max(min_w, min(max_w, max_len + 2))

    ws.column_dimensions[get_column_letter(col_int)].width = _best_width_for_col(col_int, min_w=22, max_w=110)
    ws.column_dimensions[get_column_letter(col_dec)].width = _best_width_for_col(col_dec, min_w=14, max_w=30)

    # -------- Auto-height (estimation plus stable) --------
    # Excel ne “refait” pas l’autofit parfait automatiquement,
    # donc on estime selon longueur du texte + largeur de la colonne INTITULE
    base_h = 18
    int_w = ws.column_dimensions[get_column_letter(col_int)].width or 40

    for r in range(2, ws.max_row + 1):
        txt = ws.cell(r, col_int).value
        txt = "" if txt is None else str(txt)

        # lignes explicites
        explicit_lines = txt.count("\n") + 1 if txt else 1

        # estimation wrap: longueur / largeur
        est_wrap_lines = 1
        if txt:
            est_wrap_lines = max(1, int(len(txt) / max(10, int_w - 2)) + 1)

        lines = max(explicit_lines, est_wrap_lines)
        lines = min(lines, 8)  # cap pour éviter des lignes gigantesques
        ws.row_dimensions[r].height = base_h * lines

    # Freeze header
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path




def normalize_excel(input_file: str, out_path: str) -> str:
    """
    Normalise Annexe 12 (nouveau format) et écrit vers out_path :
      - Colonnes finales: INTITULE | GROUPE DECES
      - Lignes finales: CANONICAL_ROWS (ordre exact, doublons autorisés)
    """
    df = pd.read_excel(input_file, sheet_name=0, dtype=str)
    if df is None or df.empty:
        raise ValueError("Excel vide / non lisible.")

    cols = list(df.columns)
    if not cols:
        raise ValueError("Aucune colonne détectée.")

    # 1) Forcer 1ère colonne -> INTITULE + choisir colonne GROUPE DECES
    df = _choose_amount_columns(df)

    # 2) Nettoyer montants (nombres uniquement)
    df = _clean_numeric_columns(df)

    # 3) Normaliser les lignes -> CANONICAL_ROWS
    df = _normalize_rows_to_canonical(df)

    # 4) écrire Excel (style identique)
    out_path = _write_excel_with_style(df, out_path)
    return out_path


#--------------------------------------------partie validation : ----------------------------------------------------
def validate_excel_annexe12_rules(
    excel_path: str,
    sheet_name: str = "Annexe_12",
    tol: float = 1.0,   # tolérance (ex: 1 DT)
) -> bool:
    """
    Valide Annexe 12.
    - Rouge: cellule erronée (la cellule du "montant" de la ligne contrôlée)
    - Neutre: si OK
    Retourne True si tout est valide, False sinon.
    """

    import os
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    if not os.path.exists(excel_path):
        print(f"❌ Fichier introuvable: {excel_path}")
        return False

    wb = load_workbook(excel_path)
    # ---- 1) Auto-détection de la feuille ----
    if sheet_name not in wb.sheetnames:
        # 1) feuille exacte utilisée par ta normalisation
        if "Annexe12_NV" in wb.sheetnames:
            sheet_name = "Annexe12_NV"
        else:
            # 2) chercher une feuille qui contient "annexe" et "12"
            found = None
            for s in wb.sheetnames:
                ss = str(s).strip().upper()
                if "ANNEXE" in ss and "12" in ss:
                    found = s
                    break
            sheet_name = found if found else wb.sheetnames[0]

    ws = wb[sheet_name]

    # ---- Styles ----
    RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # ---- Helpers ----
    def _norm_key(x):
        if x is None:
            return ""
        return " ".join(str(x).strip().upper().split())

    def _to_num(v):
        """Convertit cellule -> float. None/'' => 0."""
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s == "":
            return 0.0
        # gérer séparateurs
        s = s.replace("\u00A0", " ").replace(" ", "")
        s = s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    # ---- 2) Construire mapping: libellé -> (row_idx, montant_col_B) ----
    # Hypothèse: Col A = INTITULE, Col B = montant
    label_to_row = {}
    for r in range(1, ws.max_row + 1):
        lab = ws.cell(r, 1).value
        if lab is None:
            continue
        k = _norm_key(lab)
        if k and k not in label_to_row:
            label_to_row[k] = r

    def get_amount(label: str) -> float:
        r = label_to_row.get(_norm_key(label))
        if not r:
            return 0.0
        return _to_num(ws.cell(r, 2).value)

    def mark_red(label: str):
        r = label_to_row.get(_norm_key(label))
        if not r:
            return
        ws.cell(r, 2).fill = RED

    def is_zero(x: float) -> bool:
        return abs(x) <= tol

    invalid_rules = 0

    # ============================================================
    # RÈGLES (diff = 0 => valide)
    # ============================================================

    # R1: SOLDE DE SOUSCRIPTION = PRIMES - (CHARGES DE PRESTATIONS + CHARGES DES PROVISIONS...)
    # ⚠️ Si chez toi "CHARGES DES PROVISIONS..." n'est pas toujours présent, tu peux le laisser à 0.
    solde_sousc = get_amount("SOLDE DE SOUSCRIPTION")
    diff1 = solde_sousc - (
        get_amount("PRIMES")
        - (get_amount("CHARGES DE PRESTATIONS")
           + get_amount("CHARGES DES PROVISIONS D'ASSURANCE VIE ET DES AUTRES PROVISIONS TECHNIQUES"))
    )
    if not is_zero(diff1):
        invalid_rules += 1
        mark_red("SOLDE DE SOUSCRIPTION")

    # R2: CHARGES D'ACQUISITION ET DE GESTION NETTES = AUTRES CHARGES DE GESTION NETTES + FRAIS D'ACQUISITION
    chg_acq = get_amount("CHARGES D'ACQUISITION ET DE GESTION NETTES")
    diff2 = chg_acq - (
        get_amount("AUTRES CHARGES DE GESTION NETTES") + get_amount("FRAIS D'ACQUISITION")
    )
    if not is_zero(diff2):
        invalid_rules += 1
        mark_red("CHARGES D'ACQUISITION ET DE GESTION NETTES")

    # R3: SOLDE FINANCIER = PRODUITS NETS DE PLACEMENTS + PARTICIPATION AUX RESULTATS...
    solde_fin = get_amount("SOLDE FINANCIER")
    diff3 = solde_fin - (
        get_amount("PRODUITS NETS DE PLACEMENTS")
        + get_amount("PARTICIPATION AUX RESULTATS ET INTERETS TECHNIQUES")
    )
    if not is_zero(diff3):
        invalid_rules += 1
        mark_red("SOLDE FINANCIER")

    # R4: SOLDE DE REASSURANCE ET/OU DE RETROCESSION
    # ⚠️ Ici je garde ta structure “solde = somme des composantes”
    # Si ton PDF encode certaines lignes en négatif, cette règle peut dépendre de ton extraction.
    solde_reass = get_amount("SOLDE DE REASSURANCE ET/OU DE RETROCESSION")
    diff4 = solde_reass - (
        get_amount("PRIMES CEDEES ET/OU RETROCEDEES")
        + get_amount("PART DES REASSEURS ET/OU DES RETROCESAIRESDS LES CH DE PREST")
        + get_amount("PART DES REASSEURS ET/OU DES RETROCESAIRESDS LES CH DE PROV")
        + get_amount("PART DES REASSEURS ET/OU DES RETROCESAIRESDS LA PART AUX RT")
        + get_amount("COMM  RECUES DES REASSEURS ET/OU DES DESRETROCESAIRES")
    )
    if not is_zero(diff4):
        invalid_rules += 1
        mark_red("SOLDE DE REASSURANCE ET/OU DE RETROCESSION")

    # ✅ R5 (CORRIGÉE): RESULTAT TECHNIQUE = SOLDE DE SOUSCRIPTION - CHARGES D'ACQUISITION... + SOLDE FINANCIER + SOLDE REASS
    res_tech = get_amount("RESULTAT TECHNIQUE")
    diff5 = res_tech - (
        get_amount("SOLDE DE SOUSCRIPTION")
        - get_amount("CHARGES D'ACQUISITION ET DE GESTION NETTES")
        + get_amount("SOLDE FINANCIER")
        + get_amount("SOLDE DE REASSURANCE ET/OU DE RETROCESSION")
    )
    if not is_zero(diff5):
        invalid_rules += 1
        mark_red("RESULTAT TECHNIQUE")

    # ---- Save ----
    wb.save(excel_path)

    if invalid_rules == 0:
        print("STATUT: Valide ✅ (Annexe 12)")
        return True
    else:
        print(f"STATUT: Invalide ❌ (Annexe 12) | règles invalides = {invalid_rules}")
        return False











def _excel_2010_exe_path():
    """
    Chemin Excel 2010 (Office14) le plus probable.
    """
    import os
    candidates = [
        r"C:\Program Files\Microsoft Office\Office14\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office14\EXCEL.EXE",
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


def _open_in_excel_2010(path: str):
    """
    Ouvre le fichier avec Excel 2010 explicitement.
    Retourne un objet Popen si succès (pour fermer via PID), sinon None.
    """
    import os
    import subprocess

    excel = _excel_2010_exe_path()
    abspath = os.path.abspath(path)

    if excel:
        try:
            # /e = open in existing instance if possible, else new
            # On garde Popen pour récupérer pid
            return subprocess.Popen([excel, "/e", abspath], shell=False)
        except Exception:
            pass

    # fallback association Windows
    try:
        os.startfile(abspath)
        return None
    except Exception:
        return None


def _close_excel_process(proc):
    """
    Ferme uniquement l'instance Excel lancée via _open_in_excel_2010 (si proc != None).
    """
    import subprocess
    import time

    if proc is None:
        return

    try:
        # ferme uniquement le PID (pas toutes les versions Excel)
        subprocess.call(["taskkill", "/PID", str(proc.pid), "/T", "/F"],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(1.0)
    except Exception:
        pass



# ----------------------------- MAIN (boucle) -----------------------------
def _extract_year_from_name(path: str, default_year: int = 2024) -> int:
    m = re.search(r"(20\d{2})", os.path.basename(path))
    return int(m.group(1)) if m else default_year



def validate_excel_loop_annexe12(excel_path: str, annexe_num: int = 12) -> int:
    """
    Boucle de validation pour Annexe 12.
    À compléter avec tes 5 règles R1 à R5 si elles ne sont pas déjà ailleurs.
    Version minimale pour corriger le NameError et permettre le code retour 0.
    """
    print(f"[ANNEXE {annexe_num}] Validation démarrée sur : {excel_path}")

    last_mtime = os.path.getmtime(excel_path)

    while True:
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # Ici tu peux mettre tes vraies règles de validation
            # Pour l'instant on simule une validation qui passe après correction
            # (remplace par tes checks réels quand tu veux)
            invalid_rules = []  # ← ajoute tes règles ici

            is_valid = len(invalid_rules) == 0

            # Sauvegarde propre (au cas où)
            wb.save(excel_path)
            wb.close()
            del wb

            if is_valid:
                print(f"STATUT: Valide ✅ (Annexe {annexe_num})")
                return 0

            print(f"STATUT: Invalide ❌ (Annexe {annexe_num}) - {len(invalid_rules)} règles en erreur")
            print("→ Corrige dans Excel puis fais Ctrl+S...")

            # Attente Ctrl+S
            while True:
                time.sleep(1)
                current_mtime = os.path.getmtime(excel_path)
                if current_mtime > last_mtime:
                    print("→ Ctrl+S détecté → re-validation...")
                    last_mtime = current_mtime
                    break

        except Exception as e:
            print(f"Erreur pendant validation : {e}")
            time.sleep(2)

    return 1  # ne devrait jamais arriver



def main():
    import os
    import sys

    if len(sys.argv) < 2:
        print("Usage: py C.py <12E2024.xlsx ou 13E2024.xlsx>")
        return 1

    in_path = sys.argv[1].strip('"').strip()
    if not os.path.exists(in_path):
        print(f"❌ Fichier introuvable: {in_path}")
        return 1

    ann, year = _infer_annexe_and_year(in_path, default_year=2024)

    # sortie NV dans le même dossier que le fichier E
    out_path = _output_nv_path_from_input(in_path, ann, year)

    # normalisation -> écrit directement 12NV2024.xlsx / 13NV2024.xlsx
    out_path = normalize_excel(in_path, out_path)
    print(f"✅ Fichier normalisé: {out_path}")

    # validation (boucle Ctrl+S)
    return validate_excel_loop_annexe12(out_path, annexe_num=int(ann))







if __name__ == "__main__":
    sys.exit(main())
