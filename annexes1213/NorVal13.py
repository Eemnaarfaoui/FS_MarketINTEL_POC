# B.py
# ------------------------------------------------------------
# Annexe 13 (COMAR) - Normalisation + validation UNIQUEMENT via :
#   C1 = TOTAL - somme(des autres colonnes)
# - Conserve le style du tableau extrait (openpyxl, modifications "in place")
# - Ajoute une colonne C1 juste après TOTAL (si absente)
# - Met en ORANGE la cellule C1 si la ligne est invalide
# - Met en ROUGE la cellule la plus "contributrice" (|valeur| max) sur la ligne invalide
# - Boucle Excel: ouvre le fichier, attend Ctrl+S (via modification mtime), ferme Excel, revalide
# - Option: forcer Excel 2010 via EXCEL2010_PATH (si tu veux)
# ------------------------------------------------------------

import os
import re
import sys
import time
import copy
import subprocess
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================

EXPECTED_COLUMNS = [
    "CATEGORIES", "ACCEPTATION", "ACC R.D", "AUTO", "ACC.TRAV",
    "INCENDIE", "TRANSPORT", "GRELE", "GROUPE", "TOTAL"
]

EXPECTED_ROWS = [
    "PRIMES EMISES", 
    "VARIATION DES PRIMES NON ACQUISES",
    "CHARGES DE PRESTATIONS",
    "PRESTATIONS ET FRAIS PAYES",
    "CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE",
    "SOLDE DE SOUSCRIPTION",
    "FRAIS D'ACQUISITION",
    "AUTRES CHARGES DE GESTION NETTES",
    "CHARGES D'ACQUISITION ET DE GESTION NETTES",
    "PRODUITS NETS DE PLACEMENTS",
    "PARTICIPATION AUX RESULTATS",
    "SOLDE FINANCIER",
    "PRIMES CEDEES AUX REASSUREURS",
    "PART REASSUREURS DANS LES PRIMES ACQUISES",
    "PART REASSUREURS DANS LES PRESTATIONS PAYEES",
    "PART DES REAS ET/OU DES RETROC DANS LES CHARGES DE PROV POUR PRESTATION",
    "PART REASSUREURS DANS LA PARTICIPATION AUX RESULTATS",
    "COMMISSIONS RECUES DES REASSUREURS",
    "SOLDE DE REASSURANCE ET/OU DE RETROCESSION",
    "RESULTAT TECHNIQUE NON VIE",
    "INFORMATIONS COMPLEMENTAIRES",
    "PROVISIONS POUR PRIMES NON ACQUISES - ANNEE N",
    "PROVISIONS POUR PRIMES NON ACQUISES - ANNEE N",
    "PROVISIONS POUR SINSITRES A PAYER - ANNEE N",
    "PROVISIONS POUR SINSITRES A PAYER - ANNEE N",
    "AUTRES PROVISIONS TECHNIQUES CLOTURE",
    "AUTRES PROVISIONS TECHNIQUES OUVERTURE",
    "A DEDUIRE"
]

# Tolérance sur C1 (différences de somme)
TOL = 5.0  # ajuste si besoin

# Couleurs (fill)
FILL_RED = PatternFill("solid", fgColor="FF0000")
FILL_ORANGE = PatternFill("solid", fgColor="FFA500")
FILL_GREEN = PatternFill("solid", fgColor="00B050")  # vert Excel
FILL_NONE = PatternFill(fill_type=None)

# Excel 2010 (optionnel) : mets ici le chemin EXACT si tu veux forcer Office 2010
# Exemple:
# EXCEL2010_PATH = r"C:\Program Files (x86)\Microsoft Office\Office14\EXCEL.EXE"
EXCEL2010_PATH = None


# =========================
# UTILS: texte / matching
# =========================

def _norm_key(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().upper()
    # accents simples
    s = (
        s.replace("É", "E").replace("È", "E").replace("Ê", "E")
        .replace("À", "A").replace("Â", "A")
        .replace("Î", "I").replace("Ï", "I")
        .replace("Ô", "O")
        .replace("Ù", "U")
        .replace("Ç", "C")
        .replace("’", "'").replace("`", "'")
    )
    # espaces multiples
    s = re.sub(r"\s+", " ", s)
    # nettoyage ponctuation non utile
    s = re.sub(r"[•\u2010\u2011\u2012\u2013\u2014\u2212]", "-", s)  # tirets
    return s


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio() * 100.0


def best_match(value: str, expected: List[str], min_score: float = 78.0) -> Tuple[Optional[str], float]:
    k = _norm_key(value)
    best = None
    best_score = -1.0
    for e in expected:
        sc = _similarity(k, _norm_key(e))
        if sc > best_score:
            best_score = sc
            best = e
    if best_score >= min_score:
        return best, best_score
    return None, best_score


# =========================
# UTILS: nombres
# =========================

_NUM_RE = re.compile(r"[-+]?\d[\d\s.,]*")

def parse_number(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None

    # extraire la première séquence num-like
    m = _NUM_RE.search(s.replace("\u00A0", " "))
    if not m:
        return None
    token = m.group(0)

    # enlever espaces
    token = token.replace(" ", "")

    # gérer virgule/point
    # si contient , et . => on suppose . = milliers, , = décimal (ou l'inverse) : on prend dernière occurrence comme décimal
    if "," in token and "." in token:
        if token.rfind(",") > token.rfind("."):
            token = token.replace(".", "")
            token = token.replace(",", ".")
        else:
            token = token.replace(",", "")
    else:
        # si juste virgule -> décimal
        if "," in token and "." not in token:
            token = token.replace(",", ".")
        # si juste point -> décimal
        # sinon ok

    try:
        return float(token)
    except:
        return None


# =========================
# UTILS: styles (évite StyleProxy unhashable)
# =========================

import copy

def copy_cell_style(dst, src):
    """
    Copie le style d'une cellule en cassant les StyleProxy (sinon erreurs unhashable / immutable).
    """
    dst._style = copy.copy(src._style)
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy.copy(src.protection)
    dst.comment = src.comment



def ensure_c1_column_inplace(ws, header_row: int = 1):
    # trouve TOTAL
    headers = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        headers[_norm_key(v)] = c

    if "C1" in headers:
        return  # déjà là

    total_col = headers.get("TOTAL")
    if not total_col:
        # tente "Total"
        total_col = headers.get("TOTAL ") or headers.get("TOTAL:")
    if not total_col:
        raise ValueError("Colonne TOTAL introuvable (ligne d'entête).")

    insert_at = total_col + 1
    ws.insert_cols(insert_at)

    # header C1 en copiant style de TOTAL header
    src_header = ws.cell(row=header_row, column=total_col)
    dst_header = ws.cell(row=header_row, column=insert_at)
    dst_header.value = "C1"
    copy_cell_style(dst_header, src_header)

    # copier largeur
    total_letter = get_column_letter(total_col)
    c1_letter = get_column_letter(insert_at)
    ws.column_dimensions[c1_letter].width = ws.column_dimensions[total_letter].width

    # copier styles des cellules de la colonne TOTAL vers C1 (même format)
    for r in range(header_row + 1, ws.max_row + 1):
        src = ws.cell(row=r, column=total_col)
        dst = ws.cell(row=r, column=insert_at)
        copy_cell_style(dst, src)


# =========================
# Normalisation colonnes/lignes (sans casser le style)
# =========================

def normalize_columns_inplace(ws, header_row: int = 1):
    """
    Normalise les noms de colonnes pour correspondre à EXPECTED_COLUMNS.
    Règles fixes demandées :
      - COL_12 -> CONSTRUCTION
      - COL_16 -> ASSISTANCE
    + mapping de variantes fréquentes (COMAR)
    """
    # mapping direct (clé normalisée -> cible)
    direct_map = {
        _norm_key("Incendie"): "INCENDIE",
        _norm_key("Accident Travail"): "A.TRAVAIL",
        _norm_key("A.TRAVAIL"): "A.TRAVAIL",
        _norm_key("RC"): "RC",
        _norm_key("Automobile"): "AUTOMOBILE",
        _norm_key("Transport"): "TRANSPORT",
        _norm_key("Groupe"): "GROUPE",
        _norm_key("Biens"): "DOMMAGES AUX BIENS",
        _norm_key("Dommages aux biens"): "DOMMAGES AUX BIENS",
        _norm_key("Risques Agricoles"): "RISQUES AGRICOLES",
        _norm_key("d'Eploitation"): "PERTE D'EXPLOITATION",
        _norm_key("d'Exploitation"): "PERTE D'EXPLOITATION",
        _norm_key("Perte d'exploitation"): "PERTE D'EXPLOITATION",
        _norm_key("Caution"): "CAUTION",
        _norm_key("Corporel"): "A.CORPOREL",
        _norm_key("Accident corporel"): "A.CORPOREL",
        _norm_key("Acceptation"): "ACCEPTATION",
        _norm_key("Total"): "TOTAL",
        _norm_key("TOTAL"): "TOTAL",
        _norm_key("CATEGORIES"): "CATEGORIES",
        _norm_key("Categories"): "CATEGORIES",
        _norm_key("COL_12"): "CONSTRUCTION",
        _norm_key("COL_16"): "ASSISTANCE",
    }

    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        if cell.value is None:
            continue

        k = _norm_key(cell.value)
        if k in direct_map:
            cell.value = direct_map[k]
            continue

        # fuzzy match vers EXPECTED_COLUMNS (si proche)
        match, score = best_match(str(cell.value), EXPECTED_COLUMNS, min_score=78.0)
        if match:
            cell.value = match


def normalize_rows_inplace(ws, category_col: int = 1, start_row: int = 2):
    """
    Normalise les libellés de lignes (colonne CATEGORIES) par fuzzy match.
    N'altère pas la structure / style.
    """
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=category_col)
        if cell.value is None:
            continue
        raw = str(cell.value).strip()
        if not raw:
            continue

        match, score = best_match(raw, EXPECTED_ROWS, min_score=75.0)
        if match:
            cell.value = match


# =========================
# Auto-size colonnes (largeur)
# =========================

from openpyxl.utils import get_column_letter

def autosize_columns(ws, min_w: float = 10.0, max_w: float = 80.0, padding: float = 2.5):
    """
    Ajuste la largeur des colonnes selon le contenu.
    Donne plus de poids aux entêtes (ligne 1).
    """
    header_row = 1

    for col_cells in ws.columns:
        first_cell = col_cells[0]
        col_letter = get_column_letter(first_cell.column)

        max_len = 0

        # header (poids fort)
        hv = ws.cell(row=header_row, column=first_cell.column).value
        if hv is not None:
            hs = str(hv)
            hs = max(hs.splitlines(), key=len) if "\n" in hs else hs
            max_len = max(max_len, int(len(hs) * 1.35))

        # contenu
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            s = max(s.splitlines(), key=len) if "\n" in s else s
            max_len = max(max_len, len(s))

        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + padding))


def autosize_rows(ws, min_h: float = 15.0, max_h: float = 140.0):
    """
    Ajuste la hauteur des lignes selon le contenu (wrap).
    Heuristique simple et stable.
    """
    for r in range(1, ws.max_row + 1):
        max_lines = 1
        max_chars = 0

        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v)
            lines = s.splitlines() if "\n" in s else [s]
            max_lines = max(max_lines, len(lines))
            max_chars = max(max_chars, max(len(x) for x in lines))

        h = 15.0 * max_lines
        if max_chars > 45:
            h += 6.0

        ws.row_dimensions[r].height = max(min_h, min(max_h, h))




# =========================
# Validation C1: TOTAL - somme(autres colonnes)
# =========================

@dataclass
class InvalidCell:
    excel_row: int
    c1_value: float



def _find_header_map(ws, header_row: int = 1) -> Dict[str, int]:
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        m[_norm_key(v)] = c
    return m


def clear_previous_greens(ws):
    """
    Compatibilité avec l'ancien nom.
    On efface les verts/oranges/rouges appliqués par le script,
    pour que le vert disparaisse au cycle suivant.
    """
    clear_previous_marks(ws)



import copy
from openpyxl.styles import PatternFill

FILL_NONE = PatternFill(fill_type=None)

def clear_previous_marks(ws):
    """
    Efface uniquement les couleurs qu'on applique pendant la validation des LIGNES (C1):
    - vert + orange
    (on évite d'effacer d'éventuels styles d'origine)
    """
    target_rgbs = {"FFA500", "00B050"}  # ORANGE, GREEN
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            fill = getattr(cell, "fill", None)
            rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
            if fill and fill.patternType == "solid" and rgb in target_rgbs:
                cell.fill = copy.copy(FILL_NONE)


import copy
from dataclasses import dataclass
from typing import List, Dict, Optional

TOL = 5.0  # tolérance [-5, +5]

FILL_ORANGE = PatternFill("solid", fgColor="FFA500")
FILL_GREEN  = PatternFill("solid", fgColor="00B050")

@dataclass
class InvalidCell:
    excel_row: int
    c1_value: float

def validate_c1_inplace(ws, header_row: int = 1, data_start_row: int = 2) -> List[InvalidCell]:
    """
    Validation UNIQUEMENT par lignes:
      C1 = TOTAL - somme(des autres colonnes)
    Coloration:
      - C1 vert si abs(C1) <= TOL
      - C1 orange si abs(C1) > TOL
    AUCUNE autre cellule n'est coloriée.
    """
    headers = _find_header_map(ws, header_row=header_row)
    cat_col = headers.get("CATEGORIES")
    total_col = headers.get("TOTAL")
    c1_col = headers.get("C1")

    if not cat_col or not total_col or not c1_col:
        raise ValueError("Il faut CATEGORIES, TOTAL et C1 dans l'entête.")

    # colonnes numériques = toutes sauf CATEGORIES et C1
    numeric_cols = [c for k, c in headers.items() if k not in ("CATEGORIES", "C1")]

    invalids: List[InvalidCell] = []

    for r in range(data_start_row, ws.max_row + 1):
        cat = ws.cell(row=r, column=cat_col).value
        if cat is None or str(cat).strip() == "":
            continue

        total_val = parse_number(ws.cell(row=r, column=total_col).value)
        if total_val is None:
            continue

        row_sum = 0.0
        for c in numeric_cols:
            if c == total_col:
                continue
            v = parse_number(ws.cell(row=r, column=c).value)
            if v is None:
                continue
            row_sum += v

        c1 = total_val - row_sum

        c1_cell = ws.cell(row=r, column=c1_col)
        c1_cell.value = c1

        if abs(c1) <= TOL:
            c1_cell.fill = copy.copy(FILL_GREEN)
        else:
            c1_cell.fill = copy.copy(FILL_ORANGE)
            invalids.append(InvalidCell(excel_row=r, c1_value=c1))

    return invalids


import copy

def format_header_row(ws, header_row: int = 1, min_h: float = 34.0):
    """
    Agrandit la ligne des entêtes + wrap_text SANS modifier un objet style immutable.
    """
    current_h = ws.row_dimensions[header_row].height or 0
    ws.row_dimensions[header_row].height = max(current_h, min_h)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        if cell.value is None:
            continue

        # IMPORTANT: openpyxl styles sont immutables => on réassigne une COPIE
        if cell.alignment is not None:
            al = copy.copy(cell.alignment)
        else:
            from openpyxl.styles import Alignment
            al = Alignment()

        al.wrap_text = True
        al.vertical = al.vertical or "center"
        al.horizontal = al.horizontal or "center"
        cell.alignment = al




# =========================
# Excel loop (CTRL+S)
# =========================

def _open_excel_2010_or_default(file_path: str):
    """
    Ouvre le fichier dans Excel (idéalement 2010).
    Si EXCEL2010_PATH est défini, on lance cet exe.
    Sinon on tente via COM standard (Excel.Application).
    """
    try:
        import win32com.client  # pywin32
        import pythoncom
    except Exception:
        win32com = None
        pythoncom = None

    if EXCEL2010_PATH and os.path.exists(EXCEL2010_PATH):
        # Lance Excel 2010 directement
        subprocess.Popen([EXCEL2010_PATH, file_path], close_fds=True)
        return None  # pas de handle COM
    else:
        if win32com is None:
            # fallback: ouverture simple (explorer)
            os.startfile(file_path)
            return None

        pythoncom.CoInitialize()
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = True
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(os.path.abspath(file_path))
        return (xl, wb)


def _close_excel_handle(handle):
    if not handle:
        return
    xl, wb = handle
    try:
        wb.Close(SaveChanges=True)
    except:
        pass
    try:
        xl.Quit()
    except:
        pass


def wait_for_ctrl_s(file_path: str, poll: float = 0.8, timeout_sec: Optional[int] = None) -> bool:
    """
    On attend que le fichier soit sauvegardé (= mtime change).
    """
    try:
        last_mtime = os.path.getmtime(file_path)
    except:
        last_mtime = None

    start = time.time()
    while True:
        time.sleep(poll)

        if timeout_sec is not None and (time.time() - start) > timeout_sec:
            return False

        try:
            cur = os.path.getmtime(file_path)
        except:
            continue

        if last_mtime is None:
            last_mtime = cur
            continue

        if cur != last_mtime:
            return True


def save_with_retries(wb, path: str, tries: int = 10, sleep_sec: float = 0.6):
    last_err = None
    for _ in range(tries):
        try:
            wb.save(path)
            return
        except Exception as e:
            last_err = e
            time.sleep(sleep_sec)
    raise last_err


# =========================
# Pipeline Annexe 13
# =========================

def normalize_excel_annexe13_keep_style(in_path: str, out_path: str) -> str:
    wb = openpyxl.load_workbook(in_path)
    ws = wb.active  # tableau principal

    # 1) normaliser colonnes / lignes
    normalize_columns_inplace(ws, header_row=1)

    # detecter col CATEGORIES
    headers = _find_header_map(ws, header_row=1)
    cat_col = headers.get("CATEGORIES", 1)

    normalize_rows_inplace(ws, category_col=cat_col, start_row=2)

    # 2) assurer C1
    ensure_c1_column_inplace(ws, header_row=1)

    # 3) auto-size
    autosize_columns(ws)
    format_header_row(ws, header_row=1, min_h=28.0)
    autosize_columns(ws)

    save_with_retries(wb, out_path)
    return out_path

def add_c1_c2_lines(ws):
    last_row = ws.max_row
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip():
            last_row = r
            break

    start_row = last_row + 2

    rules = [
        "C1 = TOTAL - somme des autres colonnes = 0",
        "C2 = Primes acquises − (Primes émises + Variation des primes non acquises) = 0"
    ]

    for i, desc in enumerate(rules, start=start_row):
        ws.cell(row=i, column=1).value = desc
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)



def apply_colors(ws, c_rows, bad_cells):
    """
    Applique les couleurs de façon très ciblée :
    - Orange sur les cellules LHS des règles invalides (C1 ou nom de règle)
    - Rouge UNIQUEMENT sur l'intersection ligne erronée × colonne erronée
    - Pas de vert au premier tour (vert appliqué seulement dans les fonctions de validation quand OK après correction)
    """
    # Orange sur LHS invalides
    for r, c in c_rows:
        cell = ws.cell(r, c)
        
        cell.font = Font(bold=True)

    # Rouge seulement sur les intersections réelles (ligne invalide ET colonne fautive)
    invalid_rows = set(r for r, _ in c_rows)  # lignes avec erreur
    invalid_cols = set()  # colonnes avec erreur significative

    # On identifie les colonnes fautives (celles où la contribution est forte ou où C1 est hors tol)
    for r, c in bad_cells:
        invalid_cols.add(c)

    # Rouge uniquement à l'intersection
    for r in invalid_rows:
        for c in invalid_cols:
            cell = ws.cell(r, c)
            if cell.fill != FILL_ORANGE:  # ne pas écraser orange sur LHS
                cell.fill = FILL_RED
                cell.font = Font(color="FFFFFF", bold=True)

    # PAS DE VERT ICI → le vert est appliqué uniquement dans evaluate_total_c1_inplace
    # et validate_c2_to_c8_inplace quand la différence est <= tol



def validate_excel_loop_annexe13_keep_style(xlsx_path: str, tol: float = 5.0) -> int:
    """
    Validation Annexe 13 simplifiée : UNIQUEMENT C1 et C2
    - Pas de vert au premier tour
    - Vert seulement après correction quand valide
    - Boucle tant que invalide → s'arrête automatiquement quand C1 et C2 sont OK
    - Ctrl+S détecté pour revalider
    """
    print("=== Validation Annexe 13 démarrée (C1 + C2 uniquement) ===")
    print("→ Ctrl+S dans Excel pour revalider après correction")
    print("→ Sortie automatique quand C1 et C2 sont valides")

    iteration = 0

    while True:
        iteration += 1
        print(f"\n--- Itération {iteration} ---")

        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.active
        except Exception as e:
            print(f"ERREUR lecture : {e}")
            time.sleep(2)
            continue

        # Effacer toutes les couleurs précédentes
        clear_only_our_fills(ws, first_data_row=2)

        # Normalisation
        normalize_columns_inplace(ws)
        normalize_rows_inplace(ws)

        # Force C1
        ensure_c1_column_inplace(ws, header_row=1)
        c1_rows, bad_cells_c1 = evaluate_total_c1_inplace(ws, tol=tol)

        # C2 uniquement
        c_rows_c2, bad_cells_c2 = validate_c2_only_inplace(ws, tol=tol)

        # Fusion
        c_rows = c1_rows + c_rows_c2
        bad_cells = bad_cells_c1 + bad_cells_c2

        # Appliquer couleurs (orange/rouge seulement)
        apply_colors(ws, c_rows, bad_cells)

        # Ajouter lignes C1 et C2 sous le tableau
        add_c1_c2_lines(ws)

        # Ajuster colonnes
        autofit_columns_keep_style(ws)

        # Sauvegarde
        save_with_retries(wb, xlsx_path)
        wb.close()
        del wb

        # Résultat
        if not c_rows and not bad_cells:
            print("STATUT: Valide ✅ (C1 et C2 OK)")
            print("→ Validation terminée automatiquement")
            return 0  # SORT quand tout est valide

        print(f"STATUT: Invalide ❌ | {len(c_rows)} erreurs détectées")
        print("→ Ctrl+S dans Excel pour revalider...")

        # Détection Ctrl+S
        last_mtime = os.path.getmtime(xlsx_path)
        last_size = os.path.getsize(xlsx_path)
        while True:
            time.sleep(0.6)
            try:
                current_mtime = os.path.getmtime(xlsx_path)
                current_size = os.path.getsize(xlsx_path)
            except:
                continue
            if current_mtime > last_mtime or current_size != last_size:
                print("→ Ctrl+S détecté → re-validation...")
                time.sleep(1.0)  # pause pour Excel
                last_mtime = current_mtime
                last_size = current_size
                break


def validate_c2_only_inplace(ws, tol=5.0):
    c_rows = []
    bad_cells = []

    def norm(s):
        if not s:
            return ""
        s = str(s).upper()
        s = s.replace("É", "E").replace("È", "E").replace("Ê", "E")
        s = s.replace("À", "A").replace("Â", "A")
        s = s.replace("Ç", "C")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def find_row(label):
        norm_label = norm(label)
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, 1).value
            if cell and norm_label in norm(str(cell)):
                return r
        return None

    # C2 uniquement
    lhs_r = find_row("PRIMES ACQUISES")  # si la ligne existe
    if lhs_r is None:
        # Sinon on utilise SOLDE DE SOUSSCRIPTION comme proxy (souvent dérivé)
        lhs_r = find_row("SOLDE DE SOUSSCRIPTION")
        if lhs_r is None:
            print("→ Aucune ligne pour C2 trouvée (ni PRIMES ACQUISES ni SOLDE DE SOUSSCRIPTION)")
            return c_rows, bad_cells

    rhs_rows = []
    for name in ["PRIMES EMISES", "VARIATION DES PRIMES NON ACQUISES"]:
        rr = find_row(name)
        if rr is None:
            print(f"→ Ligne '{name}' non trouvée pour C2")
            return c_rows, bad_cells
        rhs_rows.append(rr)

    lhs_vals = _row_numeric_vals(ws, lhs_r)
    rhs_vals_list = [_row_numeric_vals(ws, rr) for rr in rhs_rows]

    worst_col = None
    worst_diff = 0.0

    for col in range(2, ws.max_column + 1):
        lhs_v = lhs_vals.get(col, 0.0)
        rhs_sum = sum(d.get(col, 0.0) for d in rhs_vals_list)
        diff = lhs_v - rhs_sum
        if abs(diff) > abs(worst_diff):
            worst_diff = diff
            worst_col = col

    if worst_col is None:
        return c_rows, bad_cells

    lhs_cell = ws.cell(lhs_r, worst_col)
    lhs_cell.value = round(worst_diff, 2)

    if abs(worst_diff) > tol:
        lhs_cell.fill = FILL_ORANGE
        lhs_cell.font = Font(bold=True)
        c_rows.append((lhs_r, worst_col))

        denom = abs(worst_diff) if abs(worst_diff) > 1e-9 else 0.0
        contribs = []
        for rr, rhs_vals in zip(rhs_rows, rhs_vals_list):
            v = rhs_vals.get(worst_col, 0.0)
            if v != 0:
                ratio = abs(v) / denom
                if ratio > 0.3:
                    contribs.append(rr)

        if not contribs:
            best_rr, best_abs = None, 0
            for rr, rhs_vals in zip(rhs_rows, rhs_vals_list):
                vabs = abs(rhs_vals.get(worst_col, 0.0))
                if vabs > best_abs:
                    best_abs = vabs
                    best_rr = rr
            if best_rr:
                contribs = [best_rr]

        for rr in contribs:
            bad_cells.append((rr, worst_col))
    else:
        lhs_cell.fill = FILL_GREEN
        for rr in rhs_rows:
            ws.cell(rr, worst_col).fill = FILL_GREEN

    return c_rows, bad_cells

            

def evaluate_total_c1_inplace(ws, tol=5.0, header_row=1, first_data_row=2, contrib_threshold=0.30):
    headers = _find_header_map(ws, header_row=header_row)
    cat_col = headers.get("CATEGORIES")
    total_col = headers.get("TOTAL")
    c1_col = headers.get("C1")

    if not cat_col or not total_col or not c1_col:
        raise ValueError("Il faut CATEGORIES, TOTAL et C1 dans l'entête.")

    numeric_cols = [c for k, c in headers.items() if k not in ("CATEGORIES", "C1")]

    c1_rows = []
    bad_cells = []

    for r in range(first_data_row, ws.max_row + 1):
        cat = ws.cell(r, cat_col).value
        if cat is None or str(cat).strip() == "":
            continue

        total_val = parse_number(ws.cell(r, total_col).value)
        if total_val is None:
            continue

        vals = {}
        s = 0.0
        for c in numeric_cols:
            if c == total_col:
                continue
            v = parse_number(ws.cell(r, c).value)
            v = 0.0 if v is None else float(v)
            vals[c] = v
            s += v

        c1 = float(total_val) - float(s)
        ws.cell(r, c1_col).value = round(c1, 2)

        if abs(c1) <= tol:
            continue

        c1_rows.append((r, c1_col))

        denom = abs(s) if abs(s) > 1e-9 else 0.0
        contribs = []
        if denom > 0:
            for c, v in vals.items():
                if v != 0 and (abs(v) / denom) > contrib_threshold:
                    contribs.append(c)

        if not contribs and vals:
            best_c = max(vals.keys(), key=lambda cc: abs(vals[cc]))
            contribs = [best_c]

        for cc in contribs:
            bad_cells.append((r, cc))

    return c1_rows, list(set(bad_cells))


def validate_c2_to_c8_inplace(ws, tol=5.0):
    c_rows = []
    bad_cells = []

    def norm(s):
        if not s:
            return ""
        s = str(s).upper()
        s = s.replace("É", "E").replace("È", "E").replace("Ê", "E")
        s = s.replace("À", "A").replace("Â", "A")
        s = s.replace("Ç", "C").replace("’", "'")
        s = re.sub(r"\s+", " ", s).strip()
        s = re.sub(r"[^A-Z0-9\s]", "", s)
        return s

    def find_row(label):
        norm_label = norm(label)
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, 1).value
            if cell and norm_label in norm(str(cell)):
                return r
        return None

    def mark_rule(lhs_name, rhs_names, rule_label):
        lhs_r = find_row(lhs_name)
        if lhs_r is None:
            print(f"→ Ligne '{lhs_name}' non trouvée pour {rule_label}")
            return

        rhs_rows = []
        for name in rhs_names:
            rr = find_row(name)
            if rr is None:
                print(f"→ Ligne '{name}' non trouvée pour {rule_label}")
                return
            rhs_rows.append(rr)

        lhs_vals = _row_numeric_vals(ws, lhs_r)
        rhs_vals_list = [_row_numeric_vals(ws, rr) for rr in rhs_rows]

        worst_col = None
        worst_diff = 0.0

        for col in range(2, ws.max_column + 1):
            lhs_v = lhs_vals.get(col, 0.0)
            rhs_sum = sum(d.get(col, 0.0) for d in rhs_vals_list)
            diff = lhs_v - rhs_sum
            if abs(diff) > abs(worst_diff):
                worst_diff = diff
                worst_col = col

        if worst_col is None:
            return

        lhs_cell = ws.cell(lhs_r, worst_col)
        lhs_cell.value = round(worst_diff, 2)

        if abs(worst_diff) > tol:
            lhs_cell.fill = FILL_ORANGE
            lhs_cell.font = Font(bold=True)
            c_rows.append((lhs_r, worst_col))

            denom = abs(worst_diff) if abs(worst_diff) > 1e-9 else 0.0
            contribs = []
            for rr, rhs_vals in zip(rhs_rows, rhs_vals_list):
                v = rhs_vals.get(worst_col, 0.0)
                if v != 0:
                    ratio = abs(v) / denom
                    if ratio > 0.3:
                        contribs.append(rr)

            if not contribs:
                best_rr, best_abs = None, 0
                for rr, rhs_vals in zip(rhs_rows, rhs_vals_list):
                    vabs = abs(rhs_vals.get(worst_col, 0.0))
                    if vabs > best_abs:
                        best_abs = vabs
                        best_rr = rr
                if best_rr:
                    contribs = [best_rr]

            for rr in contribs:
                bad_cells.append((rr, worst_col))
        else:
            # Vert seulement si valide
            lhs_cell.fill = FILL_GREEN
            for rr in rhs_rows:
                ws.cell(rr, worst_col).fill = FILL_GREEN

    # Règles avec libellés très proches du fichier réel
    mark_rule("SOLDE DE SOUSSCRIPTION", ["PRIMES EMISES", "VARIATION DES PRIMES NON ACQUISES"], "C2")
    mark_rule("CHARGES DE PRESTATIONS", ["PRESTATIONS ET FRAIS PAYES", "CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE"], "C3")
    mark_rule("SOLDE DE SOUSSCRIPTION", ["PRIMES EMISES", "CHARGES DE PRESTATIONS"], "C4")
    mark_rule("CHARGES D'ACQUISITION ET DE GESTION NETTES", ["FRAIS D'ACQUISITION", "AUTRES CHARGES DE GESTION NETTES"], "C5")
    mark_rule("SOLDE FINANCIER", ["PRODUITS NETS DE PLACEMENTS", "PARTICIPATION AUX RESULTATS"], "C6")
    mark_rule("SOLDE DE REASSURANCE ET/OU DE RETROCESSION", [
        "PRIMES CEDEES AUX REASSUREURS",
        "PART REASSUREURS DANS LES PRIMES ACQUISES",
        "PART REASSUREURS DANS LES PRESTATIONS PAYEES",
        "PART DES REAS ET/OU DES RETROC DANS LES CHARGES DE PROV POUR PRESTATION",
        "PART REASSUREURS DANS LA PARTICIPATION AUX RESULTATS",
        "COMMISSIONS RECUES DES REASSUREURS"
    ], "C7")
    mark_rule("RESULTAT TECHNIQUE NON VIE", [
        "SOLDE DE SOUSSCRIPTION",
        "CHARGES D'ACQUISITION ET DE GESTION NETTES",
        "SOLDE FINANCIER",
        "SOLDE DE REASSURANCE ET/OU DE RETROCESSION"
    ], "C8")

    return c_rows, bad_cells



def validate_annexe13_rules_inplace(ws, tol=5.0, header_row=1, first_data_row=2, contrib_threshold=0.30):
    """
    Applique les 8 règles C2 à C8
    - Pour chaque règle invalide : marque LHS orange + RHS contributrices rouges
    - Valide : vert sur LHS et RHS
    Retourne : c_rows (lignes à orange), bad_cells (cellules rouges)
    """
    c_rows = []
    bad_cells = []

    def _mark_rule(lhs_name, rhs_names_list):
        lhs_r = _row_index_by_name(ws, lhs_name, header_row, first_data_row)
        if lhs_r is None:
            return

        rhs_rows = []
        for name in rhs_names_list:
            rr = _row_index_by_name(ws, name, header_row, first_data_row)
            if rr is None:
                return
            rhs_rows.append(rr)

        lhs_vals = _row_numeric_vals(ws, lhs_r)
        rhs_vals_list = [_row_numeric_vals(ws, rr) for rr in rhs_rows]

        worst_col = None
        worst_diff = 0.0

        for col in range(2, ws.max_column + 1):  # saute CATEGORIES
            lhs_v = lhs_vals.get(col, 0.0)
            rhs_sum = sum(d.get(col, 0.0) for d in rhs_vals_list)
            diff = lhs_v - rhs_sum
            if abs(diff) > abs(worst_diff):
                worst_diff = diff
                worst_col = col

        if worst_col is None or abs(worst_diff) <= tol:
            # valide → vert sur LHS et tous RHS
            ws.cell(lhs_r, worst_col).fill = FILL_GREEN
            for rr in rhs_rows:
                ws.cell(rr, worst_col).fill = FILL_GREEN
            return

        # invalide → orange sur LHS
        lhs_cell = ws.cell(lhs_r, worst_col)
        lhs_cell.value = round(worst_diff, 2)
       
        lhs_cell.font = Font(bold=True)
        c_rows.append((lhs_r, worst_col))

        # rouge sur contributrices
        denom = abs(worst_diff) if abs(worst_diff) > 1e-9 else 0.0
        if denom <= 0:
            return

        contribs = []
        for rr, rhs_vals in zip(rhs_rows, rhs_vals_list):
            v = rhs_vals.get(worst_col, 0.0)
            if v != 0:
                ratio = abs(v) / denom
                if ratio > contrib_threshold:
                    contribs.append(rr)

        # fallback : la plus grande |valeur|
        if not contribs:
            best_rr, best_abs = None, 0
            for rr, rhs_vals in zip(rhs_rows, rhs_vals_list):
                vabs = abs(rhs_vals.get(worst_col, 0.0))
                if vabs > best_abs:
                    best_abs = vabs
                    best_rr = rr
            if best_rr:
                contribs = [best_rr]

        for rr in contribs:
            bad_cells.append((rr, worst_col))

    # Tes 8 règles (exactement comme tu les as listées)
    _mark_rule("PRIMES ACQUISES", ["PRIMES EMISES", "VARIATION DES PRIMES NON ACQUISES"])  # C2
    _mark_rule("CHARGES DE PRESTATIONS", ["PRESTATIONS ET FRAIS PAYES", "CHARGES DES PROVISIONS POUR PRESTATIONS DIVERSE"])  # C3
    _mark_rule("SOLDE DE SOUSCRIPTION", ["PRIMES ACQUISES", "CHARGES DE PRESTATIONS"])  # C4
    _mark_rule("CHARGES D'ACQUISITION ET DE GESTION NETTES", ["FRAIS D'ACQUISITION", "AUTRES CHARGES DE GESTION NETTES"])  # C5
    _mark_rule("SOLDE FINANCIER", ["PRODUITS NETS DE PLACEMENTS", "PARTICIPATION AUX RESULTATS"])  # C6
    _mark_rule("SOLDE DE REASSURANCE ET/OU DE RETROCESSION", [
        "PRIMES CEDEES AUX REASSUREURS",
        "PART REASSUREURS DANS LES PRIMES ACQUISES",
        "PART REASSUREURS DANS LES PRESTATIONS PAYEES",
        "PART DES REAS ET/OU DES RETROC DANS LES CHARGES DE PROV POUR PRESTATION",
        "PART REASSUREURS DANS LA PARTICIPATION AUX RESULTATS",
        "COMMISSIONS RECUES DES REASSUREURS"
    ])  # C7
    _mark_rule("RESULTAT TECHNIQUE NON VIE", [
        "SOLDE DE SOUSSCRIPTION",
        "CHARGES D'ACQUISITION ET DE GESTION NETTES",
        "SOLDE FINANCIER",
        "SOLDE DE REASSURANCE ET/OU DE RETROCESSION"
    ])  # C8

    return c_rows, bad_cells



def _num(v):
    """Convertit en float ou 0"""
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except:
        return 0.0


def _row_numeric_vals(ws, row):
    """Dict {col: valeur numérique} pour une ligne"""
    vals = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row, c).value
        vals[c] = _num(v)
    return vals


def _row_index_by_name(ws, name, header_row=1, first_data_row=2):
    """Trouve la ligne d'un libellé dans CATEGORIES"""
    name_norm = _norm_key(name)
    for r in range(first_data_row, ws.max_row + 1):
        cell = ws.cell(r, 1).value
        if cell and name_norm in _norm_key(str(cell)):
            return r
    return None


def clear_only_our_fills(ws, first_data_row=2):
    target = {"FFFF0000", "00FF0000", "FFFFA500", "00FFA500", "FF00B050", "00B050"}
    for r in range(first_data_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            try:
                rgb = (cell.fill.fgColor.rgb or "").upper()
            except Exception:
                rgb = ""
            if rgb in target:
                cell.fill = PatternFill(fill_type=None)



def autofit_columns_keep_style(ws, min_width=8.0, max_width=60.0, padding=2.0):
    # (copie ta fonction existante si elle est différente)
    # ... ton code autofit ici ...
    pass  # remplace par ta vraie implémentation


def save_with_retries(wb, path, retries=3):
    for _ in range(retries):
        try:
            wb.save(path)
            return
        except PermissionError:
            time.sleep(1)
    print("ERREUR : Impossible de sauvegarder (fichier ouvert ?)")

# =========================
# MAIN
# =========================

def _make_default_out_path(in_path: str) -> str:
    """
    Si l’utilisateur lance: py B.py "13E2024.xlsx"
    -> on travaille sur le même fichier (in-place) pour garder style + boucle.
    Mais tu peux décider de sortir un NV séparé au début.
    Ici: si le nom contient '13E2024' -> on produit '13NV2024.xlsx' une seule fois,
    puis on boucle sur ce NV.
    """
    folder = os.path.dirname(os.path.abspath(in_path))
    base = os.path.basename(in_path)
    name, ext = os.path.splitext(base)

    if "13E2024" in name.upper():
        return os.path.join(folder, "13NV2024.xlsx")
    if name.upper().endswith("E2024"):
        return os.path.join(folder, name[:-5] + "NV2024.xlsx")
    # fallback
    return os.path.join(folder, f"{name}_NV.xlsx")


def main() -> int:
    if len(sys.argv) < 2:
        print('Usage: py B.py "13E2024.xlsx"')
        return 2

    in_path = sys.argv[1]
    if not os.path.isabs(in_path):
        in_path = os.path.abspath(in_path)

    if not os.path.exists(in_path):
        print(f"Fichier introuvable: {in_path}")
        return 2

    out_path = _make_default_out_path(in_path)

    # 1) normalisation (copie style)
    if os.path.abspath(out_path) != os.path.abspath(in_path):
        print(f"➡️ Normalisation Annexe 13: {in_path}")
        out_path = normalize_excel_annexe13_keep_style(in_path, out_path)
        print(f"✅ Fichier normalisé: {out_path}")
    else:
        # Si tu veux vraiment travailler in-place sans créer NV:
        # assure-toi quand même que colonnes/lignes/C1 sont prêts
        print(f"➡️ Normalisation Annexe 13 (in-place): {in_path}")
        normalize_excel_annexe13_keep_style(in_path, in_path)
        out_path = in_path
        print(f"✅ Fichier normalisé: {out_path}")

    # 2) boucle validation C1
    return validate_excel_loop_annexe13_keep_style(out_path)


if __name__ == "__main__":
    raise SystemExit(main())