# -*- coding: utf-8 -*-
# =================================================================================================
# A.py - Extraction CMF (EXTRACTION UNIQUEMENT / sans validation)
# - Selenium headless: sélection société + récupération liens PDF
# - Download PDF
# - Insertion MySQL (cmf.document) avec UNIQUE (Societe, Nom, Annee)
# - Détection Annexe 12/13:
#       1) texte natif (PyPDF2)
#       2) OCR ciblé (dernières pages) (pdf2image + pytesseract)
#       3) fallback contenu tableaux Camelot (mots-clés)
# - Extraction tableau:
#       - Natif: camelot stream puis lattice
#       - Scanné: OCR page + parsing lignes/colonnes (simple & stable)
# - Nettoyage robuste (corrige vos problèmes COMAR et reste général):
#       ✅ Suppression colonne parasite remplie de "-" (et propagation du signe sur la colonne suivante)
#       ✅ Correction du "décalage début tableau": drop lignes parasites + détection header robuste
#       ✅ Colonnes uniques
#       ✅ Nettoyage nombres (espaces, NBSP, virgules, tirets, parenthèses...)
#       ✅ Gestion header sur 2 lignes (ex: "Dommages aux" + "Biens", "Perte" + "d'Exploitation", etc.)
# - Export Excel (openpyxl) + style simple
# - Safe save si fichier Excel est ouvert (PermissionError -> nouveau nom timestamp)
# =================================================================================================
from __future__ import annotations
import os
import re
import time
import logging
from datetime import datetime
from turtle import pd
from urllib.parse import urlparse, urlencode, parse_qs
import glob
import subprocess

import requests
import mysql.connector
from mysql.connector import Error


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

import PyPDF2
import pandas as pd
import camelot 

import pytesseract
from pdf2image import convert_from_path
from PIL import Image, ImageOps, ImageFilter
from pathlib import Path
THIS_DIR = Path(__file__).resolve().parent
NV12_SCRIPT = THIS_DIR / "NorVal12.py"
NV13_SCRIPT = THIS_DIR / "NorVal13.py"


from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


# ---------------------------------------------- Paramètres ----------------------------------------------
# SOCIETE = "COMPAGNIE MEDITERRANEENNE D'ASSURANCES ET DE REASSURANCES - COMAR -"
# ANNEE = 2024
ANNEXES = ["12", "13"]

DOC_NAMES_ACCEPTES = [
    "Etats financiers au 31/12",
    "Etats financiers consolidés au 31/12",
]

CMF_URL = "https://www.cmf.tn/consultation-des-tats-financier-des-soci-t-s-faisant-ape"

# OCR
OCR_DPI = 200
OCR_LANG = "fra"
OCR_PSM = "6"

# Excel styling
HEADER_COLOR = "0070C0"

# MySQL
MYSQL_HOST = "localhost"
MYSQL_USER = "root"
MYSQL_PASSWORD = ""
MYSQL_DB = "cmf"


# ---------------------------------------------- Logs ----------------------------------------------
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
os.environ["GLOG_minloglevel"] = "3"
os.environ["PYTHONWARNINGS"] = "ignore"

logging.basicConfig(
    filename="script.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)


# ---------------------------------------------- DB ----------------------------------------------
def create_cmf_database_and_table():
    try:
        connection = mysql.connector.connect(host=MYSQL_HOST, user=MYSQL_USER, password=MYSQL_PASSWORD)
        cursor = connection.cursor()

        logging.info("Création/mise à jour base 'cmf'...")
        print("Création/mise à jour base 'cmf'...")

        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {MYSQL_DB}")
        cursor.execute(f"USE {MYSQL_DB}")
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS document (
                id INT AUTO_INCREMENT PRIMARY KEY,
                Societe VARCHAR(255) NOT NULL,
                Nom VARCHAR(255) NOT NULL,
                Annee INT NOT NULL,
                URL VARCHAR(512) NOT NULL,
                UNIQUE KEY unique_document (Societe, Nom, Annee)
            )
            """
        )
        connection.commit()
        logging.info("Base 'cmf' et table 'document' prêtes.")
        print("Base 'cmf' et table 'document' prêtes.")
        return connection, cursor
    except Error as e:
        logging.error(f"Erreur création base : {e}")
        print(f"Erreur création base : {e}")
        return None, None


def normalize_url(url: str) -> str:
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    query_params.pop("id", None)
    query_params.pop("token", None)
    new_query = urlencode(query_params, doseq=True)
    return parsed_url._replace(query=new_query).geturl()


def extract_year_from_text(text: str):
    if not text:
        return None
    patterns = [
        r"(?:20)(1[5-9]|2[0-5])\b",
        r"\b(20\d{2})\b",
        r"(?:3112|1312)(1[5-9]|2[0-5])",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            year = match.group(1) if match.group(1) else match.group(0)
            if len(year) == 2:
                year = f"20{year}"
            try:
                yi = int(year)
                if 2015 <= yi <= 2025:
                    return str(yi)
            except Exception:
                pass
    return None


def check_document_exists(cursor, societe, nom, annee: int) -> bool:
    try:
        query = "SELECT COUNT(*) FROM document WHERE Societe = %s AND Nom = %s AND Annee = %s"
        cursor.execute(query, (societe, nom, annee))
        return (cursor.fetchone()[0] or 0) > 0
    except Error as e:
        logging.error(f"Erreur vérification document : {e}")
        print(f"Erreur vérification document : {e}")
        return False


def insert_pdf_info_cmf(connection, cursor, societe, nom_document, annee, url) -> bool:
    try:
        normalized_url = normalize_url(url)
        annee_int = int(annee)

        if not (2015 <= annee_int <= 2025):
            logging.info(f"Document {nom_document} ({annee}) ignoré (hors 2015-2025)")
            print(f"Document {nom_document} ({annee}) ignoré (hors 2015-2025)")
            return False

        if check_document_exists(cursor, societe, nom_document, annee_int):
            logging.info(f"Document {nom_document} ({annee}) existe déjà")
            print(f"Document {nom_document} ({annee}) existe déjà")
            return False

        cursor.execute(
            """
            INSERT INTO document (Societe, Nom, Annee, URL)
            VALUES (%s, %s, %s, %s)
            """,
            (societe, nom_document, annee_int, normalized_url),
        )
        connection.commit()
        logging.info(f"AJOUTÉ : {nom_document} ({annee})")
        print(f"AJOUTÉ : {nom_document} ({annee})")
        return True

    except Error as e:
        logging.error(f"Erreur insertion : {e}")
        print(f"Erreur insertion : {e}")
        return False


# ---------------------------------------------- Utilitaires ----------------------------------------------
def safe_filename(text: str) -> str:
    return re.sub(r"[^\w\s-]", "_", (text or "")).strip().replace(" ", "_")


def now_ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")




# ---------------------------------------------- Selenium / Scraping ----------------------------------------------
def make_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    service = Service(ChromeDriverManager().install(), log_path=os.devnull)
    return webdriver.Chrome(service=service, options=chrome_options)


def extract_pdfs_from_page(driver, societe: str):
    wait = WebDriverWait(driver, 25)
    pdfs = []
    target = (societe or "").strip().lower()

    try:
        logging.info(f"Traitement {societe}")
        print(f"\n--- Traitement {societe} ---")

        dropdown = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#edit_field_societesape_value_chosen a")))
        dropdown.click()
        time.sleep(0.7)

        options = wait.until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "#edit_field_societesape_value_chosen .chosen-results li")
            )
        )

        selected_option = None
        for option in options:
            if target == option.text.strip().lower():
                selected_option = option
                break
        if selected_option is None:
            for option in options:
                if target and target in option.text.strip().lower():
                    selected_option = option
                    break

        if selected_option is None:
            logging.error(f"Aucune correspondance pour '{societe}'")
            print(f"ERREUR : Société non trouvée : {societe}")
            return []

        selected_option.click()
        print(f"Sélectionnée : {selected_option.text.strip()}")

        submit = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "#edit-submit-consultation-des-tats-financier-des-soci-t-s-faisant-ape")
            )
        )
        submit.click()
        time.sleep(1.5)

        page_num = 1
        while True:
            print(f"  Page {page_num}...")
            rows = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".view-content .views-row")))

            for row in rows:
                try:
                    pdf_url = row.find_element(By.CSS_SELECTOR, "a[href$='.pdf']").get_attribute("href")
                    pdf_name = row.find_element(By.CSS_SELECTOR, ".field-name-field-p-riode .field-item").text.strip()
                    year = extract_year_from_text(pdf_url) or extract_year_from_text(pdf_name)
                    if year and year.isdigit():
                        pdfs.append({"url": pdf_url, "nom": pdf_name, "annee": year, "societe": societe})
                except Exception:
                    continue

            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, ".pager .next a:not(.disabled)")
                next_btn.click()
                page_num += 1
                time.sleep(2.0)
            except Exception:
                break

    except Exception as e:
        logging.error(f"Erreur extraction PDFs : {str(e)}")
        print(f"Erreur extraction PDFs : {str(e)}")

    return pdfs


def download_pdf(url, societe, nom, annee):
    try:
        safe_soc = safe_filename(societe)
        safe_nom = safe_filename(nom)
        filename = f"{safe_soc}_{safe_nom}_{annee}.pdf"
        filepath = os.path.join(os.getcwd(), filename)

        response = requests.get(url, stream=True, timeout=90)
        response.raise_for_status()
        with open(filepath, "wb") as f:
            f.write(response.content)

        print(f"PDF téléchargé : {filepath}")
        logging.info(f"PDF téléchargé : {filepath}")
        return filepath
    except Exception as e:
        logging.error(f"Erreur téléchargement : {str(e)}")
        print(f"Erreur téléchargement : {str(e)}")
        return None


def fetch_pdf_for_societe_annee(societe, annee, doc_names_acceptes):
    """
    Cherche sur CMF le PDF correspondant à l'année et à un des noms acceptés.
    Retourne: (pdf_path, pdf_url, pdf_nom_reel)
    """
    driver = make_driver()
    try:
        driver.get(CMF_URL)
        pdfs = extract_pdfs_from_page(driver, societe)

        pdfs_annee = [p for p in pdfs if str(p.get("annee", "")).isdigit() and int(p["annee"]) == int(annee)]
        if not pdfs_annee:
            print(f"Document {societe} {annee} non trouvé (aucun PDF année {annee})")
            return None, None, None

        # 1) match exact
        for wanted in doc_names_acceptes:
            for pdf in pdfs_annee:
                if pdf["nom"].strip().lower() == wanted.strip().lower():
                    pdf_path = download_pdf(pdf["url"], pdf["societe"], pdf["nom"], pdf["annee"])
                    return pdf_path, pdf["url"], pdf["nom"]

        # 2) contient un des libellés
        lowered_acceptes = [w.strip().lower() for w in doc_names_acceptes]
        for pdf in pdfs_annee:
            name_low = pdf["nom"].strip().lower()
            if any(w in name_low for w in lowered_acceptes):
                pdf_path = download_pdf(pdf["url"], pdf["societe"], pdf["nom"], pdf["annee"])
                return pdf_path, pdf["url"], pdf["nom"]

        # 3) fallback 1er PDF de l'année
        pdf = pdfs_annee[0]
        print(f"⚠️ Aucun nom exact trouvé. Fallback sur le 1er PDF de {annee}: {pdf['nom']}")
        pdf_path = download_pdf(pdf["url"], pdf["societe"], pdf["nom"], pdf["annee"])
        return pdf_path, pdf["url"], pdf["nom"]

    except Exception as e:
        logging.error(f"ERREUR récupération PDF {societe} {annee} : {str(e)}")
        print(f"ERREUR récupération PDF {societe} {annee} : {str(e)}")
        return None, None, None
    finally:
        try:
            driver.quit()
        except Exception:
            pass


# ---------------------------------------------- Détection Annexes 12/13 ----------------------------------------------
def normalize_text_light(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("nº", "n°").replace("no", "n°").replace("n0", "n°")
    return s


def annexe_keywords(num: str):
    n = str(num)
    return [
        f"annexe {n}",
        f"annexe{n}",
        f"annexe n° {n}",
        f"annexe n°{n}",
        f"annexe n º {n}",
        f"annexe nº {n}",
        f"note n°{n}",
        f"note n° {n}",
    ]


def preprocess_for_ocr(img: Image.Image) -> Image.Image:
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    img = img.filter(ImageFilter.SHARPEN)
    img = img.point(lambda x: 0 if x < 160 else 255, mode="1")
    return img.convert("L")


def ocr_page_text(pdf_path: str, page_num_1based: int) -> str:
    images = convert_from_path(pdf_path, first_page=page_num_1based, last_page=page_num_1based, dpi=OCR_DPI)
    if not images:
        return ""
    img = preprocess_for_ocr(images[0])
    config = f"--psm {OCR_PSM}"
    try:
        txt = pytesseract.image_to_string(img, lang=OCR_LANG, config=config)
        return normalize_text_light(txt)
    except Exception:
        return ""


def detect_annexes_by_table_content(pdf_path: str, candidate_pages):
    kw12 = [
        "engagement", "engagements", "caution", "cautionnements", "hors bilan", "garanties", "sûret", "surete",
        "provisions pour risques",
    ]
    kw13 = [
        "provision", "provisions", "sinistre", "sinistres", "primes", "technique", "techniques",
        "sap", "ibnr", "rbns", "mathematique", "mathématique", "capitaux", "réassurance", "reassurance",
    ]

    found_12 = None
    found_13 = None
    flavors = ["stream", "lattice"]

    for p in candidate_pages:
        for flavor in flavors:
            try:
                tables = camelot.read_pdf(pdf_path, flavor=flavor, pages=str(p))
                if tables.n == 0:
                    continue

                big_txt = []
                for t in tables:
                    df = t.df
                    joined = " ".join(str(x) for x in df.values.flatten() if x is not None)
                    big_txt.append(joined)

                txt = normalize_text_light(" ".join(big_txt))
                s12 = sum(1 for w in kw12 if w in txt)
                s13 = sum(1 for w in kw13 if w in txt)

                if s12 >= 2 and found_12 is None:
                    found_12 = p
                if s13 >= 2 and found_13 is None:
                    found_13 = p

                if found_12 is not None and found_13 is not None:
                    return {"Annexe_12": found_12, "Annexe_13": found_13}

            except Exception:
                continue

    return {"Annexe_12": found_12, "Annexe_13": found_13}


def search_sections_in_pdf(pdf_path: str):
    """
    1) texte natif (PyPDF2)
    2) OCR ciblé sur les dernières pages
    3) fallback: analyse du contenu des tableaux Camelot
    """
    results = {}
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_path)
        num_pages = len(pdf_reader.pages)
        print(f"Analyse PDF : {num_pages} pages")

        # 1) natif
        for num in ANNEXES:
            key = f"Annexe_{num}"
            kws = annexe_keywords(num)
            found = False

            for i in range(num_pages):
                text = pdf_reader.pages[i].extract_text() or ""
                if not text.strip():
                    continue
                t = normalize_text_light(text)
                if any(kw in t for kw in kws):
                    results[key] = (i + 1, False)
                    print(f"{key} trouvée (natif) page {i + 1}")
                    found = True
                    break

            if not found:
                results[key] = (None, None)

        if all(results.get(f"Annexe_{n}", (None, None))[0] for n in ANNEXES):
            return {k: v for k, v in results.items() if v[0] is not None}

        # 2) OCR dernières pages
        last_pages = list(range(max(1, num_pages - 10) + 1, num_pages + 1))
        for num in ANNEXES:
            key = f"Annexe_{num}"
            if results.get(key, (None, None))[0] is not None:
                continue

            print(f"Tentative OCR pour {key} (dernières pages)...")
            kws = annexe_keywords(num)
            for p in last_pages:
                txt = ocr_page_text(pdf_path, p)
                if txt and any(kw in txt for kw in kws):
                    results[key] = (p, True)
                    print(f"{key} trouvée (OCR) page {p}")
                    break

        if all(results.get(f"Annexe_{n}", (None, None))[0] for n in ANNEXES):
            return {k: v for k, v in results.items() if v[0] is not None}

        # 3) fallback tables
        print("⚠️ Fallback: détection Annexe 12/13 par contenu tableaux Camelot...")
        detected = detect_annexes_by_table_content(pdf_path, candidate_pages=last_pages)
        for key, page in detected.items():
            if page and results.get(key, (None, None))[0] is None:
                results[key] = (page, False)
                print(f"{key} détectée par contenu tableau page {page}")

        return {k: v for k, v in results.items() if v[0] is not None}

    except Exception as e:
        logging.error(f"Erreur analyse PDF : {str(e)}")
        print(f"Erreur analyse PDF : {str(e)}")
        return {}


# ---------------------------------------------- Nettoyage tableaux ----------------------------------------------
DASH_CHARS = {"-", "‐", "–", "—", "−"}  # différents tirets / minus unicode


def _normalize_text_cell(x):
    if x is None:
        return ""
    if isinstance(x, float) and pd.isna(x):
        return ""
    s = str(x)
    s = s.replace("\r", " ").replace("\n", " ")
    s = s.replace("\u00a0", " ")  # NBSP
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _is_empty(x) -> bool:
    if x is None:
        return True
    if isinstance(x, float) and pd.isna(x):
        return True
    if isinstance(x, str) and x.strip() == "":
        return True
    return False


def _is_dash_only(x) -> bool:
    if _is_empty(x):
        return False
    return _normalize_text_cell(x) in DASH_CHARS


def clean_number(val):
    """
    Convertit en int/float si possible, sinon renvoie la string nettoyée.
    Gère:
      - espaces, NBSP
      - virgules décimales
      - séparateurs de milliers (espaces, points)
      - tirets unicode comme signe négatif
      - parenthèses => négatif
      - '-' seul => vide
    """
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            if isinstance(val, float) and abs(val - int(val)) < 1e-9:
                return int(val)
            return val
        except Exception:
            return val

    s = _normalize_text_cell(val)
    if s == "":
        return ""
    if s in DASH_CHARS:
        return ""

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    if s and (s[0] in DASH_CHARS):
        negative = True
        s = s[1:].strip()

    s = re.sub(r"[A-Za-z]", "", s).strip()

    s = s.replace("\u00a0", " ").replace(" ", "")
    if s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")

    if s.count(".") > 1:
        s = s.replace(".", "")

    if not re.fullmatch(r"\d+(\.\d+)?", s):
        return _normalize_text_cell(val)

    try:
        num = float(s)
        if abs(num - int(num)) < 1e-9:
            num = int(num)
        if negative:
            return -num
        return num
    except Exception:
        return _normalize_text_cell(val)

def _excel_2010_exe_path():
    candidates = [
        r"C:\Program Files\Microsoft Office\Office14\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office14\EXCEL.EXE",
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


def _open_in_excel_2010(path: str) -> bool:
    """
    Ouvre explicitement avec Excel 2010 si présent.
    """
    try:
        excel = _excel_2010_exe_path()
        abspath = os.path.abspath(path)
        if excel:
            subprocess.Popen([excel, "/e", abspath], shell=False)
            return True
        os.startfile(abspath)  # fallback
        return True
    except Exception:
        return False


def _find_latest_nv_file(folder: str, ann: str, year: int) -> str | None:
    """
    Cherche le dernier fichier NV dans le dossier:
      ex: 12NV2024.xlsx ou 12NV2024_YYYYMMDD_HHMMSS.xlsx
    """
    pattern = os.path.join(folder, f"{ann}NV{year}*.xlsx")
    files = glob.glob(pattern)
    if not files:
        return None
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return files[0]



def _row_is_header_continuation(row_vals, min_text_cells=2, max_numeric_cells=1):
    """
    Heuristique: une ligne est une continuation d'en-têtes si:
      - elle contient au moins min_text_cells cellules avec lettres
      - elle contient au plus max_numeric_cells cellules numériques
    """
    text_cells = 0
    numeric_cells = 0

    for v in row_vals:
        s = _normalize_text_cell(v)
        if s == "":
            continue

        # numérique ?
        parsed = clean_number(s)
        if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
            numeric_cells += 1
            continue

        # texte ? (au moins une lettre)
        if re.search(r"[A-Za-zÀ-ÿ]", s):
            text_cells += 1

    return (text_cells >= min_text_cells) and (numeric_cells <= max_numeric_cells)





def _merge_multiline_headers(df, max_header_rows=3):
    """
    Corrige les cas Camelot où:
      - header sur 2 lignes dans le PDF (ex: 'Dommages aux' puis 'Biens')
      - Camelot met la 2e ligne en première ligne "data"
      - certaines colonnes restent 'COL_12', 'COL_13'...

    Stratégie:
      - on regarde les 1ères lignes de df (jusqu'à max_header_rows)
      - si une ligne ressemble à une continuation de header, on concatène cellule par cellule
        dans le nom de colonne, surtout si la colonne est un placeholder COL_x.
      - on supprime ensuite ces lignes de df
    """
    if df is None or df.empty:
        return df

    df = df.copy()

    rows_to_merge = []
    scan = min(max_header_rows, len(df))

    for i in range(scan):
        row_vals = df.iloc[i].tolist()
        if _row_is_header_continuation(row_vals):
            rows_to_merge.append(i)
        else:
            # dès qu'on tombe sur une vraie ligne data, on stop
            break

    if not rows_to_merge:
        # forcer nom 1ère colonne quand même
        df.columns = _make_unique_columns(list(df.columns), force_first_col="CATEGORIES")
        return df

    new_cols = list(df.columns)

    for ridx in rows_to_merge:
        row_vals = df.iloc[ridx].tolist()
        for j, cell in enumerate(row_vals):
            cell_s = _normalize_text_cell(cell)
            if cell_s == "":
                continue

            # ignorer si purement numérique
            parsed = clean_number(cell_s)
            if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
                continue

            cur = _normalize_text_cell(new_cols[j])

            # si placeholder COL_x => remplacer
            if re.fullmatch(r"COL_\d+(_\d+)?", cur) or cur == "":
                new_cols[j] = cell_s
            else:
                # concat sans doublonner
                if cell_s.lower() not in cur.lower():
                    new_cols[j] = (cur + " " + cell_s).strip()

    # appliquer colonnes + drop les lignes header fusionnées
    df = df.drop(index=rows_to_merge).reset_index(drop=True)
    df.columns = _make_unique_columns(new_cols, force_first_col="CATEGORIES")
    return df




def _make_unique_columns(cols, force_first_col="CATEGORIES"):
    """
    Rend les colonnes uniques.
    - Force la 1ère colonne à 'CATEGORIES'
    - Remplace les vides par COL_i
    """
    seen = {}
    out = []

    for idx, c in enumerate(cols):
        base = _normalize_text_cell(c)

        # Force 1ère colonne
        if idx == 0:
            base = force_first_col

        if base == "":
            base = f"COL_{idx+1}"

        if base not in seen:
            seen[base] = 0
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")

    return out



def _row_nonempty_count(row) -> int:
    return sum(1 for x in row if not _is_empty(_normalize_text_cell(x)))


def _row_numeric_count(row) -> int:
    n = 0
    for x in row:
        v = clean_number(x)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            n += 1
    return n


def _looks_like_header_row(row) -> bool:
    """
    Ligne "header-like": beaucoup de texte, peu de numériques.
    """
    nonempty = _row_nonempty_count(row)
    numeric = _row_numeric_count(row)
    if nonempty == 0:
        return False
    # ex: 12-16 colonnes -> numeric proche de 0
    return (numeric <= max(1, nonempty // 6)) and any(re.search(r"[A-Za-zÀ-ÿ]", _normalize_text_cell(x)) for x in row)


def _detect_best_header_row(df: pd.DataFrame) -> int:
    if df is None or df.empty:
        return 0

    best_i = 0
    best_score = -1e9
    max_scan = min(len(df), 25)

    for i in range(max_scan):
        row = [df.iloc[i, j] for j in range(df.shape[1])]
        nonempty = _row_nonempty_count(row)
        numeric = _row_numeric_count(row)

        score = (nonempty * 3) - (numeric * 5)
        row_txt = normalize_text_light(" ".join(_normalize_text_cell(x) for x in row))
        if any(k in row_txt for k in ["total", "incendie", "automobile", "groupe", "transport", "caution", "vie"]):
            score += 6

        if score > best_score:
            best_score = score
            best_i = i

    return best_i


def _drop_useless_top_rows(df: pd.DataFrame, max_drop=12) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    df = df.copy()
    df = df.applymap(_normalize_text_cell)

    drop_idx = []
    scan = min(len(df), max_drop)
    for i in range(scan):
        row = df.iloc[i].tolist()
        nonempty = _row_nonempty_count(row)
        joined = normalize_text_light(" ".join(row))

        if nonempty <= 1:
            drop_idx.append(i)
            continue

        if any(k in joined for k in ["etats financiers", "états financiers", "annexe", "notes aux", "exercice clos"]):
            if nonempty <= 2:
                drop_idx.append(i)
                continue

    if drop_idx:
        df = df.drop(index=drop_idx).reset_index(drop=True)

    return df


def _merge_wrapped_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fusion prudente retours à la ligne dans libellés (comme votre version),
    sans fusionner n'importe quoi.
    """
    if df is None or df.empty:
        return df

    df = df.copy()
    for c in df.columns:
        df[c] = df[c].map(_normalize_text_cell)

    rows = df.values.tolist()
    out = []
    i = 0
    n = len(rows)

    def _has_numbers(r):
        return any(not _is_empty(r[j]) for j in range(1, len(r)))

    def _all_empty_except_label(r):
        return all(_is_empty(r[j]) for j in range(1, len(r)))

    def _is_all_empty_row(r):
        return _is_empty(r[0]) and all(_is_empty(r[j]) for j in range(1, len(r)))

    while i < n:
        r1 = rows[i]

        if i + 2 < n:
            r2 = rows[i + 1]
            r3 = rows[i + 2]

            if _is_all_empty_row(r2):
                out.append(r1)
                i += 1
                continue

            r1_label_ok = not _is_empty(r1[0])
            r2_label_empty = _is_empty(r2[0])
            r3_label_ok = not _is_empty(r3[0])

            r1_has_nums = _has_numbers(r1)
            r2_has_nums = _has_numbers(r2)
            r3_has_nums = _has_numbers(r3)

            if (
                r1_label_ok and (not r1_has_nums) and _all_empty_except_label(r1)
                and r2_label_empty and r2_has_nums
                and r3_label_ok and (not r3_has_nums) and _all_empty_except_label(r3)
            ):
                fused_label = (str(r1[0]).strip() + " " + str(r3[0]).strip()).strip()
                new_row = r1[:]
                new_row[0] = fused_label
                for j in range(1, len(new_row)):
                    if _is_empty(new_row[j]) and not _is_empty(r2[j]):
                        new_row[j] = r2[j]
                out.append(new_row)
                i += 3
                continue

        out.append(r1)
        i += 1

    return pd.DataFrame(out, columns=df.columns)


def _drop_empty_rows_cols(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df = df.copy()
    df = df.applymap(_normalize_text_cell)

    # drop empty rows
    row_mask = []
    for i in range(len(df)):
        row = df.iloc[i].tolist()
        row_mask.append(_row_nonempty_count(row) > 0)
    df = df.loc[row_mask].reset_index(drop=True)

    # drop empty cols
    to_drop = []
    for c in df.columns:
        col_vals = df[c].tolist()
        if all(_is_empty(v) for v in col_vals):
            to_drop.append(c)
    if to_drop:
        df = df.drop(columns=to_drop, errors="ignore")

    return df


def _propagate_sign_and_drop_dash_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    ✅ Fix COMAR (et beaucoup d'assureurs):
    - La table PDF affiche souvent un "tiret" dans une colonne séparée (colonne parasite)
      et le chiffre dans la colonne suivante.
    - On propage ce signe sur la colonne suivante, puis on supprime la colonne parasite.
    """
    if df is None or df.empty:
        return df

    df = df.copy()
    for c in df.columns:
        df[c] = df[c].map(_normalize_text_cell)

    cols = list(df.columns)

    # propagation signe (col i -> col i+1)
    for ci in range(len(cols) - 1):
        left = cols[ci]
        right = cols[ci + 1]

        left_vals = df[left].tolist()
        nonempty_left = [v for v in left_vals if not _is_empty(v)]
        if len(nonempty_left) == 0:
            continue

        dash_count = sum(1 for v in nonempty_left if _is_dash_only(v))
        dash_ratio = dash_count / max(1, len(nonempty_left))
        if dash_ratio < 0.60:
            continue

        for r in range(len(df)):
            lv = df.at[r, left]
            rv = df.at[r, right]
            if _is_dash_only(lv):
                parsed = clean_number(rv)
                if isinstance(parsed, (int, float)) and parsed > 0:
                    df.at[r, right] = -parsed

    # suppression colonnes (tirets/vides >=80%)
    to_drop = []
    for c in df.columns:
        vals = df[c].tolist()
        dash_or_empty = sum(1 for v in vals if _is_empty(v) or _is_dash_only(v))
        ratio = dash_or_empty / max(1, len(vals))
        if ratio >= 0.80:
            to_drop.append(c)

    if to_drop:
        df = df.drop(columns=to_drop, errors="ignore")

    return df


def _merge_second_header_row_if_needed(df: pd.DataFrame) -> pd.DataFrame:
    """
    ✅ Fix COMAR Annexe 13 (et autres): headers sur 2 lignes.
    Exemple typique:
      ligne header 1: "Dommages aux" | "Perte" | "Accident"
      ligne header 2: "Biens"       | "d'Exploitation" | "Corporel"
    Camelot peut mettre la 2ème ligne comme première ligne de "data".
    => On la fusionne dans les noms de colonnes puis on supprime cette ligne.
    """
    if df is None or df.empty or len(df) < 1:
        return df

    first_row = df.iloc[0].tolist()
    if not _looks_like_header_row(first_row):
        return df

    # On merge cellule par cellule dans le header existant
    new_cols = list(df.columns)
    changed = False
    for j, cont in enumerate(first_row):
        cont_s = _normalize_text_cell(cont)
        if cont_s == "":
            continue
        # si c'est purement numérique, on ignore
        if re.fullmatch(r"[-\d\s.,()]+", cont_s):
            continue

        base = _normalize_text_cell(new_cols[j])
        if base.startswith("COL_") or base == "COL":
            new_cols[j] = cont_s
        else:
            # évite "Total Total" etc
            if cont_s.lower() not in base.lower():
                new_cols[j] = (base + " " + cont_s).strip()
        changed = True

    if changed:
        df = df.iloc[1:].reset_index(drop=True)
        df.columns = _make_unique_columns(new_cols)

    return df



def _row_text_numeric_stats(row_vals):
    """
    Retourne (text_cells, numeric_cells, nonempty_cells)
    text_cells: contient au moins une lettre
    numeric_cells: clean_number() donne int/float
    """
    text_cells = 0
    numeric_cells = 0
    nonempty = 0

    for v in row_vals:
        s = _normalize_text_cell(v)
        if s == "":
            continue
        nonempty += 1

        parsed = clean_number(s)
        if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
            numeric_cells += 1
        elif re.search(r"[A-Za-zÀ-ÿ]", s):
            text_cells += 1

    return text_cells, numeric_cells, nonempty




def _fill_placeholder_headers_from_top_rows(df, max_rows=4):
    """
    Objectif:
      - Corriger les colonnes restées en 'COL_12', 'COL_13'... quand le PDF a des en-têtes sur 2 lignes.
      - Exemple: 'Dommages aux' (ligne header principale) + 'Biens' (ligne suivante) => 'Dommages aux Biens'.

    Méthode:
      - On examine les max_rows premières lignes du DF (après header principal).
      - On collecte par colonne les cellules TEXT (avec lettres) -> header_parts[col]
      - On remplit PRIORITAIREMENT les colonnes placeholder (COL_x) avec la concat de ces parts.
      - On drop les lignes qui ont servi d'extension de header (car ce n'est pas de la data).
    """
    if df is None or df.empty:
        return df

    df = df.copy()

    n_scan = min(max_rows, len(df))
    if n_scan <= 0:
        return df

    cols = list(df.columns)
    header_parts = {c: [] for c in cols}
    used_rows = set()

    # helper placeholder
    def _is_placeholder(colname: str) -> bool:
        c = _normalize_text_cell(colname)
        return bool(re.fullmatch(r"COL_\d+(_\d+)?", c)) or c == ""

    # scan top rows
    for i in range(n_scan):
        row_vals = df.iloc[i].tolist()
        text_cells, numeric_cells, nonempty = _row_text_numeric_stats(row_vals)

        # Une ligne "header continuation" est généralement:
        # - pas mal de texte, peu ou pas de chiffres
        # MAIS parfois il y a 1 chiffre/total qui traîne -> on tolère un peu.
        header_like = (text_cells >= 2 and numeric_cells <= 2)

        # collecter les morceaux de texte par colonne
        contributed = False
        for j, c in enumerate(cols):
            cell = _normalize_text_cell(row_vals[j])
            if cell == "":
                continue

            # ignorer les cellules purement numériques
            parsed = clean_number(cell)
            if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
                continue

            # ignorer si pas de lettres (ex: ponctuation)
            if not re.search(r"[A-Za-zÀ-ÿ]", cell):
                continue

            # on prend surtout si la colonne est placeholder, sinon on reste prudent
            if _is_placeholder(c) or header_like:
                header_parts[c].append(cell)
                contributed = True

        if contributed and header_like:
            used_rows.add(i)

    # construire nouveaux noms de colonnes
    new_cols = []
    changed = False

    for c in cols:
        base = _normalize_text_cell(c)

        parts = header_parts.get(c, [])
        parts = [p.strip() for p in parts if p.strip()]

        if parts:
            merged = " ".join(parts)
            merged = re.sub(r"\s+", " ", merged).strip()

            # si placeholder -> remplacer complètement
            if _is_placeholder(base):
                base = merged
                changed = True
            else:
                # si header existant incomplet, concat (sans doublon)
                if merged.lower() not in base.lower():
                    base = (base + " " + merged).strip()
                    changed = True

        new_cols.append(base)

    # appliquer colonnes, forcer CATEGORIES en 1ère colonne
    new_cols = _make_unique_columns(new_cols, force_first_col="CATEGORIES")
    df.columns = new_cols

    # drop les lignes utilisées comme extension header
    if used_rows:
        df = df.drop(index=list(used_rows)).reset_index(drop=True)

    return df




def clean_table_general(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pipeline nettoyage:
      - détection header principal
      - 1ère colonne = CATEGORIES
      - complète headers multi-lignes (remplace COL_x)
      - merge libellés prudent
      - signe + suppression colonne '-' parasite
      - conversion numérique
      - ✅ réparation du décalage Vie/Total (Annexe 12 typique)
      - ✅ cellules = montants uniquement (le reste -> vide)
    """
    if df is None or df.empty:
        return df

    df = df.copy()

    df = _drop_empty_rows_cols(df)
    df = _drop_useless_top_rows(df, max_drop=12)
    df = _drop_empty_rows_cols(df)

    if df is None or df.empty:
        return df

    # --- Header principal ---
    header_i = _detect_best_header_row(df)
    raw_cols = df.iloc[header_i].tolist()
    raw_cols = [c if _normalize_text_cell(c) else f"COL_{idx+1}" for idx, c in enumerate(raw_cols)]
    raw_cols[0] = "CATEGORIES"

    df = df.iloc[header_i + 1 :].reset_index(drop=True)
    df.columns = _make_unique_columns(raw_cols, force_first_col="CATEGORIES")
    df = _drop_empty_rows_cols(df)

    if df is None or df.empty:
        return df

    # ✅ compléter headers multi-lignes (retours à la ligne)
    df = _fill_placeholder_headers_from_top_rows(df, max_rows=4)

    # merge libellés (prudent)
    df = _merge_wrapped_rows(df)

    # signe + suppression colonne '-' parasite
    df = _propagate_sign_and_drop_dash_cols(df)

    # conversion numérique
    for c in df.columns:
        df[c] = df[c].apply(clean_number)

    # ✅ réparation spécifique mais générale: Vie/Total décalé
    df = _repair_single_branch_shift(df, first_col_name="CATEGORIES")

    # ✅ cellule = montant uniquement (sinon vide)
    df = _force_numeric_cells_only(df)

    df.columns = _make_unique_columns(list(df.columns), force_first_col="CATEGORIES")
    return df




# ---------------------------------------------- Extraction tableaux ----------------------------------------------
def _repair_single_branch_shift(df: pd.DataFrame, first_col_name="CATEGORIES") -> pd.DataFrame:
    """
    Cas fréquent sur Annexe 12: colonnes [CATEGORIES, Vie, Total]
    - Sur la majorité des lignes: Vie == Total
    - Sur certaines lignes: Vie vide mais Total contient la valeur -> décalage d'extraction
    => Si la relation Vie==Total est dominante, on copie Total -> Vie quand Vie est vide.

    Cette règle est générale et ne s'applique que si:
      - exactement 2 colonnes numériques (hors CATEGORIES)
      - et l'égalité est vraie sur une majorité des lignes où les deux sont renseignées.
    """
    if df is None or df.empty:
        return df

    df = df.copy()
    cols = list(df.columns)
    if len(cols) != 3:
        return df
    if cols[0] != first_col_name:
        return df

    c1, c2 = cols[1], cols[2]

    # calculer ratio d'égalité sur lignes où les deux valeurs sont numériques
    eq = 0
    tot = 0
    for i in range(len(df)):
        v1 = df.at[i, c1]
        v2 = df.at[i, c2]
        if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
            tot += 1
            if v1 == v2:
                eq += 1

    # si pas assez de lignes comparables, ne rien faire
    if tot < 5:
        return df

    ratio = eq / tot

    # si majorité forte (ex: Vie == Total dans 70%+), on répare les lignes décalées
    if ratio >= 0.70:
        for i in range(len(df)):
            v1 = df.at[i, c1]
            v2 = df.at[i, c2]
            v1_empty = (v1 == "" or v1 is None)
            v2_num = isinstance(v2, (int, float)) and not isinstance(v2, bool)
            if v1_empty and v2_num:
                df.at[i, c1] = v2

    return df


def _force_numeric_cells_only(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pour toutes les colonnes sauf la 1ère (CATEGORIES):
      - si clean_number() donne un nombre -> on garde
      - sinon -> cellule vide ""
    """
    if df is None or df.empty:
        return df

    df = df.copy()
    cols = list(df.columns)

    for c in cols[1:]:
        df[c] = df[c].apply(lambda x: clean_number(x))
        df[c] = df[c].apply(lambda v: v if isinstance(v, (int, float)) and not isinstance(v, bool) else "")

    return df

def extract_native_pdf(pdf_path, page_num):
    try:
        print(f"\nExtraction tableau page {page_num} (natif)...")

        tables = None
        try:
            tables = camelot.read_pdf(pdf_path, flavor="stream", pages=str(page_num))
        except Exception:
            tables = None

        if tables is None or tables.n == 0:
            try:
                tables = camelot.read_pdf(pdf_path, flavor="stream", pages=str(page_num))
            except Exception:
                tables = None

        if tables is None or tables.n == 0:
            print(f"Aucun tableau page {page_num}")
            return None

        results = []
        for i, table in enumerate(tables, 1):
            try:
                df_raw = table.df
                print(f"DEBUG raw shape = {df_raw.shape}")
                df_clean = clean_table_general(df_raw)
                print(f"DEBUG clean shape = {None if df_clean is None else df_clean.shape}")

                if df_clean is None or df_clean.empty or df_clean.shape[0] < 2 or df_clean.shape[1] < 2:
                    continue
                results.append((f"Tableau_{i}_Page_{page_num}", df_clean))
            except Exception:
                continue

        if not results:
            print("⚠️ Tables détectées mais contenu inutilisable après nettoyage.")
            return None

        print(f"  -> {len(results)} tableau(x) retenu(s)")
        return results

    except Exception as e:
        logging.error(f"Erreur extraction natif page {page_num} : {str(e)}")
        print(f"Erreur extraction natif page {page_num} : {str(e)}")
        return None


def extract_scanned_pdf(pdf_path, page_num):
    """
    OCR page entière puis parsing simple par espaces (stable).
    """
    try:
        print(f"\nExtraction tableau page {page_num} (scanné/OCR)...")

        images = convert_from_path(pdf_path, first_page=page_num, last_page=page_num, dpi=OCR_DPI)
        if not images:
            print(f"Impossible convertir page {page_num}")
            return None

        image = preprocess_for_ocr(images[0])
        text = pytesseract.image_to_string(image, lang=OCR_LANG, config=f"--psm {OCR_PSM}")
        lines = [line.strip() for line in text.split("\n") if line.strip()]

        table_data = []
        for line in lines:
            parts = re.split(r"\s{2,}", line)
            parts = [p.strip() for p in parts if p.strip()]
            if len(parts) >= 2:
                table_data.append(parts)

        if not table_data:
            print(f"Aucun tableau OCR page {page_num}")
            return None

        max_cols = max(len(row) for row in table_data)
        for row in table_data:
            row.extend([""] * (max_cols - len(row)))

        df_raw = pd.DataFrame(table_data)
        df_clean = clean_table_general(df_raw)

        if df_clean is None or df_clean.empty:
            print(f"⚠️ Table OCR nettoyée vide page {page_num}")
            return None

        return [(f"Tableau_OCR_Page_{page_num}", df_clean)]

    except Exception as e:
        logging.error(f"Erreur extraction scanné page {page_num} : {str(e)}")
        print(f"Erreur extraction scanné page {page_num} : {str(e)}")
        return None


# ---------------------------------------------- Export Excel ----------------------------------------------
def _safe_save_workbook(wb: Workbook, path: str) -> str:
    try:
        wb.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        new_path = f"{base}_{now_ts()}{ext}"
        wb.save(new_path)
        return new_path


def export_to_excel(tables, output_name, societe):
    try:
        safe_societe = safe_filename(societe)
        dossier = os.path.join(os.getcwd(), safe_societe)
        os.makedirs(dossier, exist_ok=True)
        chemin_final = os.path.join(dossier, output_name)

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        data_font = Font(name="Arial", size=10)

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for name, df in tables:
            ws = wb.create_sheet(title=str(name)[:31])

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)

                    if isinstance(value, float) and pd.isna(value):
                        cell.value = ""
                    else:
                        cell.value = value

                    cell.font = data_font
                    cell.border = border

                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        # ✅ ne pas centrer les noms de lignes (col 1)
                        cell.alignment = data_alignment_left if c_idx == 1 else data_alignment_center

            # auto width
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value is not None and str(cell.value) != "":
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(60, (max_length + 2) * 1.2)

            for r in range(1, ws.max_row + 1):
                ws.row_dimensions[r].height = 18

        saved_path = _safe_save_workbook(wb, chemin_final)
        print(f"✅ Excel créé : {saved_path}")
        logging.info(f"Fichier Excel créé : {saved_path}")
        return True, saved_path

    except Exception as e:
        logging.error(f"Erreur export : {str(e)}")
        print(f"Erreur export : {str(e)}")
        return False, None


import subprocess
import sys

def run_c_normalisation(excel_annexe_path: str, ann: str = "12", year: int = 2024) -> int:
    """
    Lance C.py en lui passant le chemin Excel (ex: 12E2024.xlsx).
    Attend la fin, puis ouvre le NV généré (12NV2024.xlsx) dans Excel 2010.
    """
    try:
        excel_annexe_path = os.path.abspath(excel_annexe_path)
        folder = os.path.dirname(excel_annexe_path)

        cmd = [sys.executable, str(NV12_SCRIPT), excel_annexe_path]
        print(f"➡️ Lancement C.py : {' '.join(cmd)}")
        res = subprocess.run(cmd, capture_output=False)
        rc = int(res.returncode or 0)

        # Ouvrir automatiquement le fichier NV (si créé)
        nv_path = _find_latest_nv_file(folder, ann=str(ann), year=int(year))
        if nv_path and os.path.exists(nv_path):
            opened = _open_in_excel_2010(nv_path)
            if not opened:
                print(f"⚠️ Impossible d'ouvrir automatiquement: {nv_path} (ouvre-le manuellement).")
        else:
            print("⚠️ Aucun fichier NV trouvé après exécution de C.py.")

        return rc

    except Exception as e:
        print(f"❌ Erreur lancement C.py : {e}")
        logging.error(f"Erreur lancement C.py : {e}")
        return 1


def run_b_processing(excel_annexe13_path: str) -> int:
    """
    Lance B.py en lui passant le chemin de l'Excel Annexe 13 (ex: 13E2024.xlsx).
    Retourne le code de retour de B.py.
    """
    try:
        excel_annexe13_path = os.path.abspath(excel_annexe13_path)

        if not os.path.exists(excel_annexe13_path):
            print(f"⚠️ Annexe 13 introuvable: {excel_annexe13_path} -> B.py non lancé.")
            return 1

        cmd = [sys.executable, str(NV13_SCRIPT), excel_annexe13_path]
        print(f"➡️ Lancement B.py : {' '.join(cmd)}")
        res = subprocess.run(cmd, capture_output=False)
        return int(res.returncode or 0)

    except Exception as e:
        print(f"❌ Erreur lancement B.py : {e}")
        logging.error(f"Erreur lancement B.py : {e}")
        return 1


# ---------------------------------------------- Main ----------------------------------------------
def main():
    start_time = time.time()
    logging.info(f"Démarrage script - {time.strftime('%H:%M:%S')}")
    print(f"\n=== Démarrage {time.strftime('%H:%M:%S')} ===")

    # 1) Télécharger le PDF ciblé (société + année + type document)
    pdf_path, pdf_url, pdf_nom_reel = fetch_pdf_for_societe_annee(SOCIETE, ANNEE, DOC_NAMES_ACCEPTES)

    # 2) DB
    connection, cursor = create_cmf_database_and_table()
    if not connection or not cursor:
        logging.error("Échec connexion base")
        print("Échec connexion base")
        return

    try:
        # 3) Insérer l'info du PDF téléchargé
        if pdf_url and pdf_nom_reel:
            insert_pdf_info_cmf(connection, cursor, SOCIETE, pdf_nom_reel, ANNEE, pdf_url)

        # 4) Stop si pas de PDF
        if not (pdf_path and os.path.exists(pdf_path)):
            print("PDF non disponible → extraction annulée")
            logging.warning("PDF non disponible")
            return

        print(f"\n=== Analyse PDF {SOCIETE} {ANNEE} ===")
        sections_found = search_sections_in_pdf(pdf_path)

        if not sections_found:
            print("Aucune annexe détectée (12/13).")
            logging.warning("Aucune annexe détectée (12/13).")
            return

        annexe12_export_path = None  # 12E2024.xlsx
        annexe13_export_path = None  # 13E2024.xlsx

        # 5) Extraction + export pour chaque annexe détectée
        for section_key, (page_num, is_scanned) in sections_found.items():
            section_display = {"Annexe_12": "Annexe 12", "Annexe_13": "Annexe 13"}.get(section_key, section_key)
            print(f"\n--- {section_display} (page {page_num}) ---")
            logging.info(f"{section_display} page {page_num}")

            # extraction table
            if is_scanned:
                tables = extract_scanned_pdf(pdf_path, page_num)
            else:
                tables = extract_native_pdf(pdf_path, page_num)
                # fallback OCR si natif échoue
                if not tables:
                    print("⚠️ Extraction natif vide -> tentative OCR fallback...")
                    tables = extract_scanned_pdf(pdf_path, page_num)

            if not tables:
                print(f"❌ Aucun tableau exploitable pour {section_display}")
                logging.warning(f"Aucun tableau exploitable pour {section_display}")
                continue

            # ✅ Noms courts demandés
            if section_key == "Annexe_12":
                output_name = f"12E{ANNEE}.xlsx"
            elif section_key == "Annexe_13":
                output_name = f"13E{ANNEE}.xlsx"
            else:
                output_name = f"{section_key}_E{ANNEE}.xlsx"

            ok, saved_path = export_to_excel(tables, output_name, SOCIETE)

            if not ok or not saved_path:
                print(f"❌ Échec export {section_display}")
                logging.error(f"Échec export {section_display}")
                continue

            print(f"✅ Export réussi: {saved_path}")
            logging.info(f"Export réussi: {saved_path}")

            # mémoriser chemins
            if section_key == "Annexe_12":
                annexe12_export_path = saved_path
            elif section_key == "Annexe_13":
                annexe13_export_path = saved_path

        # 6) Lancer C.py seulement sur Annexe 12 (si extrait)
        if annexe12_export_path and os.path.exists(annexe12_export_path):
            rc_c = run_c_normalisation(annexe12_export_path)
            print(f"✅ C.py (Annexe 12) terminé avec code retour = {rc_c}")
            logging.info(f"C.py (Annexe 12) terminé avec code retour = {rc_c}")

            # ✅ Scénario demandé : si Annexe 12 validée (rc==0) => lancer B.py sur Annexe 13
            if rc_c == 0:
                if annexe13_export_path and os.path.exists(annexe13_export_path):
                    rc_b = run_b_processing(annexe13_export_path)
                    print(f"✅ B.py (Annexe 13) terminé avec code retour = {rc_b}")
                    logging.info(f"B.py (Annexe 13) terminé avec code retour = {rc_b}")
                else:
                    print("ℹ️ Annexe 13 non exportée -> B.py non lancé.")
                    logging.info("Annexe 13 non exportée -> B.py non lancé.")
            else:
                print("ℹ️ Annexe 12 non validée -> B.py non lancé.")
                logging.info("Annexe 12 non validée -> B.py non lancé.")

        else:
            print("ℹ️ Annexe 12 non exportée -> C.py non lancé.")
            logging.info("Annexe 12 non exportée -> C.py non lancé.")

        print(f"\n=== Terminé en {time.time() - start_time:.2f} s ===")
        logging.info(f"Script terminé en {time.time() - start_time:.2f} s")

    except Exception as e:
        logging.error(f"ERREUR GLOBALE : {str(e)}")
        print(f"\n=== ERREUR GLOBALE : {str(e)} ===")

    finally:
        try:
            if connection and connection.is_connected():
                cursor.close()
                connection.close()
                logging.info("Connexion MySQL fermée")
                print("\n=== MySQL fermé ===")
        except Exception:
            pass



def run_for(societe: str, annee: int) -> int:
    global SOCIETE, ANNEE
    SOCIETE = societe
    ANNEE = int(annee)
    return int(main() or 0)



if __name__ == "__main__":
    import sys

    # Usage: python Extraction1213.py "<societe>" 2024
    if len(sys.argv) >= 3:
        SOCIETE = sys.argv[1]
        ANNEE = int(sys.argv[2])

    sys.exit(int(main() or 0))


