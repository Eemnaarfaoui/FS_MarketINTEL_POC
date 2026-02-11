# Configuration et Imports
import os
import re
import time
import mysql.connector
from mysql.connector import Error
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse, urlencode, parse_qs
from datetime import datetime
import PyPDF2
import requests
import pandas as pd
import camelot
import fitz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
import logging
import pyodbc

# -----------------------------------------------------Partie 1 : Configuration des logs ----------------------------------------------------------------------
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"   # Ne montrer aucun message, sauf les erreurs graves (TensorFlow )
os.environ["GLOG_minloglevel"] = "3"       # Similaire, mais elle concerne une autre bibliothèque appelée (GLOG)
os.environ["PYTHONWARNINGS"] = "ignore"
logging.basicConfig(filename='script.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')  #configure le système de journaux (logs) de Python

# *********** Fixation sur la société STAR : ************
societes_assurances = [
    "Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR -"
]

# -----------------------------------------------------Partie 2 : Gestion Base de données ---------------------------------------------------------------------
def create_cmf_database_and_table():
    try:
        # Connect to SQL Server (Windows Auth)
        connection = pyodbc.connect(
            'DRIVER={ODBC Driver 18 for SQL Server};'
            'SERVER=localhost;'
            'Trusted_Connection=yes;'
            'Encrypt=no;',
            autocommit=True  # Disable encryption to avoid SSL certificate issues
        )
        cursor = connection.cursor()
        
        logging.info("Création/mise à jour de la base de données 'cmf'...")
        print("Création/mise à jour de la base de données 'cmf'...")
        
        # 1️⃣ Create database if it doesn't exist
        cursor.execute("""
        IF DB_ID('cmf') IS NULL
            CREATE DATABASE cmf
        """)
        
        # Switch to the database
        cursor.execute("USE cmf")
        
        # 2️⃣ Create table if it doesn't exist
        cursor.execute("""
        IF OBJECT_ID('document', 'U') IS NULL
        CREATE TABLE document (
            id INT IDENTITY(1,1) PRIMARY KEY,
            Societe NVARCHAR(255) NOT NULL,
            Nom NVARCHAR(255) NOT NULL,
            Annee INT NOT NULL,
            URL NVARCHAR(512) NOT NULL,
            CONSTRAINT unique_document UNIQUE (Societe, Nom, Annee)
        )
        """)
        
        # Commit changes
        connection.commit()
        
        logging.info("Base de données 'cmf' et table 'document' prêtes.")
        print("Base de données 'cmf' et table 'document' prêtes.")
        
        # Return connection and cursor in case you want to use them
        return connection, cursor

    except Exception as e:
        logging.error(f"Erreur lors de la création de la base : {e}")
        print(f"Erreur lors de la création de la base : {e}")
        return None, None
    
# -----------------------------------------------------Partie 3 : Fonctions à utiliser dans la logique du main code --------------------------------------------------------
    


    # *********** Fonction N°1(Appel FN°7-8): Récupération pdf 2024 de STAR depuis CMF après scraping : ************

def fetch_star_2024_pdf():
    societe = "Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR -"
    nom = "Etats financiers au 31/12"
    annee = 2024
    
    chrome_options = Options()  # crée un objet chrome_options qui permet de configurer un navigateur web automatisé (Google Chrome). 
    chrome_options.add_argument("--headless=new")  # sans interface visible
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    service = Service(ChromeDriverManager().install(), log_path=os.devnull)
    
    driver = webdriver.Chrome(service=service, options=chrome_options)
    try:
        driver.get("https://www.cmf.tn/consultation-des-tats-financier-des-soci-t-s-faisant-ape")
        pdfs = extract_pdfs_from_page(driver, societe)
        
        for pdf in pdfs:
            if pdf["nom"] == nom and int(pdf["annee"]) == annee:
                pdf_path = download_pdf(pdf["url"], pdf["societe"], pdf["nom"], pdf["annee"])
                if pdf_path:
                    logging.info(f"PDF Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024 téléchargé avant la création de la base: {pdf_path}")
                    print(f"PDF Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024 téléchargé avant la création de la base: {pdf_path}")
                    return pdf_path, pdf["url"]
                else:
                    logging.error("Échec du téléchargement du PDF Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024")
                    print("Échec du téléchargement du PDF Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024")
                    return None, None
        logging.warning("Document Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024 non trouvé sur le site CMF")
        print("Document Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024 non trouvé sur le site CMF")
        return None, None
    except Exception as e:
        logging.error(f"ERREUR lors de la récupération du PDF Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024: {str(e)}")
        print(f"ERREUR lors de la récupération du PDF Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024: {str(e)}")
        return None, None
    finally:
        driver.quit()

 # *********** Fonction N°2(Appel FN°5-4): La fonction d'ajout du doc dans la base : ************

def insert_pdf_info_cmf(connection, cursor, societe, nom_document, annee, url):
    try:
        normalized_url = normalize_url(url)
        
        try:
            annee_int = int(annee)
            if not (2015 <= annee_int <= 2025):
                logging.info(f"Document {nom_document} ({annee}) ignoré (hors période 2015-2025).")
                print(f"Document {nom_document} ({annee}) ignoré (hors période 2015-2025).")
                return False
        except ValueError:
            logging.error(f"Année invalide pour {nom_document}: {annee}")
            print(f"Année invalide pour {nom_document}: {annee}")
            return False
            
        if check_document_exists(cursor, societe, nom_document, annee_int):
            logging.info(f"Document {nom_document} ({annee}) pour {societe} existe déjà.")
            print(f"Document {nom_document} ({annee}) pour {societe} existe déjà.")
            return False
            
        insert_query = """
        INSERT INTO document (Societe, Nom, Annee, URL)
        VALUES (?, ?, ?, ?)
        """
        cursor.execute(insert_query, (societe, nom_document, annee_int, normalized_url))
        connection.commit()
        logging.info(f"AJOUTÉ: {nom_document} ({annee}) pour {societe}")
        print(f"AJOUTÉ: {nom_document} ({annee}) pour {societe}")
        return True
        
    except Error as e:
        logging.error(f"Erreur lors de l'insertion pour {societe} : {e}")
        print(f"Erreur lors de l'insertion pour {societe} : {e}")
        return False
    
# *********** Fonction N°3 (Appel FN°9-13-14-15): Recherche le document STAR 2024 et extrait automatiquement le tableau de l'annexe 13: ************


def search_document(connection, cursor):
    logging.info("Recherche de l'annexe 13 pour Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024")
    print("\n=== Recherche de l'annexe 13 ===")
    
    try:
        societe = "Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR -"
        nom = "Etats financiers au 31/12"
        annee = 2024
        
        # Vérifier si le document existe dans la base
        query = """
        SELECT id, Societe, Nom, Annee, URL 
        FROM document 
        WHERE Societe = ? 
        AND Nom = ?
        AND Annee = ?
        """
        cursor.execute(query, (societe, nom, annee))
        result = cursor.fetchone()
        
        if result:
            doc_id, doc_societe, doc_nom, doc_annee, doc_url = result
            logging.info(f"Document trouvé - ID: {doc_id}, Société: {doc_societe}, Nom: {doc_nom}, Année: {doc_annee}, URL: {doc_url}")
            print(f"\n=== Document trouvé ===\nID: {doc_id}\nSociété: {doc_societe}\nNom: {doc_nom}\nAnnée: {doc_annee}\nURL: {doc_url}")
            
            # Vérifier si le PDF existe localement
            safe_societe = re.sub(r'[^\w\s-]', '_', doc_societe).replace(' ', '_')
            safe_nom = re.sub(r'[^\w\s-]', '_', doc_nom).replace(' ', '_')
            pdf_path = os.path.join(os.getcwd(), f"{safe_societe}_{safe_nom}_{doc_annee}.pdf")
            if not os.path.exists(pdf_path):
                logging.warning("PDF Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024 non trouvé localement, il devrait avoir été téléchargé avant")
                print("PDF Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024 non trouvé localement, il devrait avoir été téléchargé avant")
                return
            
            page_num, is_scanned = search_annexe13_in_pdf(pdf_path)
            if page_num:
                logging.info(f"Annexe 13 trouvée à la page {page_num} (ID: {doc_id})")
                print(f"\n=== Résultat ===\nL'annexe 13 a été trouvée à la page {page_num} (ID: {doc_id})")
                
                # Extraire automatiquement le tableau
                if is_scanned:
                    tables = extract_scanned_pdf(pdf_path, page_num)
                else:
                    tables = extract_native_pdf(pdf_path, page_num)
                
                if tables:
                    output_name = f"{safe_societe}_Annexe13_{doc_annee}_extracted.xlsx"
                    if export_to_excel(tables, output_name):
                        logging.info(f"Fichier Excel créé : {os.path.abspath(output_name)}")
                        print(f"\nFichier Excel créé : {os.path.abspath(output_name)}")

#------------------------------------------------ Appel au code B avec le fichier Excel généré------------------------------------------------------------
                        import subprocess
                        try:
                            subprocess.run(["python", "B.py", os.path.abspath(output_name)], check=True)
                            logging.info(f"Code B exécuté avec le fichier {output_name}")
                            print(f"Code B exécuté avec le fichier {output_name}")
                        except subprocess.CalledProcessError as e:
                            logging.error(f"Erreur lors de l'exécution du code B : {str(e)}")
                            print(f"Erreur lors de l'exécution du code B : {str(e)}")
                    else:
                        logging.error("Échec de l'exportation vers Excel")
                        print("\nÉchec de l'exportation vers Excel")
                else:
                    logging.warning("Aucun tableau détecté à la page de l'annexe 13")
                    print("\nAucun tableau détecté à la page de l'annexe 13")
            else:
                logging.warning(f"L'annexe 13 n'a pas été trouvée dans le document ID {doc_id}")
                print(f"\n=== Résultat ===\nL'annexe 13 n'a pas été trouvée dans le document ID {doc_id}")
        else:
            logging.error("Aucun document trouvé pour Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024")
            print("\nAucun document trouvé avec ces critères pour Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR - 2024")
            
    except Exception as e:
        logging.error(f"ERREUR LORS DE LA RECHERCHE : {str(e)}")
        print(f"\n=== ERREUR LORS DE LA RECHERCHE : {str(e)} ===")





# *********** Fonction N°4 :S'assurer que le document existe dans la base cmf : ************

def check_document_exists(cursor, societe, nom, annee):
    try:
        query = """
        SELECT COUNT(*) FROM document 
        WHERE Societe = ? 
        AND Nom = ?
        AND Annee = ?
        """
        cursor.execute(query, (societe, nom, annee))
        return cursor.fetchone()[0] > 0
    except Error as e:
        logging.error(f"Erreur lors de la vérification : {e}")
        print(f"Erreur lors de la vérification : {e}")
        return False





# ***********  Fonction N°5: Extraction des URLs de PDFs et leurs normalisation en supprimant les paramètres inutiles : ************

def normalize_url(url):
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    if 'id' in query_params or 'token' in query_params:
        query_params.pop('id', None)
        query_params.pop('token', None)
    new_query = urlencode(query_params, doseq=True)
    return parsed_url._replace(query=new_query).geturl()


# ***********  Fonction N°6: Extrait l'année à partir d'un URL, en priorisant 2015-2025. : ************

def extract_year_from_text(text):
    patterns = [
        r'(?:20)(1[5-9]|2[0-5])\b',  # 2015-2025
        r'\b(20\d{2})\b',            # 4-digit years starting with 20
        r'(?:3112|1312)(1[5-9]|2[0-5])',  # Specific patterns like 3112 or 1312
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            year = match.group(1) if match.group(1) else match.group(0)
            if len(year) == 2:
                year = f"20{year}"
            if 2015 <= int(year) <= 2025:
                return year
    return None

# ***********  Fonction N°7(Appel FN°6): Extraction des pdf pour une société donnée : ************

def extract_pdfs_from_page(driver, societe):
    """Commencer par Extraire leurs URLs"""
    wait = WebDriverWait(driver, 20)
    pdfs = []
    
    try:
        logging.info(f"Traitement de {societe}")
        print(f"\n--- Traitement de {societe} ---")
        
        dropdown = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#edit_field_societesape_value_chosen a")))
        dropdown.click()
        time.sleep(1)
        
        options = wait.until(EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "#edit_field_societesape_value_chosen .chosen-results li")))
        
        selected_option = None
        for option in options:
            if societe.lower() == option.text.strip().lower():
                selected_option = option
                break
        
        if not selected_option:
            logging.error(f"Aucune correspondance exacte trouvée pour '{societe}'")
            print(f"ERREUR: Aucune correspondance exacte trouvée pour '{societe}'")
            return []
        
        selected_option.click()
        logging.info(f"Sélectionnée: {selected_option.text.strip()}")
        print(f"Sélectionnée: {selected_option.text.strip()}")

        submit = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "#edit-submit-consultation-des-tats-financier-des-soci-t-s-faisant-ape")))
        submit.click()
        time.sleep(2)

        page_num = 1
        while True:
            logging.info(f"Traitement page {page_num}")
            print(f"  Traitement page {page_num}...")
            rows = wait.until(EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, ".view-content .views-row")))
            
            for row in rows:
                try:
                    pdf_url = row.find_element(By.CSS_SELECTOR, "a[href$='.pdf']").get_attribute("href")
                    pdf_name = row.find_element(
                        By.CSS_SELECTOR, ".field-name-field-p-riode .field-item").text.strip()
                    
                    year = extract_year_from_text(pdf_url) or extract_year_from_text(pdf_name)
                    
                    if year and year.isdigit():
                        pdfs.append({
                            "url": pdf_url,
                            "nom": pdf_name,
                            "annee": year,
                            "societe": societe
                        })
                        logging.info(f"Trouvé: {pdf_name} ({year})")
                        print(f"  Trouvé: {pdf_name} ({year})")

                except (NoSuchElementException, StaleElementReferenceException) as e:
                    logging.warning(f"Erreur mineure lors du traitement d'une ligne: {str(e)}")
                    continue

            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, ".pager .next a:not(.disabled)")
                next_btn.click()
                page_num += 1
                time.sleep(3)
            except:
                break

    except Exception as e:
        logging.error(f"ERREUR lors de l'extraction pour {societe}: {str(e)}")
        print(f"ERREUR lors de l'extraction pour {societe}: {str(e)}")
    
    return pdfs







# ***********  Fonction N°8: Téléchargement des PDFs dans le dossier du projet : ************

def download_pdf(url, societe, nom, annee):
    try:
        # Nettoyer les caractères invalides pour le nom du fichier
        safe_societe = re.sub(r'[^\w\s-]', '_', societe).replace(' ', '_')
        safe_nom = re.sub(r'[^\w\s-]', '_', nom).replace(' ', '_')
        filename = f"{safe_societe}_{safe_nom}_{annee}.pdf"
        filepath = os.path.join(os.getcwd(), filename)
        
        # Créer le répertoire ken moch mawjoud
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        response = requests.get(url, stream=True, timeout=30)
        response.raise_for_status()
        with open(filepath, 'wb') as f:
            f.write(response.content)
        logging.info(f"PDF téléchargé : {filepath}")
        print(f"PDF téléchargé : {filepath}")
        return filepath
    except Exception as e:
        logging.error(f"Erreur lors du téléchargement du PDF {url}: {str(e)}")
        print(f"Erreur lors du téléchargement du PDF {url}: {str(e)}")
        return None

# ***********  Fonction N°9 :  Recherche l'annexe 13 dans un PDF et retourne le numéro de page et le type du pdf (natif/scanné) : ************

def search_annexe13_in_pdf(pdf_path):
    try:
        start_time = time.time()
        logging.info(f"Analyse du PDF: {pdf_path}")
        print(f"\nAnalyse du PDF: {pdf_path}")
        
                               # Étape 1 : Essayer l'extraction de texte natif
        pdf_reader = PyPDF2.PdfReader(pdf_path)
        num_pages = len(pdf_reader.pages)
        
        logging.info(f"Nombre de pages: {num_pages}")
        print(f"Nombre de pages: {num_pages}")
        
        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            
            if text:
                text = text.lower()
                text = re.sub(r'\s+', ' ', text)
                if any(keyword in text for keyword in ["annexe 13", "annexe13", "annexe n° 13", "annexe n°13", "note n°13"]):
                    logging.info(f"Annexe 13 trouvée à la page {page_num + 1} (PDF natif)")
                    print(f"Annexe 13 trouvée à la page {page_num + 1} (PDF natif)")
                    print(f"Temps de traitement: {time.time() - start_time:.2f} secondes")
                    return page_num + 1, False  # Si False donc PDF natif
        
                                  # Étape 2 : Essayer OCR pour les PDFs scannés (Pas encore fonctionnel)
        logging.info("Aucun texte natif trouvé, tentative d'extraction OCR...")
        print("Aucun texte natif trouvé, tentative d'extraction OCR...")
        images = convert_from_path(pdf_path, dpi=100)
        
        for page_num, image in enumerate(images, 1):
            image = image.resize((int(image.width * 0.25), int(image.height * 0.25)), Image.Resampling.LANCZOS)
            text = pytesseract.image_to_string(image, lang='fra', config='--psm 6')
            if text:
                text = text.lower()
                text = re.sub(r'\s+', ' ', text)
                if any(keyword in text for keyword in ["annexe 13", "annexe13", "annexe n° 13", "annexe n°13", "note n°13"]):
                    logging.info(f"Annexe 13 trouvée à la page {page_num} (PDF scanné via OCR)")
                    print(f"Annexe 13 trouvée à la page {page_num} (PDF scanné via OCR)")
                    print(f"Temps de traitement: {time.time() - start_time:.2f} secondes")
                    return page_num, True  # ken True donc PDF scanné
        
        logging.warning("Annexe 13 non trouvée dans le document")
        print("Annexe 13 non trouvée dans le document")
        print(f"Temps de traitement: {time.time() - start_time:.2f} secondes")
        return None, None
            
    except Exception as e:
        logging.error(f"Erreur lors de l'analyse du PDF: {str(e)}")
        print(f"Erreur lors de l'analyse du PDF: {str(e)}")
        return None, None

# ***********  Fonction N°10:  Fonction qui vérifie si le pdf est scanné (pas encore utilisée) : ************

def is_scanned_pdf(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            if page.get_text().strip():
                return False
        return True
    except Exception:
        return True


# ***********  Fonction N°11:  Cette fonction c'est pour forcer que le type des cellule soit numérique et non pas une chaine: ************

def clean_number(text):
    """Converts French-formatted numbers (Par exemple: '1 234,56') to int, forcing integer type."""
    if isinstance(text, str):
        # Remove spaces, replace comma with dot, remove non-numeric suffixes
        cleaned = re.sub(r'\s+', '', text).replace(',', '.').rstrip('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ')
        try:
            # Convert directly to int after handling decimal
            return int(float(cleaned))
        except ValueError:
            return text
    return int(text) if isinstance(text, (int, float)) else text


# *********** Fonction N°12:(PDF natifs) convertit les cellules en int : ************

def clean_native_table(df):
    """Cleans the table from native PDFs and converts numbers to int type."""
    for i in range(len(df)):
        if sum(1 for cell in df.iloc[i] if str(cell).strip()) >= 2:
            df.columns = df.iloc[i]
            df = df.iloc[i+1:].reset_index(drop=True)
            break
    # Convert columns to int where possible
    for col in df.columns:
        df[col] = df[col].apply(clean_number)
        if df[col].apply(lambda x: isinstance(x, (int, float))).all():
            df[col] = pd.to_numeric(df[col], errors='coerce', downcast='integer').astype('Int64')  # Use Int64 to handle NaN
    return df


# *********** Fonction N°13: Extrait les tableaux d'une page spécifique des PDF natifs : ************

def extract_native_pdf(pdf_path, page_num):
    try:
        logging.info(f"Extraction du tableau de la page {page_num} (PDF natif)")
        print(f"\n Extraction du tableau de la page {page_num}...")
        tables = camelot.read_pdf(pdf_path, flavor='stream', pages=str(page_num))
        
        if tables.n == 0:
            logging.warning(f"Aucun tableau détecté à la page {page_num}")
            print(f"\n Aucun tableau détecté à la page {page_num}")
            return None
            
        results = []
        for i, table in enumerate(tables, 1):
            logging.info(f"Tableau {i} détecté à la page {page_num}")
            print(f"\n Tableau {i} détecté à la page {page_num}")
            cleaned_df = clean_native_table(table.df)
            results.append((f"Tableau_{i}_Page_{page_num}", cleaned_df))
        
        return results
        
    except Exception as e:
        logging.error(f"Erreur d'extraction à la page {page_num}: {str(e)}")
        print(f"\n Erreur d'extraction à la page {page_num}: {str(e)}")
        return None



# *********** Fonction N°14: Extrait les tableaux d'une page spécifique des PDF SCANNEs et convertit les nbr en int : ************

def extract_scanned_pdf(pdf_path, page_num):
    try:
        logging.info(f"Extraction du tableau de la page {page_num} (PDF scanné)")
        print(f"\n🔍 Extraction du tableau de la page {page_num} (PDF scanné)...")
        start_time = time.time()
        images = convert_from_path(pdf_path, first_page=page_num, last_page=page_num, dpi=100)
        
        if not images:
            logging.error(f"Impossible de convertir la page {page_num} en image")
            print(f"\n Impossible de convertir la page {page_num} en image")
            return None
        
        image = images[0]
        image = image.resize((int(image.width * 0.25), int(image.height * 0.25)), Image.Resampling.LANCZOS)
        text = pytesseract.image_to_string(image, lang='fra', config='--psm 6')
        
        if not text.strip():
            logging.warning(f"Aucun texte détecté à la page {page_num} via OCR")
            print(f"\n Aucun texte détecté à la page {page_num} via OCR")
            return None
        
        debug_file = f"ocr_debug_page_{page_num}.txt"
        with open(debug_file, 'w', encoding='utf-8') as f:
            f.write(text)
        logging.info(f"Texte brut OCR sauvegardé dans : {os.path.abspath(debug_file)}")
        print(f"\n Texte brut OCR sauvegardé dans : {os.path.abspath(debug_file)}")
        
        text = re.sub(r'\s*\n\s*', '\n', text.strip())
        lines = text.split('\n')
        table_data = []
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            columns = re.split(r'\s{2,}', line)
            if len(columns) > 1:
                cleaned_columns = [clean_number(col.strip()) for col in columns if col.strip()]
                if len(cleaned_columns) >= 2:
                    table_data.append(cleaned_columns)
                elif table_data:
                    table_data[-1].extend(cleaned_columns)
        
        if table_data:
            max_cols = max(len(row) for row in table_data)
            for row in table_data:
                while len(row) < max_cols:
                    row.append('')
            df = pd.DataFrame(table_data)
            # Convert columns to int where possible
            for col in df.columns:
                df[col] = df[col].apply(clean_number)
                if df[col].apply(lambda x: isinstance(x, (int, float))).all():
                    df[col] = pd.to_numeric(df[col], errors='coerce', downcast='integer').astype('Int64')  # Use Int64 to handle NaN
            logging.info(f"Tableau construit à partir du texte OCR à la page {page_num} ({len(table_data)} lignes, {max_cols} colonnes)")
            print(f"\n Tableau construit à partir du texte OCR à la page {page_num} ({len(table_data)} lignes, {max_cols} colonnes)")
            print(f"Temps d'extraction: {time.time() - start_time:.2f} secondes")
            return [(f"Tableau_1_Page_{page_num}", df)]
        
        logging.warning(f"Aucun tableau détecté à la page {page_num} via OCR")
        print(f"\n Aucun tableau détecté à la page {page_num} via OCR")
        return None
        
    except Exception as e:
        logging.error(f"Erreur d'extraction à la page {page_num}: {str(e)}")
        print(f"\n Erreur d'extraction à la page {page_num}: {str(e)}")
        return None


# *********** Fonction N°15: Exportation Tableau extrait vers Excel avec format numérique  : ************

def export_to_excel(tables, output_name):
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        data_font = Font(name='Arial', size=10)
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for name, df in tables:
            ws = wb.create_sheet(title=name[:31])
            
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    # Forcer la conversion en entier sans décimales
                    try:
                        numeric_value = float(str(value).replace(' ', '').replace(',', '.'))
                        cell.value = int(numeric_value)  # Forcer entier
                        cell.number_format = '0'  # Format entier strict sans décimales
                    except (ValueError, TypeError):
                        cell.value = value
                    cell.font = data_font
                    cell.border = border
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
            
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
            
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 20
        
        wb.save(output_name)
        logging.info(f"Fichier Excel créé : {os.path.abspath(output_name)}")
        print(f"\nFichier Excel créé avec en-têtes centrés : {os.path.abspath(output_name)}")
        return True
    except Exception as e:
        logging.error(f"Erreur d'export : {str(e)}")
        print(f"\nErreur d'export : {str(e)}")
        return False
    


# *********** Fonction N°16: Traitement d'une sté spécifique pour extraire et enregistrer ses PDFs : ************

def process_societe(societe, connection, cursor):
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    service = Service(ChromeDriverManager().install(), log_path=os.devnull)
    
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.get("https://www.cmf.tn/consultation-des-tats-financier-des-soci-t-s-faisant-ape")
    
    pdfs = extract_pdfs_from_page(driver, societe)
    inserted_count = 0
    
    for pdf in pdfs:
        if insert_pdf_info_cmf(connection, cursor, pdf["societe"], pdf["nom"], pdf["annee"], pdf["url"]):
            inserted_count += 1
    
    driver.quit()
    return inserted_count


# *********** Fonction N°17: Vérifie les documents manquants : ************

def check_missing_documents(connection, cursor):
    try:
        logging.info("Vérification des documents manquants")
        print("\n=== Vérification des documents manquants ===")
        total_inserted = 0
        
        # Prioriser la société STAR 
        prioritized_societe = "Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR -"
        societes = [prioritized_societe]
        
        for societe in societes:
            logging.info(f"Vérification pour la société: {societe}")
            print(f"\n>>> Vérification pour la société: {societe}")
            
            cursor.execute("SELECT Annee FROM document WHERE Societe = %s", (societe,))
            existing_years = {row[0] for row in cursor.fetchall()}
            
            missing_years = set(range(2015, 2026)) - existing_years
            
            if missing_years:
                logging.info(f"Documents manquants pour {societe}: Années {sorted(missing_years)}")
                print(f"Documents manquants pour {societe}: Années {sorted(missing_years)}")
                
                chrome_options = Options()
                chrome_options.add_argument("--headless=new")
                chrome_options.add_argument("--no-sandbox")
                chrome_options.add_argument("--disable-dev-shm-usage")
                chrome_options.add_argument("--disable-gpu")
                service = Service(ChromeDriverManager().install(), log_path=os.devnull)
                
                driver = webdriver.Chrome(service=service, options=chrome_options)
                driver.get("https://www.cmf.tn/consultation-des-tats-financier-des-soci-t-s-faisant-ape")
                
                pdfs = extract_pdfs_from_page(driver, societe)
                inserted_count = 0
                
                for pdf in pdfs:
                    if int(pdf["annee"]) in missing_years:
                        if insert_pdf_info_cmf(connection, cursor, pdf["societe"], pdf["nom"], pdf["annee"], pdf["url"]):
                            inserted_count += 1
                
                driver.quit()
                total_inserted += inserted_count
                logging.info(f"Résultat pour {societe}: {inserted_count} documents manquants ajoutés")
                print(f"=== Résultat pour {societe}: {inserted_count} documents manquants ajoutés ===")
            else:
                logging.info(f"Aucun document manquant pour {societe} (2015-2025)")
                print(f"Aucun document manquant pour {societe} (2015-2025)")
            
            time.sleep(2)
        
        logging.info(f"TOTAL DOCUMENTS MANQUANTS AJOUTÉS: {total_inserted}")
        print(f"\n=== TOTAL DOCUMENTS MANQUANTS AJOUTÉS: {total_inserted} ===")
        return total_inserted
        
    except Exception as e:
        logging.error(f"ERREUR LORS DE LA VÉRIFICATION DES DOCUMENTS MANQUANTS : {str(e)}")
        print(f"\n=== ERREUR LORS DE LA VÉRIFICATION DES DOCUMENTS MANQUANTS : {str(e)} ===")
        return 0


# -----------------------------------------------------Partie 4 : Partie code main d'exécution --------------------------------------------------------

def main():
    start_time = time.time()
    logging.info(f"Démarrage du script à {time.strftime('%H:%M:%S')}")
    print(f"\n=== Démarrage à {time.strftime('%H:%M:%S')} ===")
    

     # *********** Etape 1: (Fonction N°1) Récupération pdf 2024 de STAR depuis CMF après scraping et avant création de la base : ************

    pdf_path, star_url = fetch_star_2024_pdf()
    star_societe = "Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR -"
    star_nom = "Etats financiers au 31/12"
    star_annee = 2024



     # *********** Etape 2: (Partie 2)  Gestion Base de données : ************

    connection, cursor = create_cmf_database_and_table()
    if not connection or not cursor:
        logging.error("Échec de la connexion à la base de données")
        print("Échec de la connexion à la base de données")
        return
    


    # *********** Étape 3: (Fonction N°2-16-17) La fonction d'ajout du doc dans la base (Si inexistant) 
    #                                           +  Traitement d'une sté spécifique pour extraire et enregistrer ses PDFs : ************

    if star_url:
        insert_pdf_info_cmf(connection, cursor, star_societe, star_nom, star_annee, star_url)
        
    
    try:
        # Vérifier si la base est vide
        cursor.execute("SELECT COUNT(*) FROM document")
        count = cursor.fetchone()[0]
        
        if count == 0:
            logging.info("Base de données vide, début de l'alimentation")
            print("\n=== Début de l'alimentation de la base ===")
            total_inserted = 0
            
            #-------- Prioriser Société STAR pour l'alimentation ------------
            prioritized_societe = "Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR -"
            societes = [prioritized_societe]
            
            for societe in societes:
                logging.info(f"Traitement de la société: {societe}")
                print(f"\n>>> Traitement de la société: {societe}")
                inserted = process_societe(societe, connection, cursor)
                total_inserted += inserted
                logging.info(f"Résultat pour {societe}: {inserted} documents ajoutés")
                print(f"=== Résultat pour {societe}: {inserted} documents ajoutés ===")
                time.sleep(2)
            
            logging.info(f"TOTAL: {total_inserted} documents ajoutés")
            print(f"\n=== TOTAL: {total_inserted} documents ajoutés ===")
        else:
            logging.info(f"La base contient déjà {count} documents, vérification des documents manquants")
            print(f"\nLa base contient déjà {count} documents.")
            check_missing_documents(connection, cursor)
        


         # *********** Étape 4: (Fonction N°3) Recherche le document STAR 2024 et extrait automatiquement le tableau de l'annexe 13 : ************

        # Rechercher et traiter le document STAR 2024
        search_document(connection, cursor)
        
        logging.info(f"Script terminé en {time.time() - start_time:.2f} secondes")
        print(f"\n=== Script terminé en {time.time() - start_time:.2f} secondes ===")
        
    except Exception as e:
        logging.error(f"ERREUR GLOBALE : {str(e)}")
        print(f"\n=== ERREUR GLOBALE : {str(e)} ===")
    finally:
        if connection:
            cursor.close()
            connection.close()
            logging.info("Connexion MySQL fermée")
            print("\n=== Connexion MySQL fermée ===")

if __name__ == "__main__":
    try:
        import camelot
        import pandas
        import fitz
        import openpyxl
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import Image
        main()
    except ImportError as e:
        logging.error(f"Packages manquants : {str(e)}")
        print(f"\nPackages manquants : {str(e)}")
        print("Installez-les avec : pip install camelot-py pandas pymupdf openpyxl pytesseract pdf2image pillow")
        input("Appuyez sur Entrée pour quitter...")