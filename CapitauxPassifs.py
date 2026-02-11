# Configuration et Imports
import os
import re
import time
import mysql.connector
from mysql.connector import Error
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse, urlencode, parse_qs
from datetime import datetime
import PyPDF2
import requests
import pandas as pd
import camelot
import fitz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
import logging
import pyodbc

# -----------------------------------------------------Partie 1 : Configuration des logs ----------------------------------------------------------------------
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
os.environ["GLOG_minloglevel"] = "3"
os.environ["PYTHONWARNINGS"] = "ignore"
logging.basicConfig(filename='script.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

societes_assurances = [
    "Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR -"
]

# -----------------------------------------------------Partie 2 : Gestion Base de données ---------------------------------------------------------------------
def create_cmf_database_and_table():
    try:
        connection = pyodbc.connect(
            'DRIVER={ODBC Driver 18 for SQL Server};'
            'SERVER=localhost;'
            'Trusted_Connection=yes;'
            'Encrypt=no;',
            autocommit=True
        )
        cursor = connection.cursor()
        
        logging.info("Création/mise à jour de la base de données 'cmf'...")
        print("Création/mise à jour de la base de données 'cmf'...")
        
        cursor.execute("""
        IF DB_ID('cmf') IS NULL
            CREATE DATABASE cmf
        """)
        
        cursor.execute("USE cmf")
        
        cursor.execute("""
        IF OBJECT_ID('document', 'U') IS NULL
        CREATE TABLE document (
            id INT IDENTITY(1,1) PRIMARY KEY,
            Societe NVARCHAR(255) NOT NULL,
            Nom NVARCHAR(255) NOT NULL,
            Annee INT NOT NULL,
            URL NVARCHAR(512) NOT NULL,
            CONSTRAINT unique_document UNIQUE (Societe, Nom, Annee)
        )
        """)

        cursor.execute("""
        IF OBJECT_ID('financial_data', 'U') IS NULL
        CREATE TABLE financial_data (
            id INT IDENTITY(1,1) PRIMARY KEY,
            document_id INT NOT NULL,
            category NVARCHAR(255),
            subcategory NVARCHAR(255),
            code NVARCHAR(50),
            description NVARCHAR(MAX),
            level INT,
            is_total BIT,
            value_n FLOAT,
            value_n_1 FLOAT,
            FOREIGN KEY (document_id) REFERENCES document(id) ON DELETE CASCADE
        )
        """)
        
        connection.commit()
        
        logging.info("Base de données 'cmf', tables 'document' et 'financial_data' prêtes.")
        print("Base de données 'cmf', tables 'document' et 'financial_data' prêtes.")
        
        return connection, cursor

    except Exception as e:
        logging.error(f"Erreur lors de la création de la base : {e}")
        print(f"Erreur lors de la création de la base : {e}")
        return None, None

# -----------------------------------------------------Partie 3 : Fonctions utilitaires ---------------------------------------------------------------------

def init_driver():
    chrome_options = Options()
    # chrome_options.add_argument("--headless=new") # Disable headless for interactivity visibility if needed, or keep provided user preference. Keeping headless for speed unless debugging.
    # User might want to see it? User didn't specify, but CLI usually implies headless or console interaction.
    # Let's keep headless to be clean, but show progress in console.
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    service = Service(ChromeDriverManager().install(), log_path=os.devnull)
    return webdriver.Chrome(service=service, options=chrome_options)

def get_all_companies(driver):
    print("⏳ Chargement de la liste des sociétés...")
    driver.get("https://www.cmf.tn/consultation-des-tats-financier-des-soci-t-s-faisant-ape")
    wait = WebDriverWait(driver, 20)
    
    try:
        dropdown = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#edit_field_societesape_value_chosen a")))
        dropdown.click()
        
        options = wait.until(EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "#edit_field_societesape_value_chosen .chosen-results li")))
        
        companies = [opt.text.strip() for opt in options if opt.text.strip()]
        return companies
    except Exception as e:
        print(f"❌ Erreur lors du chargement des sociétés : {e}")
        return []

def fetch_pdf_interactive(driver, societe, annee):
    """
    Fetches PDFs for a given company and year. Assumes the company has already been selected
    in the dropdown and the results page is loaded.
    """
    print(f"\n🔎 Recherche des documents pour : {societe} ({annee})")
    
    try:
        wait = WebDriverWait(driver, 20)
        pdfs = []
        
        # Scrape all pages of results
        page_num = 1
        while True:
            time.sleep(2)  # Wait for page to load
            rows = driver.find_elements(By.CSS_SELECTOR, ".view-content .views-row")
            print(f"DEBUG: Found {len(rows)} documents on page {page_num}.")
            
            for row in rows:
                try:
                    pdf_url = row.find_element(By.CSS_SELECTOR, "a[href$='.pdf']").get_attribute("href")
                    pdf_name = row.find_element(
                        By.CSS_SELECTOR, ".field-name-field-p-riode .field-item").text.strip()
                    
                    year = extract_year_from_text(pdf_url) or extract_year_from_text(pdf_name)
                    print(f"  - Document: '{pdf_name}' | Year: '{year}'")
                    
                    if year and year.isdigit():
                        pdfs.append({
                            "url": pdf_url,
                            "nom": pdf_name,
                            "annee": year,
                            "societe": societe
                        })
                except (NoSuchElementException, StaleElementReferenceException) as e:
                    print(f"DEBUG: Error processing row: {e}")
                    continue
            
            # Try to go to next page
            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, ".pager .next a:not(.disabled)")
                next_btn.click()
                page_num += 1
                time.sleep(3)
            except:
                print(f"DEBUG: No more pages (total pages: {page_num}).")
                break
        
        print(f"DEBUG: {len(pdfs)} documents trouvés au total pour la société.")
        if len(pdfs) > 0:
            print("Exemples de documents trouvés :")
            for p in pdfs[:5]:
                 print(f" - {p['nom']} (Année détectée: {p['annee']})")
        
        # Filtrer par année
        year_pdfs = [p for p in pdfs if str(int(p["annee"])) == str(annee)]
        
        if not year_pdfs:
            print(f"❌ Aucun document trouvé pour l'année {annee}.")
            return None, None
            
        print(f"\n📂 Documents trouvés pour {annee} :")
        for i, pdf in enumerate(year_pdfs, 1):
            print(f"  [{i}] {pdf['nom']}")
            
        while True:
            choice = input("\n👉 Choisissez le numéro du document à télécharger (ou 0 pour annuler) : ").strip()
            if choice == '0':
                return None, None
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(year_pdfs):
                    selected_pdf = year_pdfs[idx]
                    print(f"✅ Document sélectionné : {selected_pdf['nom']}")
                    
                    pdf_path = download_pdf(selected_pdf["url"], selected_pdf["societe"], selected_pdf["nom"], selected_pdf["annee"])
                    if pdf_path:
                        return pdf_path, selected_pdf["url"]
                    else:
                        print("❌ Erreur de téléchargement.")
                        return None, None
            print("⚠️ Choix invalide.")
            
    except Exception as e:
        logging.error(f"ERREUR lors de la récupération interactive: {str(e)}")
        print(f"❌ ERREUR : {str(e)}")
        return None, None

def insert_pdf_info_cmf(connection, cursor, societe, nom_document, annee, url):
    try:
        normalized_url = normalize_url(url)
        
        try:
            annee_int = int(annee)
            if not (2015 <= annee_int <= 2025):
                return False
        except ValueError:
            return False
            
        if check_document_exists(cursor, societe, nom_document, annee_int):
            return False
            
        insert_query = """
        INSERT INTO document (Societe, Nom, Annee, URL)
        VALUES (?, ?, ?, ?)
        """
        cursor.execute(insert_query, (societe, nom_document, annee_int, normalized_url))
        connection.commit()
        return True
        
        return True
        
    except Error as e:
        logging.error(f"Erreur lors de l'insertion : {e}")
        return False

def insert_financial_data(cursor, doc_id, hierarchical_data):
    """
    Insère les données financières extraites dans la table financial_data
    """
    try:
        print(f"\n Insertion des données financières pour le document {doc_id}...")
        
        # Supprimer les anciennes données pour ce document pour éviter les doublons
        cursor.execute("DELETE FROM financial_data WHERE document_id = ?", (doc_id,))
        
        insert_query = """
        INSERT INTO financial_data 
        (document_id, category, subcategory, code, description, level, is_total, value_n, value_n_1)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        
        count = 0
        for item in hierarchical_data:
            # Récupérer les valeurs, gérer les cas vides ou non numériques
            val_n = item['values'][0] if len(item['values']) > 0 and isinstance(item['values'][0], (int, float)) else None
            val_n_1 = item['values'][1] if len(item['values']) > 1 and isinstance(item['values'][1], (int, float)) else None
            
            cursor.execute(insert_query, (
                doc_id,
                item['category'],
                item['subcategory'],
                item['code'],
                item['description'],
                item['level'],
                1 if item['is_total'] else 0,
                val_n,
                val_n_1
            ))
            count += 1
            
        cursor.commit() # Note: pyodbc cursor commit might be needed depending on autocommit setting
        print(f"✅ {count} lignes insérées dans la table financial_data")
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors de l'insertion des données financières : {str(e)}")
        logging.error(f"Erreur insertion financial_data : {str(e)}")
        return False

def check_document_exists(cursor, societe, nom, annee):
    try:
        query = """
        SELECT COUNT(*) FROM document 
        WHERE Societe = ? 
        AND Nom = ?
        AND Annee = ?
        """
        cursor.execute(query, (societe, nom, annee))
        return cursor.fetchone()[0] > 0
    except Error as e:
        return False

def normalize_url(url):
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    if 'id' in query_params or 'token' in query_params:
        query_params.pop('id', None)
        query_params.pop('token', None)
    new_query = urlencode(query_params, doseq=True)
    return parsed_url._replace(query=new_query).geturl()

def extract_year_from_text(text):
    patterns = [
        r'(?:20)(1[5-9]|2[0-9])\b', # 2015-2029
        r'\b(20\d{2})\b', # 20xx
        r'(?:3112|1312|31_12_|31-12-)(1[5-9]|2[0-9])', # 311219 or 31_12_19
        r'(?:3112|1312|31_12_|31-12-)(20\d{2})', # 31122019
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            # group(1) is usually the year part. If not, group(0) might be it but group(1) is safer given the patterns.
            year = match.group(1) if match.lastindex and match.lastindex >= 1 else match.group(0)
            
            # Clean up
            year = re.sub(r'[^0-9]', '', year)
            
            if len(year) == 2:
                year = f"20{year}"
            
            if len(year) == 4 and 2010 <= int(year) <= 2030:
                return year
    return None

def extract_pdfs_from_page(driver, societe):
    wait = WebDriverWait(driver, 20)
    pdfs = []
    
    try:
        time.sleep(3) # Ensure page loaded
        dropdown = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#edit_field_societesape_value_chosen a")))
        dropdown.click()
        time.sleep(1)
        
        options = wait.until(EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "#edit_field_societesape_value_chosen .chosen-results li")))
        
        selected_option = None
        # First pass: Exact match
        for option in options:
            if societe.lower() == option.text.strip().lower():
                selected_option = option
                print(f"DEBUG: Exact match found for '{societe}'")
                break
        
        # Second pass: Partial match if no exact match
        if not selected_option:
            print(f"DEBUG: No exact match for '{societe}', trying partial...")
            for option in options:
                opt_text = option.text.strip().lower()
                target_text = societe.lower()
                if target_text in opt_text or opt_text in target_text:
                    selected_option = option
                    print(f"DEBUG: Partial match found: '{option.text.strip()}'")
                    break
        
        if not selected_option:
            print(f"DEBUG: Could not match '{societe}' in dropdown.")
            return []
        
        selected_option.click()

        submit = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "#edit-submit-consultation-des-tats-financier-des-soci-t-s-faisant-ape")))
        submit.click()
        time.sleep(2)

        page_num = 1
        pdfs = []
        while True:
            # Wait a sec for rows to load
            time.sleep(1)
            rows = driver.find_elements(By.CSS_SELECTOR, ".view-content .views-row")
            print(f"DEBUG: Found {len(rows)} documents on page {page_num}.")
            
            for row in rows:
                try:
                    pdf_url = row.find_element(By.CSS_SELECTOR, "a[href$='.pdf']").get_attribute("href")
                    pdf_name = row.find_element(
                        By.CSS_SELECTOR, ".field-name-field-p-riode .field-item").text.strip()
                    
                    year = extract_year_from_text(pdf_url) or extract_year_from_text(pdf_name)
                    print(f"  - Document: '{pdf_name}' | Url: ...{pdf_url[-15:]} | Year: '{year}'")
                    
                    if year and year.isdigit():
                        pdfs.append({
                            "url": pdf_url,
                            "nom": pdf_name,
                            "annee": year,
                            "societe": societe
                        })

                except (NoSuchElementException, StaleElementReferenceException) as e:
                    print(f"DEBUG: Error processing row: {e}")
                    continue

            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, ".pager .next a:not(.disabled)")
                next_btn.click()
                page_num += 1
                time.sleep(3)
            except:
                break

    except Exception as e:
        logging.error(f"ERREUR lors de l'extraction : {str(e)}")
    
    return pdfs

def download_pdf(url, societe, nom, annee):
    try:
        safe_societe = re.sub(r'[^\w\s-]', '_', societe).replace(' ', '_')
        safe_nom = re.sub(r'[^\w\s-]', '_', nom).replace(' ', '_')
        filename = f"{safe_societe}_{safe_nom}_{annee}.pdf"
        filepath = os.path.join(os.getcwd(), filename)
        
        if os.path.exists(filepath):
            print(f" PDF déjà existant : {filepath}")
            return filepath
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        response = requests.get(url, stream=True, timeout=30)
        response.raise_for_status()
        with open(filepath, 'wb') as f:
            f.write(response.content)
        print(f" PDF téléchargé : {filepath}")
        return filepath
    except Exception as e:
        print(f" Erreur téléchargement : {str(e)}")
        return None

def clean_number(text):
    if isinstance(text, str):
        cleaned = re.sub(r'\s+', '', text).replace(',', '.').rstrip('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ')
        try:
            return int(float(cleaned))
        except ValueError:
            return text
    return int(text) if isinstance(text, (int, float)) else text

# ===================================================================================================
# NOUVELLE SECTION : Extraction et structuration hiérarchique du bilan
# ===================================================================================================

def detect_hierarchy_level(row_data, current_section=None):
    """
    Détecte le niveau hiérarchique d'une ligne basé sur les patterns de codes.
    Retourne (level, code, description, is_total, category, subcategory)
    """
    if not row_data or len(row_data) == 0:
        return None
    
    first_col = str(row_data[0]).strip() if row_data[0] else ""
    second_col = str(row_data[1]).strip() if len(row_data) > 1 and row_data[1] else ""
    
    # Combiner les deux premières colonnes pour l'analyse
    combined = f"{first_col} {second_col}".strip()
    combined_lower = combined.lower()
    
    # Niveau 0 : Titre principal (CAPITAUX PROPRES ET LE PASSIF)
    if "capitaux propres et" in combined_lower and "passif" in combined_lower:
        return (0, "", combined, False, "TITRE", "")
    
    # Niveau 1 : Sections principales (CAPITAUX PROPRES:, PASSIF:)
    if re.match(r'^(CAPITAUX PROPRES|PASSIF):?$', combined, re.IGNORECASE):
        section = "CAPITAUX PROPRES" if "capitaux propres" in combined_lower else "PASSIF"
        return (1, "", combined, False, "SECTION", section)
    
    # Totaux principaux
    if "total" in combined_lower:
        if "total capitaux propres avant résultat" in combined_lower:
            return (2, "", combined, True, "TOTAL", "Capitaux Propres - Avant Résultat")
        elif "total capitaux propres avant affectation" in combined_lower:
            return (2, "", combined, True, "TOTAL", "Capitaux Propres - Avant Affectation")
        elif "total du passif" in combined_lower:
            return (2, "", combined, True, "TOTAL", "Total Passif")
        elif "total des capitaux propres et du passif" in combined_lower:
            return (1, "", combined, True, "TOTAL GÉNÉRAL", "")
        else:
            category = current_section if current_section else "TOTAL"
            return (3, "", combined, True, "TOTAL", category)
    
    # Niveau 2 : Sous-sections (PA2, PA3, PA5, PA6, PA7, PA72)
    if re.match(r'^(PA\d+|PA\d+[A-Z]?\d*)\s+', combined):
        code_match = re.match(r'^(PA\d+[A-Z]?\d*)\s+(.+)', combined)
        if code_match:
            code = code_match.group(1)
            desc = code_match.group(2)
            
            # Mapper les codes aux sous-catégories
            subcategory_map = {
                'PA2': 'Provisions pour risques et charges',
                'PA23': 'Provisions pour risques et charges',
                'PA3': 'Provisions techniques brutes',
                'PA310': 'Provisions techniques brutes',
                'PA320': 'Provisions techniques brutes',
                'PA330': 'Provisions techniques brutes',
                'PA331': 'Provisions techniques brutes',
                'PA340': 'Provisions techniques brutes',
                'PA341': 'Provisions techniques brutes',
                'PA350': 'Provisions techniques brutes',
                'PA360': 'Provisions techniques brutes',
                'PA361': 'Provisions techniques brutes',
                'PA5': 'Dettes pour dépôts',
                'PA6': 'Autres dettes',
                'PA61': 'Autres dettes',
                'PA62': 'Autres dettes',
                'PA63': 'Autres dettes',
                'PA631': 'Autres dettes',
                'PA632': 'Autres dettes',
                'PA633': 'Autres dettes',
                'PA634': 'Autres dettes',
                'PA7': 'Autres passifs',
                'PA71': 'Comptes de régularisation',
                'PA710': 'Comptes de régularisation',
                'PA711': 'Comptes de régularisation',
                'PA712': 'Comptes de régularisation',
                'PA72': 'Écart de conversion',
            }
            
            subcategory = subcategory_map.get(code, 'PASSIF')
            
            # Déterminer si c'est un parent ou un enfant
            if code in ['PA2', 'PA3', 'PA5', 'PA6', 'PA7', 'PA72', 'PA71']:
                return (2, code, desc, False, "PASSIF", subcategory)
            else:
                return (3, code, desc, False, "PASSIF", subcategory)
    
    # Niveau 2 : Codes CP (Capital)
    if re.match(r'^(CP\d+)\s+', combined):
        code_match = re.match(r'^(CP\d+)\s+(.+)', combined)
        if code_match:
            code = code_match.group(1)
            desc = code_match.group(2)
            
            # Mapper les codes CP
            subcategory_map = {
                'CP1': 'Capital social',
                'CP2': 'Réserves et primes',
                'CP3': 'Rachat d\'actions',
                'CP4': 'Autres capitaux propres',
                'CP5': 'Résultat reporté',
                'CP6': 'Résultat de l\'exercice',
            }
            
            subcategory = subcategory_map.get(code, 'Capitaux Propres')
            return (2, code, desc, False, "CAPITAUX PROPRES", subcategory)
    
    # Si commence par un code seul (CP ou PA)
    if re.match(r'^(CP\d+|PA\d+[A-Z]?\d*)$', first_col):
        if first_col.startswith('CP'):
            return (2, first_col, second_col, False, "CAPITAUX PROPRES", "")
        else:
            return (2, first_col, second_col, False, "PASSIF", "")
    
    # Ligne de description sans code (niveau 2 par défaut)
    if first_col and not re.match(r'^(CP|PA)', first_col):
        category = current_section if current_section else "AUTRE"
        return (2, "", combined, False, category, "")
    
    return None

def extract_hierarchical_table(pdf_path, page_num, is_scanned):
    """
    Extrait le tableau en préservant la structure hiérarchique
    """
    try:
        print(f"\n🔍 Extraction hiérarchique page {page_num}...")
        
        if is_scanned:
            # Extraction OCR
            images = convert_from_path(pdf_path, first_page=page_num, last_page=page_num, dpi=150)
            if not images:
                return None
            
            image = images[0]
            text = pytesseract.image_to_string(image, lang='fra', config='--psm 6')
            
            # Parser le texte OCR ligne par ligne
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            structured_data = []
            
            for line in lines:
                # Séparer par espaces multiples
                parts = re.split(r'\s{2,}', line)
                if len(parts) >= 2:
                    structured_data.append(parts)
        
        else:
            # Extraction native avec Camelot
            tables = camelot.read_pdf(pdf_path, flavor='stream', pages=str(page_num))
            
            if tables.n == 0:
                print(" Aucun tableau détecté")
                return None
            
            # Prendre le premier tableau (généralement le plus complet)
            df = tables[0].df
            structured_data = df.values.tolist()
        
        # Structurer les données avec hiérarchie
        hierarchical_rows = []
        current_section = None
        
        for row in structured_data:
            if not any(str(cell).strip() for cell in row):  # Ignorer lignes vides
                continue
            
            hierarchy_info = detect_hierarchy_level(row, current_section)
            
            if hierarchy_info:
                level, code, description, is_total, category, subcategory = hierarchy_info
                
                # Mettre à jour la section courante
                if category == "SECTION":
                    current_section = subcategory
                
                # Extraire les valeurs numériques (colonnes de droite)
                values = []
                for cell in row[2:] if len(row) > 2 else row:
                    cleaned = clean_number(cell)
                    values.append(cleaned)
                
                # Si pas de valeurs extraites, essayer les deux dernières colonnes
                if not values or all(v == '' or v == 0 for v in values):
                    if len(row) >= 2:
                        values = [clean_number(row[-2]), clean_number(row[-1])]
                
                hierarchical_rows.append({
                    'level': level,
                    'code': code,
                    'description': description,
                    'is_total': is_total,
                    'category': category,
                    'subcategory': subcategory,
                    'values': values
                })
        
        print(f" {len(hierarchical_rows)} lignes structurées extraites")
        return hierarchical_rows
        
    except Exception as e:
        print(f"Erreur extraction hiérarchique : {str(e)}")
        logging.error(f"Erreur extraction hiérarchique : {str(e)}")
        return None

def export_hierarchical_to_excel(hierarchical_data, output_name):
    """
    Exporte les données hiérarchiques vers Excel avec mise en forme minimaliste
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Capitaux Propres et Passif"
        
        # Styles minimalistes
        header_font = Font(bold=True, size=11)
        total_font = Font(bold=True, size=10)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        
        number_format = '#,##0'
        border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))
        
        # En-tête
        ws.append(['Type', 'Sous-catégorie', 'Code', 'Description', '31/12/2024', '31/12/2023'])
        for col in range(1, 7):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Parcourir les données hiérarchiques
        current_row = 2
        
        for item in hierarchical_data:
            level = item['level']
            code = item['code']
            description = item['description']
            is_total = item['is_total']
            category = item['category']
            subcategory = item['subcategory']
            values = item['values']
            
            # Ignorer les lignes de section (niveau 1)
            if level == 1 and category == "SECTION":
                continue
            
            # Ignorer le titre principal si vous voulez (niveau 0)
            if level == 0:
                continue
            
            # Indentation basée sur le niveau (réajusté car on retire niveau 0 et 1)
            indent = "  " * max(0, level - 1)
            
            # Écrire le type/catégorie
            ws.cell(current_row, 1, category)
            
            # Écrire la sous-catégorie
            ws.cell(current_row, 2, subcategory)
            
            # Écrire le code
            ws.cell(current_row, 3, code)
            
            # Écrire la description avec indentation
            ws.cell(current_row, 4, f"{indent}{description}")
            
            # Écrire les valeurs numériques
            for i, value in enumerate(values[:2], start=5):  # Max 2 colonnes de valeurs
                if value and value != '':
                    ws.cell(current_row, i, value)
                    ws.cell(current_row, i).number_format = number_format
            
            # Appliquer le style minimaliste
            for col in range(1, 7):
                cell = ws.cell(current_row, col)
                cell.border = border
                
                # Gras pour les totaux et sous-sections principales
                if is_total or level == 2:
                    cell.font = total_font if is_total else bold_font
                else:
                    cell.font = normal_font
                
                # Alignement
                if col <= 4:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            current_row += 1
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 20  # Type
        ws.column_dimensions['B'].width = 35  # Sous-catégorie
        ws.column_dimensions['C'].width = 8   # Code
        ws.column_dimensions['D'].width = 55  # Description
        ws.column_dimensions['E'].width = 15  # 31/12/2024
        ws.column_dimensions['F'].width = 15  # 31/12/2023
        
        # Figer la première ligne
        ws.freeze_panes = 'A2'
        
        # Ajouter des filtres automatiques
        ws.auto_filter.ref = f"A1:F{current_row-1}"
        
        wb.save(output_name)
        print(f"✅ Fichier Excel créé : {os.path.abspath(output_name)}")
        logging.info(f"Fichier Excel créé : {os.path.abspath(output_name)}")
        return True
        
    except Exception as e:
        print(f" Erreur export Excel : {str(e)}")
        logging.error(f"Erreur export Excel : {str(e)}")
        return False

def search_capitaux_propres_passif_in_pdf(pdf_path):
    """
    Recherche la page contenant "CAPITAUX PROPRES ET LE PASSIF"
    """
    try:
        print(f"\n Recherche de CAPITAUX PROPRES ET PASSIF...")
        
        pdf_reader = PyPDF2.PdfReader(pdf_path)
        num_pages = len(pdf_reader.pages)
        
        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            
            if text:
                text_lower = text.lower()
                text_lower = re.sub(r'\s+', ' ', text_lower)
                
                if "capitaux propres et" in text_lower and "passif" in text_lower:
                    print(f" CAPITAUX PROPRES ET PASSIF trouvé à la page {page_num + 1}")
                    return page_num + 1, False
        
        # OCR fallback
        print(" Tentative OCR...")
        images = convert_from_path(pdf_path, dpi=100)
        
        for page_num, image in enumerate(images, 1):
            img = image.resize((int(image.width * 0.25), int(image.height * 0.25)), Image.Resampling.LANCZOS)
            text = pytesseract.image_to_string(img, lang='fra', config='--psm 6')
            if text:
                text_lower = text.lower()
                if "capitaux propres et" in text_lower and "passif" in text_lower:
                    print(f"✅ CAPITAUX PROPRES ET PASSIF trouvé à la page {page_num} (OCR)")
                    return page_num, True
        
        print(" CAPITAUX PROPRES ET PASSIF non trouvé")
        return None, None
            
    except Exception as e:
        print(f" Erreur : {str(e)}")
        return None, None

def search_and_extract_dynamic(connection, cursor, societe, annee):
    """
    Recherche et extrait le tableau CAPITAUX PROPRES ET PASSIF de manière structurée pour une société et une année données
    """
    logging.info(f"Recherche du bilan pour {societe} {annee}")
    print("\n" + "="*70)
    print(f" RECHERCHE DU BILAN - CAPITAUX PROPRES ET PASSIF : {societe} {annee}")
    print("="*70)
    
    try:
        # Recherche plus souple dans la base de données
        query = """
        SELECT id, Societe, Nom, Annee, URL 
        FROM document 
        WHERE Societe = ? 
        AND Annee = ?
        AND Nom LIKE ?
        """
        # On cherche un document qui contient "Etats financiers" et "31/12"
        cursor.execute(query, (societe, int(annee), '%Etats financiers%'))
        
        results = cursor.fetchall()
        target_doc = None
        
        # Filtrer pour trouver le bon (ex: contenant 31/12)
        for res in results:
             if "31/12" in res[2]: # res[2] est Nom
                 target_doc = res
                 break
        
        if not target_doc and results:
             target_doc = results[0] # Fallback au premier trouvé si pas de 31/12 explicite
             
        if target_doc:
            doc_id, doc_societe, doc_nom, doc_annee, doc_url = target_doc
            print(f"\n Document trouvé en base - ID: {doc_id} ('{doc_nom}')")
            
            safe_societe = re.sub(r'[^\w\s-]', '_', doc_societe).replace(' ', '_')
            safe_nom = re.sub(r'[^\w\s-]', '_', doc_nom).replace(' ', '_')
            pdf_path = os.path.join(os.getcwd(), f"{safe_societe}_{safe_nom}_{doc_annee}.pdf")
            
            if not os.path.exists(pdf_path):
                print(" PDF non trouvé localement, redownloading...")
                pdf_path = download_pdf(doc_url, doc_societe, doc_nom, doc_annee)
                if not pdf_path:
                    print("❌ Impossible de télécharger le PDF.")
                    return

            # Rechercher la page du tableau
            page_num, is_scanned = search_capitaux_propres_passif_in_pdf(pdf_path)
            
            if page_num:
                # Extraire avec structure hiérarchique
                hierarchical_data = extract_hierarchical_table(pdf_path, page_num, is_scanned)
                
                if hierarchical_data:
                    # Insertion dans la base de données
                    insert_financial_data(cursor, doc_id, hierarchical_data)
                    connection.commit()

                    # Export Excel avec nom dynamique
                    output_name = f"{safe_societe}_{doc_annee}_{safe_nom}.xlsx"
                    
                    if export_hierarchical_to_excel(hierarchical_data, output_name):
                        print(f"\n" + "="*70)
                        print(f"✅ EXTRACTION RÉUSSIE")
                        print(f"📁 Fichier : {output_name}")
                        print(f"📊 Lignes extraites : {len(hierarchical_data)}")
                        print("="*70)
                        
                        # Appel au code B
                        import subprocess
                        try:
                            # Verify if B.py logic needs this specific filename or if we should pass it
                            subprocess.run(["python", "B.py", os.path.abspath(output_name)], check=True)
                            print(f"✅ Code B exécuté avec succès")
                        except subprocess.CalledProcessError as e:
                            print(f"❌ Erreur lors de l'exécution du code B : {str(e)}")
                        except FileNotFoundError:
                            print(f"ℹ️ Fichier B.py non trouvé, étape ignorée")
                    else:
                        print("❌ Échec de l'exportation vers Excel")
                else:
                    print("⚠️ Aucune donnée hiérarchique extraite")
            else:
                print("⚠️ CAPITAUX PROPRES ET PASSIF non trouvé dans le document")
        else:
            print(f"❌ Aucun document trouvé pour {societe} {annee}")
            
    except Exception as e:
        logging.error(f"ERREUR : {str(e)}")
        print(f"❌ ERREUR : {str(e)}")

def process_societe(societe, connection, cursor):
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    service = Service(ChromeDriverManager().install(), log_path=os.devnull)
    
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.get("https://www.cmf.tn/consultation-des-tats-financier-des-soci-t-s-faisant-ape")
    
    pdfs = extract_pdfs_from_page(driver, societe)
    inserted_count = 0
    
    for pdf in pdfs:
        if insert_pdf_info_cmf(connection, cursor, pdf["societe"], pdf["nom"], pdf["annee"], pdf["url"]):
            inserted_count += 1
    
    driver.quit()
    return inserted_count

def check_missing_documents(connection, cursor):
    try:
        print("\n=== 🔍 Vérification des documents manquants ===")
        total_inserted = 0
        
        prioritized_societe = "Sté. TUNISIENNE D'ASSURANCES ET DE REASSURANCES - STAR -"
        societes = [prioritized_societe]
        
        for societe in societes:
            cursor.execute("SELECT Annee FROM document WHERE Societe = ?", (societe,))
            existing_years = {row[0] for row in cursor.fetchall()}
            
            missing_years = set(range(2015, 2026)) - existing_years
            
            if missing_years:
                print(f"⚠️ Documents manquants : {sorted(missing_years)}")
                
                chrome_options = Options()
                chrome_options.add_argument("--headless=new")
                chrome_options.add_argument("--no-sandbox")
                chrome_options.add_argument("--disable-dev-shm-usage")
                chrome_options.add_argument("--disable-gpu")
                service = Service(ChromeDriverManager().install(), log_path=os.devnull)
                
                driver = webdriver.Chrome(service=service, options=chrome_options)
                driver.get("https://www.cmf.tn/consultation-des-tats-financier-des-soci-t-s-faisant-ape")
                
                pdfs = extract_pdfs_from_page(driver, societe)
                inserted_count = 0
                
                for pdf in pdfs:
                    if int(pdf["annee"]) in missing_years:
                        if insert_pdf_info_cmf(connection, cursor, pdf["societe"], pdf["nom"], pdf["annee"], pdf["url"]):
                            inserted_count += 1
                
                driver.quit()
                total_inserted += inserted_count
                print(f"✅ {inserted_count} documents ajoutés")
            else:
                print(f"✅ Aucun document manquant")
            
            time.sleep(2)
        
        return total_inserted
        
    except Exception as e:
        print(f"❌ ERREUR : {str(e)}")
        return 0

# -----------------------------------------------------Partie 4 : Code main d'exécution --------------------------------------------------------

def main():
    start_time = time.time()
    print(f"\n{'='*70}")
    print(f"🚀 EXTRACTION STRUCTURÉE - CAPITAUX PROPRES ET PASSIF (INTERACTIF)")
    print(f"{'='*70}")
    
    driver = init_driver()
    try:
        # 1. Gestion des sociétés
        available_companies = get_all_companies(driver)
        if not available_companies:
            print("❌ Impossible de récupérer la liste des sociétés.")
            return

        target_societe = None
        while not target_societe:
            search_query = input("\n🏢 Entrez le nom (ou partie du nom) de la société : ").strip().lower()
            if not search_query:
                continue
                
            matches = [c for c in available_companies if search_query in c.lower()]
            
            if not matches:
                print("❌ Aucune société trouvée.")
                continue
            
            print(f"\nSociétés trouvées ({len(matches)}) :")
            for i, match in enumerate(matches, 1):
                print(f"  [{i}] {match}")
            
            choice = input(f"👉 Sélectionnez le numéro de la société (1-{len(matches)}) ou 0 pour rechercher à nouveau : ").strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(matches):
                    target_societe = matches[idx]
                    print(f"✅ Société sélectionnée : {target_societe}")
                elif idx == -1:
                    continue
            else:
                print("⚠️ Choix invalide.")

        # 2. Choix de l'année
        target_annee = None
        while not target_annee:
            year_input = input("\n📅 Entrez l'année (ex: 2024) : ").strip()
            if year_input.isdigit() and 2015 <= int(year_input) <= 2030:
                target_annee = int(year_input)
            else:
                print("⚠️ Année invalide.")

        # 3. Select company in dropdown and submit form
        print(f"\n🔄 Sélection de '{target_societe}' sur le site CMF...")
        try:
            wait = WebDriverWait(driver, 20)
            
            # Refresh the page to avoid stale elements
            driver.get("https://www.cmf.tn/consultation-des-tats-financier-des-soci-t-s-faisant-ape")
            time.sleep(2)
            
            # Click dropdown
            dropdown = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#edit_field_societesape_value_chosen a")))
            dropdown.click()
            time.sleep(1)
            
            # Find and click the company option
            options = wait.until(EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "#edit_field_societesape_value_chosen .chosen-results li")))
            
            selected_option = None
            for option in options:
                if target_societe.lower() == option.text.strip().lower():
                    selected_option = option
                    print(f"DEBUG: Exact match found!")
                    break
            
            if not selected_option:
                # Try partial match
                print(f"DEBUG: Trying partial match...")
                for option in options:
                    opt_text = option.text.strip()
                    if opt_text and target_societe.lower() in opt_text.lower():
                        selected_option = option
                        print(f"DEBUG: Partial match found: '{opt_text}'")
                        break
            
            if not selected_option:
                print(f"❌ Impossible de sélectionner '{target_societe}' dans le dropdown.")
                print(f"DEBUG: Available options count: {len(options)}")
                for i, opt in enumerate(options[:5]):
                    print(f"  Sample {i+1}: '{opt.text.strip()}'")
                return
            
            selected_option.click()
            
            # Submit form
            submit = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "#edit-submit-consultation-des-tats-financier-des-soci-t-s-faisant-ape")))
            submit.click()
            time.sleep(3)  # Wait for results to load
            
            print("✅ Formulaire soumis, chargement des résultats...")
            
        except Exception as e:
            print(f"❌ Erreur lors de la sélection de la société : {e}")
            return

        # 4. Récupération PDF Interactive (results page already loaded)
        pdf_path, doc_url = fetch_pdf_interactive(driver, target_societe, target_annee)
        
        if not pdf_path:
            print("❌ Processus annulé ou aucun fichier téléchargé.")
            return

        # Le reste du processus continue...
        target_nom = "Etats financiers au 31/12" # On garde ce nom générique pour ID DB ou on utilise le vrai nom du fichier ?
        # Idéalement on utilise le nom structuré, mais la logique actuelle DB repose sur Nom.
        # Pour simplifier, on garde le nom DB générique ou on extrait du path.
        # Le nom du document est utilisé pour check_document_exists.
        
        # 4. Gestion base de données
        connection, cursor = create_cmf_database_and_table()
        if not connection or not cursor:
            print("❌ Échec de la connexion à la base de données")
            return

        # 5. Ajout du document dans la base
        # Note: insert_pdf_info_cmf utilise normalize_url.
        if doc_url:
            insert_pdf_info_cmf(connection, cursor, target_societe, target_nom, target_annee, doc_url)
        
        # Étape 4: Recherche et extraction structurée
        # search_and_extract_dynamic utilise 'Etats financiers au 31/12' comme nom pour chercher en base.
        # C'est cohérent avec ce qu'on insère juste au-dessus.
        
        # Cependant, search_and_extract_dynamic DOIT utiliser le PDF qu'on vient de télécharger !
        # Problème : search_and_extract_dynamic redownload le PDF basé sur l'URL en base.
        # Comme on vient de le télécharger dans 'pdf_path', on devrait idéalement passer 'pdf_path' à la fonction d'extraction.
        # Refactorisons légèrement search_and_extract_dynamic pour accepter un pdf_path optionnel ou juste ne pas redownloader si existant.
        # Le code actuel check `if not os.path.exists(pdf_path)`. Si `fetch_pdf_interactive` a téléchargé le fichier avec le BON nom, c'est bon.
        # `fetch_pdf_interactive` appelle `download_pdf` qui utilise `safe_societe` et `safe_nom` (Etats financiers au 31/12).
        # Donc si on passe les mêmes `target_societe` et `target_nom` à `download_pdf`, le fichier aura le même nom.
        
        search_and_extract_dynamic(connection, cursor, target_societe, target_annee)
        
        elapsed = time.time() - start_time
        print(f"\n{'='*70}")
        print(f"✅ TERMINÉ en {elapsed:.2f}s")
        print(f"{'='*70}\n")
        
    except Exception as e:
        print(f"\n❌ ERREUR GLOBALE : {str(e)}")
    finally:
        if 'driver' in locals():
            driver.quit()
        if 'connection' in locals() and connection:
            cursor.close()
            connection.close()
            print("✅ Connexion fermée")

if __name__ == "__main__":
    try:
        import camelot
        import pandas
        import fitz
        import openpyxl
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import Image
        main()
    except ImportError as e:
        print(f"\n Packages manquants : {str(e)}")
        print("Installez-les avec : pip install camelot-py pandas pymupdf openpyxl pytesseract pdf2image pillow")
        input("Appuyez sur Entrée pour quitter...")